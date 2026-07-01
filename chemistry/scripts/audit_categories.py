#!/usr/bin/env python3
"""Phase 0 audit for the 2026-07-01 chemistry-dashboard-corrections spec.

READ-ONLY: makes NO changes to any parquet. Produces a review workbook +
markdown summary so Muhannad can approve the category rules, the name-group
merges (filter water / شطة), and the water failed-test recovery BEFORE Phase 1
applies anything.

The DRAFT rules below (canonical vocabulary, name keywords, per-section valid
allow-list, name-group merges) are proposals for review; after sign-off they
migrate into chemistry/scripts/categories.py.

Run: microbiology/.venv/bin/python chemistry/scripts/audit_categories.py
"""
from __future__ import annotations
import glob, re, sys
from collections import defaultdict, Counter
from pathlib import Path
import pandas as pd
import openpyxl

ROOT = Path(__file__).resolve().parents[1]           # chemistry/
CLEANED = ROOT / "cleaned"
RAW = ROOT / "raw"
OUT_XLSX = ROOT / "reports" / "category_audit_2026-07-01.xlsx"
OUT_MD = ROOT / "reports" / "category_audit_2026-07-01.md"

SECTIONS = ["aflatoxins", "food_chemistry", "heavy_metals", "honey",
            "hormones_antibiotics", "pesticides", "water_analysis"]

# ---------------------------------------------------------------- draft rules
# Canonical category vocabulary (clean Arabic labels).
C_CEREAL="الحبوب والبقوليات"; C_SPICE="البهارات والصوصات"; C_RTE="الأطعمة الجاهزة للأكل"
C_FRVEG="الفواكه والخضار"; C_SWEET="الحلويات والشوكولاتة"; C_BEV="المشروبات"
C_MEAT="اللحوم والدواجن"; C_FISH="الأسماك والمأكولات البحرية"; C_DAIRY="الحليب ومنتجات الألبان"
C_FAT="الدهون والزيوت"; C_FEED="الأعلاف"; C_HONEY="عسل"
W_TAP="مياه الحنفية"; W_FILTER="مياه فلتر"; W_DRINK="مياه شرب/معبأة"
C_UNKNOWN="غير مصنّف"

# Substring keyword -> canonical, matched against the (bilingual) raw category
# text. First hit wins. Handles trailing-quote junk and English+Arabic strings.
CAT_KEYWORDS = [
    ("فلتر", W_FILTER),
    ("حنفي", W_TAP), ("tap water", W_TAP),
    ("معبأ", W_DRINK), ("شرب", W_DRINK), ("bottled", W_DRINK),
    ("unbottled", W_DRINK), ("غير المعبأ", W_DRINK), ("متحرك", W_DRINK), ("drinking", W_DRINK),
    ("حبوب", C_CEREAL), ("بقول", C_CEREAL), ("cereal", C_CEREAL), ("legume", C_CEREAL),
    ("بهار", C_SPICE), ("صوص", C_SPICE), ("spice", C_SPICE), ("sauce", C_SPICE),
    ("جاهز", C_RTE), ("ready to eat", C_RTE),
    ("فواكه", C_FRVEG), ("خضار", C_FRVEG), ("fruit", C_FRVEG), ("vegetable", C_FRVEG),
    ("حلوي", C_SWEET), ("شوكولا", C_SWEET), ("شكولا", C_SWEET), ("sweet", C_SWEET), ("chocolate", C_SWEET),
    ("مشروب", C_BEV), ("beverage", C_BEV),
    ("لحوم", C_MEAT), ("دواجن", C_MEAT), ("meat", C_MEAT), ("poultry", C_MEAT),
    ("أسماك", C_FISH), ("اسماك", C_FISH), ("مأكولات", C_FISH), ("fish", C_FISH), ("seafood", C_FISH),
    ("ألبان", C_DAIRY), ("البان", C_DAIRY), ("حليب", C_DAIRY), ("dairy", C_DAIRY), ("milk", C_DAIRY),
    ("دهون", C_FAT), ("زيوت", C_FAT), ("oil", C_FAT), ("fat", C_FAT),
    ("اعلاف", C_FEED), ("أعلاف", C_FEED), ("fodder", C_FEED), ("feed", C_FEED),
    ("عسل", C_HONEY), ("honey", C_HONEY),
]

# Name keywords for rows with NO raw category (all of 2024). Reuses the
# dashboard's keyword table, mapped to the canonical vocabulary.
NAME_KEYWORDS = [
    ("فلتر", W_FILTER),
    ("موي", W_TAP), ("مياه", W_TAP), ("مياة", W_TAP), ("ماء", W_TAP), ("حنفي", W_TAP),
    ("سمك", C_FISH), ("تون", C_FISH), ("جمبري", C_FISH), ("روبيان", C_FISH), ("سلمون", C_FISH), ("بلطي", C_FISH),
    ("لحم", C_MEAT), ("دجاج", C_MEAT), ("فروج", C_MEAT), ("شاورما", C_MEAT), ("كباب", C_MEAT),
    ("شط", C_SPICE), ("صلصة", C_SPICE), ("صوص", C_SPICE), ("خل", C_SPICE), ("بهار", C_SPICE),
    ("فلفل", C_SPICE), ("كركم", C_SPICE), ("زنجبيل", C_SPICE), ("هيل", C_SPICE), ("قرفة", C_SPICE), ("كمون", C_SPICE),
    ("ارز", C_CEREAL), ("أرز", C_CEREAL), ("رز", C_CEREAL), ("قمح", C_CEREAL), ("عدس", C_CEREAL),
    ("حمص", C_CEREAL), ("فول", C_CEREAL), ("فاصولي", C_CEREAL), ("ذرة", C_CEREAL), ("شعير", C_CEREAL),
    ("لوز", C_CEREAL), ("فستق", C_CEREAL), ("كاجو", C_CEREAL), ("بندق", C_CEREAL), ("جوز", C_CEREAL),
    ("سمسم", C_CEREAL), ("ترمس", C_CEREAL), ("جريش", C_CEREAL), ("بصل مجفف", C_SPICE),
    ("عصير", C_BEV), ("شاي", C_BEV), ("قهوة", C_BEV), ("كركدي", C_BEV), ("نسكافيه", C_BEV),
    ("حليب", C_DAIRY), ("لبن", C_DAIRY), ("جبن", C_DAIRY), ("زبادي", C_DAIRY), ("قشطة", C_DAIRY),
    ("عسل", C_HONEY),
    ("زيت", C_FAT), ("سمن", C_FAT),
    ("مربى", C_SWEET), ("شوكولا", C_SWEET), ("حلاوة", C_SWEET), ("كاكاو", C_SWEET), ("بسكويت", C_SWEET),
    ("علف", C_FEED), ("اعلاف", C_FEED),
    ("بيض", C_RTE),
]

# Per-section VALID canonical categories (draft — Muhannad to confirm).
# "review" categories are allowed but flagged for a closer look.
SECTION_VALID = {
    "aflatoxins":           {C_CEREAL, C_SPICE, C_RTE, C_SWEET},
    "food_chemistry":       {C_CEREAL, C_SPICE, C_RTE, C_FRVEG, C_SWEET, C_BEV,
                             C_MEAT, C_FISH, C_DAIRY, C_FAT, C_FEED, C_HONEY},
    "heavy_metals":         {C_CEREAL, C_SPICE, C_RTE, C_FRVEG, C_SWEET, C_BEV,
                             C_MEAT, C_FISH, C_DAIRY, C_FAT, C_FEED, C_HONEY,
                             W_TAP, W_FILTER, W_DRINK},
    "honey":                {C_HONEY, C_RTE},
    "hormones_antibiotics": {C_MEAT, C_FISH, C_DAIRY},
    "pesticides":           {C_FRVEG, C_CEREAL, C_SPICE, C_RTE, C_DAIRY, C_FAT},
    "water_analysis":       {W_TAP, W_FILTER, W_DRINK},
}
SECTION_REVIEW = {   # allowed but worth a look
    "aflatoxins": {C_FRVEG},   # dried fruit ok, fresh veg not
}

# ------------------------------------------------------------- helpers
def norm(s) -> str:
    return "" if s is None else str(s).strip().strip('"').strip().lower()

def cat_canonical(raw) -> str | None:
    if raw is None or (isinstance(raw, float)) or str(raw) in ("<NA>", "nan", "None", ""):
        return None
    s = norm(raw)
    if not s:
        return None
    for kw, canon in CAT_KEYWORDS:
        if kw.lower() in s:
            return canon
    return C_UNKNOWN

def name_canonical(name) -> str | None:
    s = norm(name)
    if not s:
        return None
    for kw, canon in NAME_KEYWORDS:
        if kw.lower() in s:
            return canon
    return None

SHATTA = re.compile(r"شط[ةه]")           # matches شطة/شطه/الشطة; excludes وشط (coffee)
def name_group(name) -> str | None:
    """D4/D5 display-name grouping. Returns a group label or None (keep name)."""
    s = norm(name)
    if not s:
        return None
    if "فلتر" in s:
        return W_FILTER
    if SHATTA.search(s):
        return "شطة"
    return None

def load(sec):
    out = {}
    for f in sorted(glob.glob(str(CLEANED / f"chem_{sec}_*.parquet"))):
        yr = re.search(r"_(\d{4})\.parquet", f).group(1)
        out[yr] = pd.read_parquet(f)
    return out

# ------------------------------------------------------- red-cell extraction (M1 2024)
def extract_water_2024_redcells():
    """Return {sample_id: [failed params]} from red-filled cells in the 2024
    water monthly sheets (red fill = FFFF0000)."""
    f = RAW / "2024" / "Water analysis section.xlsx"
    if not f.exists():
        return {}, "raw file missing"
    wb = openpyxl.load_workbook(f)
    skip = {"Copy of Water analysis section", "invalid samples 2024"}
    META_COLS = {"no.", "receiving date", "sampling date", "sample name", "sample id",
                 "facility name", "municipality name", "municipality name ", "district name",
                 "street name", "sample notes", "testing notes", "license number", "رقم الرخصة",
                 "analysis section", "valid/ invalid", "valid/invalid", "valid /invalid",
                 "matched/not matched", "matched/ not matched", "compliant / non-compliant",
                 "invalid test", "non-compliant test", "no. of tests", "no of tests",
                 "qc for invalid test", "result qc for invalid test", ""}
    failed = defaultdict(list)
    for sh in wb.sheetnames:
        if sh in skip:
            continue
        ws = wb[sh]
        # locate header row + Sample ID column
        header_row = None; headers = {}
        for r in range(1, min(6, ws.max_row) + 1):
            vals = {c.column: (str(c.value).split("\n")[0].strip() if c.value else "")
                    for c in ws[r]}
            if any(v.lower() == "sample id" for v in vals.values()):
                header_row = r; headers = vals; break
        if header_row is None:
            continue
        sid_col = next((col for col, h in headers.items() if h.lower() == "sample id"), None)
        if sid_col is None:
            continue
        for row in ws.iter_rows(min_row=header_row + 1):
            sid_cell = ws.cell(row=row[0].row, column=sid_col)
            sid = str(sid_cell.value).strip() if sid_cell.value else None
            if not sid:
                continue
            for c in row:
                fl = c.fill
                fg = getattr(fl.fgColor, "rgb", None) if (fl and fl.patternType == "solid") else None
                if fg == "FFFF0000":
                    h = headers.get(c.column, "")
                    if h and h.lower() not in META_COLS:
                        failed[sid.lower()].append(h)
    # dedupe preserve order
    return {k: list(dict.fromkeys(v)) for k, v in failed.items()}, "ok"

# ------------------------------------------------------- build audit tables
def main():
    dist_rows, suspect_rows, unclassified_rows = [], [], []
    filter_rows, shatta_rows = [], []
    for sec in SECTIONS:
        for yr, df in load(sec).items():
            has_cat = df["sample_category"].notna().any()
            for r in df.itertuples(index=False):
                raw = getattr(r, "sample_category", None)
                nm = getattr(r, "sample_name", None)
                sid = getattr(r, "sample_id", None)
                canon = cat_canonical(raw) or name_canonical(nm) or C_UNKNOWN
                valid = SECTION_VALID.get(sec, set())
                review = SECTION_REVIEW.get(sec, set())
                if canon == C_UNKNOWN:
                    status = "UNCLASSIFIED"          # name-guesser gap, not a mislabel
                elif canon in valid:
                    status = "ok"
                elif canon in review:
                    status = "review"
                else:
                    status = "SUSPECT"               # known category, wrong for this section
                dist_rows.append((sec, yr, ("" if raw is None else str(raw)),
                                  canon, "name" if (raw is None or pd.isna(raw)) else "category", status))
                if status in ("SUSPECT", "review"):
                    suspect_rows.append((sec, yr, str(sid), str(nm),
                                         ("" if raw is None else str(raw)), canon, status))
                if status == "UNCLASSIFIED":
                    unclassified_rows.append((sec, yr, str(nm)))
                grp = name_group(nm)
                if grp == W_FILTER:
                    filter_rows.append((sec, yr, str(sid), str(nm)))
                elif grp == "شطة":
                    shatta_rows.append((sec, yr, str(sid), str(nm)))

    # distribution collapsed to counts
    dist_counter = Counter((s, y, raw, canon, src, st)
                           for (s, y, raw, canon, src, st) in dist_rows)
    dist_df = pd.DataFrame(
        [(s, y, raw, canon, src, st, n) for (s, y, raw, canon, src, st), n in dist_counter.items()],
        columns=["section", "year", "raw_category", "proposed_canonical",
                 "source", "status", "count"]
    ).sort_values(["section", "year", "status", "count"], ascending=[True, True, True, False])

    suspect_df = pd.DataFrame(suspect_rows, columns=[
        "section", "year", "sample_id", "sample_name", "raw_category",
        "proposed_canonical", "status"])

    # top unclassified names per section/year (coverage gap → extend rules Phase 1)
    unc_df = pd.DataFrame(
        Counter((s, y, nm) for (s, y, nm) in unclassified_rows).items(),
        columns=["key", "count"])
    if len(unc_df):
        unc_df[["section", "year", "sample_name"]] = pd.DataFrame(unc_df["key"].tolist(), index=unc_df.index)
        unc_df = unc_df.drop(columns="key")[["section", "year", "sample_name", "count"]] \
                       .sort_values(["section", "year", "count"], ascending=[True, True, False])

    filt_df = pd.DataFrame(
        Counter((s, y, nm) for (s, y, sid, nm) in filter_rows).items(),
        columns=["key", "count"])
    if len(filt_df):
        filt_df[["section", "year", "sample_name"]] = pd.DataFrame(filt_df["key"].tolist(), index=filt_df.index)
        filt_df = filt_df.drop(columns="key")[["section", "year", "sample_name", "count"]]
        filt_df["merged_into"] = W_FILTER

    shat_df = pd.DataFrame(
        Counter((s, y, nm) for (s, y, sid, nm) in shatta_rows).items(),
        columns=["key", "count"])
    if len(shat_df):
        shat_df[["section", "year", "sample_name"]] = pd.DataFrame(shat_df["key"].tolist(), index=shat_df.index)
        shat_df = shat_df.drop(columns="key")[["section", "year", "sample_name", "count"]]
        shat_df["merged_into"] = "شطة"

    # water M1
    m1_2025 = []
    w25 = load("water_analysis").get("2025")
    if w25 is not None:
        inv = w25[w25["validity_status"].astype(str) == "invalid"]
        for r in inv.itertuples(index=False):
            it = getattr(r, "invalid_test", None)
            m1_2025.append((str(getattr(r, "sample_id", "")), str(getattr(r, "sample_name", "")),
                            "" if it is None or pd.isna(it) else str(it),
                            "surfaced" if (it is not None and not pd.isna(it) and str(it).strip()) else "STILL EMPTY"))
    m1_2025_df = pd.DataFrame(m1_2025, columns=["sample_id", "sample_name", "invalid_test", "outcome"])

    red_map, red_status = extract_water_2024_redcells()
    m1_2024 = []
    w24 = load("water_analysis").get("2024")
    if w24 is not None:
        inv = w24[w24["validity_status"].astype(str) == "invalid"]
        for r in inv.itertuples(index=False):
            sid = str(getattr(r, "sample_id", ""))
            recovered = red_map.get(sid.lower(), [])
            m1_2024.append((sid, str(getattr(r, "sample_name", "")),
                            ", ".join(recovered),
                            "recovered" if recovered else "no red cell found"))
    m1_2024_df = pd.DataFrame(m1_2024, columns=["sample_id", "sample_name", "failed_tests_from_red", "outcome"])

    valid_df = pd.DataFrame(
        [(sec, ", ".join(sorted(SECTION_VALID.get(sec, set()))),
          ", ".join(sorted(SECTION_REVIEW.get(sec, set()))))
         for sec in SECTIONS],
        columns=["section", "valid_categories (draft)", "review_categories (draft)"])

    # ------------------------------------------------------ write xlsx
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as xw:
        valid_df.to_excel(xw, sheet_name="proposed_valid_cats", index=False)
        dist_df.to_excel(xw, sheet_name="category_distribution", index=False)
        suspect_df.to_excel(xw, sheet_name="suspect_rows", index=False)
        (unc_df if len(unc_df) else pd.DataFrame(columns=["section","year","sample_name","count"])).to_excel(xw, sheet_name="unclassified_names", index=False)
        (filt_df if len(filt_df) else pd.DataFrame(columns=["section","year","sample_name","count","merged_into"])).to_excel(xw, sheet_name="merge_filter_water", index=False)
        (shat_df if len(shat_df) else pd.DataFrame(columns=["section","year","sample_name","count","merged_into"])).to_excel(xw, sheet_name="merge_shatta", index=False)
        m1_2025_df.to_excel(xw, sheet_name="water_M1_2025", index=False)
        m1_2024_df.to_excel(xw, sheet_name="water_M1_2024_red", index=False)

    # ------------------------------------------------------ markdown summary
    n_suspect = (suspect_df["status"] == "SUSPECT").sum() if len(suspect_df) else 0
    n_review = (suspect_df["status"] == "review").sum() if len(suspect_df) else 0
    n_unc = len(unclassified_rows)
    lines = []
    lines.append("# Chemistry category audit — 2026-07-01 (Phase 0, no changes applied)\n")
    lines.append(f"Workbook: `{OUT_XLSX.name}` (8 sheets). Rules are DRAFT — confirm before Phase 1.\n")
    lines.append("Three buckets, kept separate on purpose:\n")
    lines.append("- **SUSPECT** = a known category that is WRONG for the section (the real mislabels — D1/D2/D3).")
    lines.append("- **review** = allowed but worth a look (e.g. fruit/veg in aflatoxins — dried ok, fresh not).")
    lines.append("- **UNCLASSIFIED** = 2024 rows with no source category the name-guesser couldn't place — a coverage gap, NOT a mislabel. Phase 1 fixes with section-aware guessing.\n")
    lines.append("## SUSPECT — genuine cross-section mislabels (D1/D2/D3)\n")
    lines.append(f"- **{n_suspect} SUSPECT** + **{n_review} review**. See `suspect_rows` sheet.\n")
    if len(suspect_df):
        by = suspect_df.groupby(["section", "year", "proposed_canonical", "status"]).size().reset_index(name="n")
        for _, x in by.sort_values(["status","n"], ascending=[True,False]).iterrows():
            lines.append(f"  - {x['status']}: {x['section']} {x['year']} → {x['proposed_canonical']} ({x['n']})")
    lines.append(f"\n## UNCLASSIFIED — 2024 name-guesser coverage gap ({n_unc} rows)\n")
    lines.append("- Not mislabels — no source category exists in 2024. Top unmatched names are in the `unclassified_names` sheet; Phase 1 will extend the section-aware name rules to absorb them.")
    if len(unc_df):
        gg = unc_df.groupby(["section","year"])["count"].sum().reset_index(name="n")
        for _, x in gg.sort_values("n", ascending=False).iterrows():
            lines.append(f"  - {x['section']} {x['year']}: {x['n']} rows")
    lines.append("\n## D4 — filter-water name merge → «مياه فلتر»\n")
    lines.append(f"- {len(filter_rows)} rows across {len(filt_df) if len(filt_df) else 0} distinct names. See `merge_filter_water`.\n")
    lines.append("## D5 — شطة name merge → «شطة»\n")
    lines.append(f"- {len(shatta_rows)} rows across {len(shat_df) if len(shat_df) else 0} distinct names. See `merge_shatta`.\n")
    lines.append("## M1 — water failed-test recovery\n")
    if len(m1_2025_df):
        surf = (m1_2025_df["outcome"] == "surfaced").sum()
        lines.append(f"- 2025: {len(m1_2025_df)} invalid water samples, **{surf}** get failed-tests surfaced from `invalid_test` "
                     f"({len(m1_2025_df)-surf} still empty). See `water_M1_2025`.")
    if len(m1_2024_df):
        rec = (m1_2024_df["outcome"] == "recovered").sum()
        lines.append(f"- 2024: {len(m1_2024_df)} invalid water samples, **{rec}** recovered from red cells "
                     f"(red-scan: {red_status}). See `water_M1_2024_red`.")
    lines.append("\n## Proposed per-section valid categories (draft — confirm)\n")
    for sec in SECTIONS:
        v = ", ".join(sorted(SECTION_VALID.get(sec, set())))
        rv = SECTION_REVIEW.get(sec, set())
        extra = f"  _(review: {', '.join(sorted(rv))})_" if rv else ""
        lines.append(f"- **{sec}**: {v}{extra}")
    OUT_MD.write_text("\n".join(lines), encoding="utf-8")

    print(f"wrote {OUT_XLSX}")
    print(f"wrote {OUT_MD}")
    print(f"SUSPECT={n_suspect}  review={n_review}  filter_water_rows={len(filter_rows)}  "
          f"shatta_rows={len(shatta_rows)}  red2024_samples={sum(1 for x in m1_2024 if x[3]=='recovered')}")

if __name__ == "__main__":
    sys.exit(main())
