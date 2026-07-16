"""Shared utilities for chemistry-section cleaners (v2).

Every cleaner reads a section schema (YAML) describing logical columns + their
header aliases. This module provides the year-agnostic helpers all cleaners
reuse. Includes every correctness fix from the v1.5 audit pass.
"""
from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
import yaml
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SCHEMA_DIR = ROOT / "schemas"
CLEAN_DIR  = ROOT / "cleaned"
REPORT_DIR = ROOT / "reports"
RAW_DIR    = ROOT / "raw"

ND_TOKENS    = {"N.D", "ND", "n.d", "n.d.", "Not detected", "غير موجودة", "غير موجود", "غيرموجودة"}
NA_TOKENS    = {"NA", "N/A", "-", "—", "not applicable", "كل المبيدات اقل 0.01",
                "تراكيز المبيدات أقل من 0.01"}
EMPTY_TOKENS = {"", "  "}

ZERO_WIDTH = "​‌‍﻿"
RTL_MARKS  = "‎‏‪‫‬‭‮⁦⁧⁨⁩"
NBSP       = " "

ARABIC_INDIC = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")
EASTERN      = str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789")

SAMPLE_ID_RE      = re.compile(r"^([A-Za-z]+[A-Za-z0-9]*)-(\d+)-([Rr])(\d+)$")
SAMPLE_ID_RE_4SEG = re.compile(r"^([A-Za-z]+(?:-[A-Za-z0-9]+)*)-(\d+)-([Rr])(\d+)$")

MONTH_PATTERNS = {
    "jan":1,"january":1,"feb":2,"february":2,
    "mar":3,"march":3,"apr":4,"april":4,"may":5,
    "jun":6,"june":6,"jul":7,"july":7,"aug":8,"august":8,
    "sep":9,"sept":9,"september":9,
    "oct":10,"october":10,"nov":11,"november":11,
    "dec":12,"des":12,"december":12,
}


# ---------------------------------------------------------------------------
# Text & numeric helpers
# ---------------------------------------------------------------------------
def normalise_text(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    if not isinstance(v, str):
        return v
    s = unicodedata.normalize("NFKC", v)
    for ch in ZERO_WIDTH + RTL_MARKS:
        s = s.replace(ch, "")
    s = s.replace(NBSP, " ").replace("\t", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s or None


def header_en(v: Any) -> str | None:
    """English half of bilingual 'English\\nArabic' header, normalised."""
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    s = str(v)
    first = s.split("\n", 1)[0]
    return normalise_text(first)


def header_full(v: Any) -> str | None:
    """Full bilingual header normalised (newlines → space). Disambiguates
    columns sharing the same English half (e.g. 4 sensory 'اختبار حسي' cols)."""
    if v is None:
        return None
    if isinstance(v, float) and pd.isna(v):
        return None
    return normalise_text(str(v))


def parse_nd_to_float(v: Any) -> tuple[float | None, bool]:
    """For RESULT columns. N.D → (0.0, True). Numeric → (float, False).
    NA / empty → (None, False)."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None, False
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v), False
    if not isinstance(v, str):
        return None, False
    s = v.strip()
    if not s or s in EMPTY_TOKENS:
        return None, False
    s_clean = re.sub(r"\s+", " ", s)
    if s_clean in ND_TOKENS:
        return 0.0, True
    if s_clean in NA_TOKENS:
        return None, False
    body = s_clean
    for p in ("<=", ">=", "≤", "≥", "<", ">"):
        if body.startswith(p):
            body = body[len(p):].strip()
            break
    body = re.sub(r"^(Max|max|Min|min)\s*", "", body)
    body = body.translate(ARABIC_INDIC).translate(EASTERN)
    body = body.replace(",", "").replace("٬", "").replace("٫", ".")
    body = re.sub(r"[^\d.\-eE]", "", body)
    if body in ("", "-", ".", "-.", "."):
        return None, False
    try:
        return float(body), False
    except ValueError:
        return None, False


def parse_limit_to_float(v: Any) -> float | None:
    """For LIMIT columns. N.D in a limit means 'limit not set' → None
    (never 0.0, which would cause false exceedance flags)."""
    out, is_nd = parse_nd_to_float(v)
    if is_nd:
        return None
    return out


def parse_date(v: Any) -> date | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if not isinstance(v, str):
        return None
    s = v.strip()
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d-%m-%Y", "%d/%m/%Y",
                "%m/%d/%Y", "%d-%m-%y", "%d/%m/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    m = re.search(r"(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})", s)
    if m:
        dd, mm, yy = m.groups()
        try:
            yy_int = int(yy)
            if yy_int < 100:
                yy_int += 2000
            return date(yy_int, int(mm), int(dd))
        except ValueError:
            return None
    return None


def normalise_sample_id(v: Any) -> str | None:
    if v is None or (isinstance(v, float) and pd.isna(v)) or not isinstance(v, str):
        return None
    s = v.strip()
    m = SAMPLE_ID_RE.match(s)
    if m:
        prefix, digits, r, vnum = m.groups()
        return f"{prefix.lower()}-{digits}-r{vnum}"
    m = SAMPLE_ID_RE_4SEG.match(s)
    if m:
        prefix, digits, r, vnum = m.groups()
        return f"{prefix.lower()}-{digits}-r{vnum}"
    return s.lower() or None


def sheet_year_month(sheet_name: str, default_year: int) -> tuple[int, int] | None:
    """Extract (year, month) from a sheet name like 'Jan-2025', 'March 2025',
    'Feb-25', 'JAN-2026' etc. Returns None for non-monthly sheets."""
    s = sheet_name.lower().strip()
    s = re.sub(r"[-_/\.]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    parts = s.split(" ")
    mn = None
    for p in parts:
        p2 = re.sub(r"\d+", "", p).strip()
        if p2 in MONTH_PATTERNS:
            mn = MONTH_PATTERNS[p2]
            break
    if mn is None:
        return None
    year = None
    for p in parts:
        digs = re.findall(r"\d+", p)
        for d in digs:
            if len(d) == 4:
                year = int(d)
                break
            elif len(d) == 2:
                year = 2000 + int(d)
                break
        if year is not None:
            break
    return (year or default_year, mn)


# ---------------------------------------------------------------------------
# Workbook iteration
# ---------------------------------------------------------------------------
def iter_data_sheets(path: Path, year: int, schema: dict):
    """Yields (sheet_name, (year, month), rows_list_of_dicts).

    Filters applied:
      * sheet name starts with 'final' (case-insensitive)  → skip (summary tabs)
      * sheet name contains any schema.skip_sheets token   → skip
      * non-monthly sheet name when require_monthly_sheet  → skip
      * detected year != file's declared year              → skip (year purity)
      * sheet with no data rows                            → skip naturally

    Section may set 'single_sheet: true' to allow a non-monthly single sheet
    (used by Honey which has one big sheet covering the whole year).

    Header alias matching is case-insensitive on Latin characters, and tries
    the bilingual full header if header_en fails.
    """
    skip_tokens = [t.lower() for t in schema.get("skip_sheets", [])]
    only_sheets = [t.strip().lower() for t in schema.get("only_sheets", [])]
    header_row_max = int(schema.get("header_row_max", 5))
    single_sheet   = bool(schema.get("single_sheet", False))
    require_month  = bool(schema.get("require_monthly_sheet", True)) and not single_sheet

    wb = load_workbook(path, data_only=True, read_only=False, keep_links=False)
    for sn in wb.sheetnames:
        sn_low = sn.lower().lstrip()
        if sn_low.startswith("final"):
            continue
        if any(t in sn_low for t in skip_tokens):
            continue
        if only_sheets:
            sn_norm = sn.strip().lower()
            if not any(sn_norm == t or sn_norm.startswith(t) for t in only_sheets):
                continue

        ws = wb[sn]
        max_r = ws.max_row or 0
        max_c = ws.max_column or 0
        if max_r < 2 or max_c < 3:
            continue

        ym = sheet_year_month(sn, default_year=year)
        if ym is None and require_month:
            continue
        if ym is not None and ym[0] != year:
            continue
        # Single-sheet sections fall through with ym=None → synthesise (year, 1).
        if ym is None and single_sheet:
            ym = (year, 1)

        # Locate header row.
        header_row = None
        for r in range(1, min(header_row_max, max_r) + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_c + 1)]
            non_null = sum(1 for v in row_vals if v is not None)
            if non_null >= max(3, max_c // 3):
                header_row = r
                break
        if header_row is None:
            continue

        header_vals = [ws.cell(row=header_row, column=c).value for c in range(1, max_c + 1)]
        header_enh  = [header_en(v) for v in header_vals]
        header_full_list = [header_full(v) for v in header_vals]
        header_enh_lc    = [h.lower() if h else None for h in header_enh]
        header_full_lc   = [h.lower() if h else None for h in header_full_list]

        col_map: dict[str, int] = {}
        used_indices: set[int] = set()
        for logical, aliases in schema.get("columns", {}).items():
            for alias in aliases:
                alias_norm = normalise_text(alias)
                if alias_norm is None:
                    continue
                alias_lc = alias_norm.lower()
                found_i = None
                for i, h in enumerate(header_enh_lc):
                    if h and h == alias_lc and i not in used_indices:
                        found_i = i
                        break
                if found_i is None:
                    for i, h in enumerate(header_full_lc):
                        if h and h == alias_lc and i not in used_indices:
                            found_i = i
                            break
                if found_i is not None:
                    col_map[logical] = found_i
                    used_indices.add(found_i)
                    break

        # Paired-column groups (e.g. Heavy Metals, Hormones): result column +
        # immediately-next limit column (+ optional QC% column).
        for pair in schema.get("paired_columns", []):
            r_aliases = pair.get("result_aliases") or [pair.get("result_alias")]
            r_aliases = [a for a in r_aliases if a]
            limit_aliases = pair.get("limit_aliases") or [pair.get("limit_alias", "Limit (ppb)")]
            limit_aliases_lc = [normalise_text(a).lower() for a in limit_aliases if normalise_text(a)]
            qc_alias_lc = (normalise_text(pair.get("qc_alias", "QC.RC%")) or "").lower()
            found_col = None
            for ra in r_aliases:
                ra_lc = (normalise_text(ra) or "").lower()
                for i, h in enumerate(header_enh_lc):
                    if h and h == ra_lc and i not in used_indices:
                        found_col = i
                        break
                if found_col is not None:
                    break
            if found_col is None:
                continue
            col_map[pair["result_key"]] = found_col
            used_indices.add(found_col)
            if found_col + 1 < len(header_enh_lc) and header_enh_lc[found_col + 1] in limit_aliases_lc:
                col_map[pair["limit_key"]] = found_col + 1
                used_indices.add(found_col + 1)
            if pair.get("qc_key") and found_col + 2 < len(header_enh_lc) and header_enh_lc[found_col + 2] == qc_alias_lc:
                col_map[pair["qc_key"]] = found_col + 2
                used_indices.add(found_col + 2)

        # Each data row → dict of logical names.
        rows = []
        for r in range(header_row + 1, max_r + 1):
            row_dict = {}
            any_value = False
            for logical, ci in col_map.items():
                v = ws.cell(row=r, column=ci + 1).value
                if v is not None:
                    any_value = True
                row_dict[logical] = v
            if any_value:
                rows.append(row_dict)
        if rows:
            yield sn, ym, rows
    wb.close()


def load_schema(section: str) -> dict:
    path = SCHEMA_DIR / f"chem_{section}.yaml"
    with path.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def write_parquet(records: list[dict], path: Path, string_cols: list[str], int_cols: list[str] | None = None) -> int:
    if not records:
        return 0
    df = pd.DataFrame.from_records(records)
    for c in string_cols:
        if c in df.columns:
            df[c] = pd.array(
                [None if (v is None or (isinstance(v, float) and pd.isna(v))) else str(v) for v in df[c]],
                dtype="string",
            )
    for c in (int_cols or []):
        if c in df.columns:
            df[c] = pd.array(
                [int(v) if isinstance(v, (int, float)) and not isinstance(v, bool) and not pd.isna(v) else pd.NA for v in df[c]],
                dtype="Int64",
            )
    if "sampling_date" in df.columns:
        df["sampling_date"] = pd.to_datetime(df["sampling_date"], errors="coerce").astype("datetime64[ms]")
    if "testing_date" in df.columns:
        df["testing_date"] = pd.to_datetime(df["testing_date"], errors="coerce").astype("datetime64[ms]")
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(path, compression="zstd", index=False)
    return len(df)
