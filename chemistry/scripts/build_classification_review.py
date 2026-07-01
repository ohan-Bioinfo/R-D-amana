#!/usr/bin/env python3
"""Compile every classification/mapping table into one review workbook + md so
Muhannad can review the taxonomy and guide changes. READ-ONLY.

Covers: GSO 1016 categories, canonical→GSO map, per-section valid categories
(with proposed aflatoxin tightening), name-keyword rules, name-groups, and the
municipality→sector map (with chemistry-municipality coverage).

Run: microbiology/.venv/bin/python chemistry/scripts/build_classification_review.py
"""
from __future__ import annotations
import glob, re
from pathlib import Path
import pandas as pd
import yaml

ROOT = Path(__file__).resolve().parents[1]
import sys; sys.path.insert(0, str(Path(__file__).parent))
import categories as C
import sectors as SEC
# NOTE: do NOT import build_dashboard — its module-level import breaks the
# openpyxl xlsx writer. The GSO list below mirrors its CHEM_TO_GSO values.

OUT_XLSX = ROOT / "reports" / "classification_review_2026-07-01.xlsx"
OUT_MD = ROOT / "reports" / "classification_review_2026-07-01.md"

# --- GSO 1016 category list (the official values used by build_dashboard's
#     CHEM_TO_GSO bridge) ---
gso_categories = sorted({
    "Fruit and Vegetables", "Cereals; Legumes and their Products",
    "Tomato Concentrates, Sauces, Vinegar, Spices and Herbs", "Ready to Eat Foods",
    "Meat, Poultry and its Products", "Chocolate, Sweets and their Ingredients",
    "Dairy Products", "Beverages", "Fish and Shellfish their Products",
    "Egg and Egg Products", "Fats and Oils", "Drinking Water",
    "Jelly, Jam and Marmalade",
    "Infants, Children and Certain Categories of Dietetic Foods",
    "Animal Feed", "Miscellaneous Foods",
})

# --- canonical → GSO map (my canonical vocab bridged to GSO) ---
CANON_TO_GSO = {
    C.C_CEREAL: "Cereals; Legumes and their Products",
    C.C_SPICE:  "Tomato Concentrates, Sauces, Vinegar, Spices and Herbs",
    C.C_RTE:    "Ready to Eat Foods",
    C.C_FRVEG:  "Fruit and Vegetables",
    C.C_SWEET:  "Chocolate, Sweets and their Ingredients",
    C.C_BEV:    "Beverages",
    C.C_MEAT:   "Meat, Poultry and its Products",
    C.C_FISH:   "Fish and Shellfish their Products",
    C.C_DAIRY:  "Dairy Products",
    C.C_FAT:    "Fats and Oils",
    C.C_FEED:   "Animal Feed",
    C.C_HONEY:  "Ready to Eat Foods",
    C.W_TAP:    "Drinking Water",
    C.W_FILTER: "Drinking Water",
    C.W_DRINK:  "Drinking Water",
    C.C_MISC:   "Miscellaneous Foods",
}

# --- proposed aflatoxin tightening (per Muhannad 2026-07-01: no RTE/meat/bev) ---
AFLATOXIN_PROPOSED_VALID = {C.C_CEREAL, C.C_SPICE, C.C_SWEET}   # +nuts fold into cereals
AFLATOXIN_PROPOSED_REVIEW = {C.C_FRVEG}                          # dried fruit only

# --- sector taxonomy (from microbio schema) ---
SCHEMA = ROOT.parent / "microbiology" / "schemas" / "lab_data_2025_v1.yaml"
st = yaml.safe_load(open(SCHEMA, encoding="utf-8"))["sector_taxonomy"]
sub_to_sector = {sub: sec["name"] for sec in st["sectors"] for sub in sec.get("sub_municipalities", [])}


def strip_baladiya(m):
    if not m: return None
    return re.sub(r"^\s*بلدية\s*", "", str(m)).strip()


def main():
    # ---- sheet 1: GSO categories
    gso_df = pd.DataFrame({"GSO_1016_category": gso_categories})

    # ---- sheet 2: canonical → GSO
    canon_df = pd.DataFrame(
        [(k, CANON_TO_GSO.get(k, "?")) for k in
         [C.C_CEREAL, C.C_SPICE, C.C_RTE, C.C_FRVEG, C.C_SWEET, C.C_BEV, C.C_MEAT,
          C.C_FISH, C.C_DAIRY, C.C_FAT, C.C_FEED, C.C_HONEY, C.W_TAP, C.W_FILTER,
          C.W_DRINK, C.C_MISC]],
        columns=["canonical_category (mine)", "GSO_1016_category"])

    # ---- sheet 3: per-section valid categories (current + proposed)
    sec_rows = []
    for sec, valid in C.SECTION_VALID.items():
        rev = C.SECTION_REVIEW.get(sec, set())
        proposed_v = AFLATOXIN_PROPOSED_VALID if sec == "aflatoxins" else valid
        proposed_r = AFLATOXIN_PROPOSED_REVIEW if sec == "aflatoxins" else rev
        sec_rows.append((sec, "، ".join(sorted(valid)), "، ".join(sorted(rev)),
                         "، ".join(sorted(proposed_v)), "، ".join(sorted(proposed_r)),
                         "REMOVE RTE/meat/bev; nuts→cereals" if sec == "aflatoxins" else ""))
    sec_df = pd.DataFrame(sec_rows, columns=[
        "section", "valid_now", "review_now", "valid_PROPOSED", "review_PROPOSED", "note"])

    # ---- sheet 4: name-keyword rules (name → canonical)
    kw_df = pd.DataFrame(C.NAME_KEYWORDS, columns=["name_keyword", "→ canonical"])

    # ---- sheet 5: name-groups (D4/D5 + proposed فلفل)
    ng_rows = [("فلتر (in name)", C.W_FILTER, "applied", "filter water"),
               ("شط[ةه]", "شطة", "applied", "hot sauce variants"),
               ("فلفل (proposed)", "فلفل", "PROPOSED", "black/white/red/cherry pepper — 74 variants, 463 rows in aflatoxin")]
    ng_df = pd.DataFrame(ng_rows, columns=["match", "group_label", "status", "note"])

    # ---- sheet 6: municipality → sector (coverage over chemistry data)
    muni = {}
    for f in glob.glob(str(ROOT / "cleaned" / "chem_*.parquet")):
        df = pd.read_parquet(f)
        if "municipality" in df.columns:
            for m in df["municipality"].dropna().astype(str):
                muni[m] = muni.get(m, 0) + 1
    map_rows = []
    for m, n in sorted(muni.items(), key=lambda x: -x[1]):
        sec, flag = SEC.sector_for(m)
        map_rows.append((m, strip_baladiya(m), sec or f"★ {flag}", n))
    map_df = pd.DataFrame(map_rows, columns=["municipality (raw)", "stripped", "sector", "rows"])

    # ---- sector taxonomy reference
    tax_rows = [(sec["name"], "، ".join(sec.get("sub_municipalities", []))) for sec in st["sectors"]]
    tax_df = pd.DataFrame(tax_rows, columns=["sector", "sub_municipalities"])

    # Write via openpyxl directly — pandas ExcelWriter is unreliable with the
    # installed pandas 3.0 / openpyxl 3.1.5 combo ("At least one sheet must be
    # visible"). Direct openpyxl is what export_to_xlsx.py uses successfully.
    from openpyxl import Workbook
    from openpyxl.styles import Font
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    sheets = [
        ("GSO_1016_categories", gso_df), ("canonical_to_GSO", canon_df),
        ("section_valid_cats", sec_df), ("name_keyword_rules", kw_df),
        ("name_groups", ng_df), ("sector_taxonomy", tax_df),
        ("municipality_to_sector", map_df),
    ]
    for name, df in sheets:
        ws = wb.create_sheet(name[:31])
        ws.append(list(df.columns))
        for c in ws[1]:
            c.font = Font(bold=True)
        for _, row in df.iterrows():
            ws.append(["" if pd.isna(v) else str(v) for v in row])
        ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)

    unmapped = map_df[map_df["sector"] == "★ unmapped"]      # true junk only
    private = map_df[map_df["sector"] == "★ private"]
    lines = ["# Chemistry classification review — 2026-07-01\n",
             f"Workbook `{OUT_XLSX.name}` (7 sheets) for your review/guidance.\n",
             "## GSO 1016 categories (official 15)\n"]
    lines += [f"- {c}" for c in gso_categories]
    lines.append("\n## Per-section valid categories — PROPOSED change (aflatoxin)\n")
    lines.append("- **aflatoxins now:** " + "، ".join(sorted(C.SECTION_VALID['aflatoxins'])))
    lines.append("- **aflatoxins PROPOSED:** " + "، ".join(sorted(AFLATOXIN_PROPOSED_VALID)) +
                 f"  (review: {'، '.join(sorted(AFLATOXIN_PROPOSED_REVIEW))})")
    lines.append("  - Removes RTE / meat / beverage. The 197 'RTE' aflatoxin rows are actually NUTS "
                 "(لوز/فستق/كاجو) → reclassify to الحبوب والبقوليات by name.")
    lines.append("\n## Name-groups\n")
    lines.append("- Applied: فلتر → «مياه فلتر» ، شط[ةه] → «شطة»")
    lines.append("- **PROPOSED:** فلفل → «فلفل» (74 variants / 463 rows fragment the aflatoxin top-10)")
    lines.append("\n## Municipality → sector coverage (chemistry)\n")
    lines.append(f"- {len(map_df)} distinct municipality values → mapped after normalization; "
                 f"**{len(unmapped)} true-junk values ({unmapped['rows'].sum()} rows)** remain "
                 f"(sample names leaked into the municipality column); "
                 f"{len(private)} private-sample values ({private['rows'].sum()} rows) have no sector.")
    lines.append("- Sector column NOW added to the chemistry parquets. Rows with no municipality "
                 "at all (mostly 2024) get flag `no_municipality` and no sector.")
    if len(unmapped):
        lines.append("\n### True-junk municipality values (sample names in wrong column)\n")
        for _, r in unmapped.head(25).iterrows():
            lines.append(f"  - {r['municipality (raw)']}  ({r['rows']} rows)")
    OUT_MD.write_text("\n".join(lines), encoding="utf-8")
    print(f"wrote {OUT_XLSX}\nwrote {OUT_MD}")
    print(f"municipalities={len(map_df)} unmapped={len(unmapped)} ({unmapped['rows'].sum()} rows)")


if __name__ == "__main__":
    main()
