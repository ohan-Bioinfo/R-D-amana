"""Generic chemistry cleaner — section-driven via YAML schemas (v2).

Run:
    .venv/bin/python scripts/clean_chemistry.py --section <name>
    .venv/bin/python scripts/clean_chemistry.py --section all

Sections (loaded from chemistry/schemas/chem_<section>.yaml):
    aflatoxins, food_chemistry, heavy_metals, honey,
    hormones_antibiotics, pesticides, water_analysis

Output:
    chemistry/cleaned/chem_<section>_<year>.parquet  (wide, per-sample)
    chemistry/reports/chem_<section>_<year>.md       (audit summary)
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any

import pandas as pd

sys.path.insert(0, str(Path(__file__).parent))
from _common import (
    RAW_DIR, SCHEMA_DIR, CLEAN_DIR, REPORT_DIR,
    iter_data_sheets, load_schema,
    normalise_text, parse_nd_to_float, parse_limit_to_float, parse_date,
    normalise_sample_id, write_parquet,
)
import categories  # canonical sample-category classification (2026-07-01 spec)


def water_redcell_failed_map(src_path: Path) -> dict[str, str]:
    """M1 (water 2024): the failing parameter of each invalid water sample is
    marked only as a red cell-fill (FFFF0000) in the monthly sheets. Return
    {sample_id_lower: "param1|param2|..."} recovered from those red cells."""
    import openpyxl
    META = {"no.", "receiving date", "sampling date", "sample name", "sample id",
            "facility name", "municipality name", "municipality name ", "district name",
            "street name", "sample notes", "testing notes", "license number", "رقم الرخصة",
            "analysis section", "valid/ invalid", "valid/invalid", "valid /invalid",
            "matched/not matched", "matched/ not matched", "compliant / non-compliant",
            "invalid test", "non-compliant test", "no. of tests", "no of tests",
            "qc for invalid test", "result qc for invalid test", ""}
    out: dict[str, list[str]] = {}
    try:
        wb = openpyxl.load_workbook(src_path)
    except Exception:
        return {}
    for ws in wb.worksheets:
        name = ws.title.lower()
        if name.startswith("copy of") or name.startswith("invalid samples"):
            continue
        header_row = None; headers: dict[int, str] = {}
        for r in range(1, min(6, ws.max_row) + 1):
            vals = {c.column: (str(c.value).split("\n")[0].strip() if c.value else "") for c in ws[r]}
            if any(v.lower() == "sample id" for v in vals.values()):
                header_row = r; headers = vals; break
        if header_row is None:
            continue
        sid_col = next((col for col, h in headers.items() if h.lower() == "sample id"), None)
        if sid_col is None:
            continue
        for row in ws.iter_rows(min_row=header_row + 1):
            sid_cell = ws.cell(row=row[0].row, column=sid_col)
            sid = str(sid_cell.value).strip().lower() if sid_cell.value else None
            if not sid:
                continue
            for c in row:
                fl = c.fill
                fg = getattr(fl.fgColor, "rgb", None) if (fl and fl.patternType == "solid") else None
                if fg == "FFFF0000":
                    h = headers.get(c.column, "")
                    if h and h.lower() not in META:
                        out.setdefault(sid, [])
                        if h not in out[sid]:
                            out[sid].append(h)
    return {k: "|".join(v) for k, v in out.items() if v}

VALIDITY_TRUE_TOKENS  = ("مطابق", "matched", "valid", "vaild", "صالح", "compliant")
VALIDITY_FALSE_TOKENS = ("غير مطابق", "غيرمطابق", "not matched", "invalid", "invaild",
                         "غير صالح", "غيرصالح",
                         "non-compliant", "non compliant", "noncompliant")
NON_VALIDITY_TOKENS   = ("مرفوض", "غير صالح للتحليل", "ولم يتم إستلام",
                         "no specification", "raw_meat")


def map_validity(v: Any) -> tuple[bool | None, str | None]:
    """Maps source validity vocabulary to (bool, flag). FALSE tokens tested
    first since 'غير مطابق' / 'non-compliant' contain TRUE substrings."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None, None
    s = str(v).strip()
    if not s:
        return None, None
    s_norm = " ".join(s.split()).lower()
    if any(tok in s_norm for tok in VALIDITY_FALSE_TOKENS):
        return False, None
    if any(tok in s_norm for tok in VALIDITY_TRUE_TOKENS):
        return True, None
    if any(tok in s_norm for tok in NON_VALIDITY_TOKENS):
        return None, f"validity_non_verdict:{s[:40]}"
    return None, f"validity_unrecognised:{s[:40]}"


def derive_test_result(result_raw: Any, limit_raw: Any, direction: str = "max") -> dict:
    """Per-test record: numeric, is_nd, limit_numeric, exceeds_limit.

    direction='max' (default): fail when value > limit (most tests).
    direction='min': fail when value < limit (honey purity tests where the
    limit is a minimum threshold — e.g. Glucose+Fructose must be ≥ 60%)."""
    rnum, is_nd = parse_nd_to_float(result_raw)
    lnum = parse_limit_to_float(limit_raw)
    exceeds = None
    if rnum is not None and lnum is not None:
        exceeds = bool(rnum > lnum) if direction == "max" else bool(rnum < lnum)
    return {"numeric": rnum, "is_nd": is_nd, "limit_numeric": lnum, "exceeds_limit": exceeds}


def clean_section(section: str, year: int) -> tuple[int, dict]:
    schema = load_schema(section)
    applies_to = schema.get("applies_to", {})
    src_filename = applies_to.get(year) or applies_to.get(str(year))
    if not src_filename:
        print(f"[skip] section={section} year={year}: no source file in schema", file=sys.stderr)
        return 0, {}
    # Look in raw/<year>/<filename> first (preferred layout); fall back to
    # raw/<filename> (flat legacy layout) so older schemas still work.
    candidates = [RAW_DIR / str(year) / src_filename, RAW_DIR / src_filename]
    src_path = next((p for p in candidates if p.exists()), None)
    if src_path is None:
        print(f"[skip] missing source: tried {[str(p) for p in candidates]}", file=sys.stderr)
        return 0, {}

    print(f"=== {section} {year}: {src_filename} ===")
    records: list[dict] = []
    audit = {"sheets_processed": 0, "rows_in": 0, "rows_dropped_empty": 0,
             "rows_out": 0, "flags": Counter(),
             "tests_in_panel": [t["name"] for t in schema.get("tests", [])]}

    tests_def = schema.get("tests", [])
    test_keys: set[str] = set()
    for tdef in tests_def:
        test_keys.add(tdef["result"])
        if tdef.get("limit"):
            test_keys.add(tdef["limit"])
    base_keys = {
        "receiving_date", "sampling_date", "testing_date",
        "sample_id", "sample_name", "sample_category",
        "facility_name", "district_name", "street_name", "municipality",
        "license_number", "analysis_section",
        "validity_raw", "invalid_test", "sample_notes", "testing_notes",
        # Section-specific (pesticides long format) — handled by their own block.
        "pesticide_name", "concentration_ppm", "limit_ppm",
        "specification", "analyzer", "num_pesticides", "sample_code",
        "qc_pesticide",
    }
    schema_cols = list(schema.get("columns", {}).keys())
    # paired_columns.qc_key fields are mapped by the iter_data_sheets paired-
    # column logic but aren't listed under columns:. They should still be
    # passed through to the parquet (e.g. silver_qc, aluminum_qc — the per-
    # metal QC.RC% values).
    qc_keys = []
    for pair in schema.get("paired_columns", []):
        if pair.get("qc_key"):
            qc_keys.append(pair["qc_key"])
    passthrough_keys = [k for k in schema_cols if k not in test_keys and k not in base_keys] + qc_keys

    # For single-sheet sections (e.g. honey), the sheet name doesn't encode the
    # month — each row's own sampling_date is the only correct source. Detect this
    # so we can derive sheet_year_month from the row's date instead of the synthetic
    # (year, 1) bucket the iterator hands us.
    is_single_sheet = bool(schema.get("single_sheet", False))

    # M1: water failed-tests have no numeric basis (limits live in GSO ref), so
    # recover them from the source instead. 2024 → red cell-fills; 2025 → the
    # lab-entered `invalid_test` column (surfaced per-row below).
    water_redcells: dict[str, str] = {}
    if section == "water_analysis" and int(year) == 2024:
        water_redcells = water_redcell_failed_map(src_path)
        print(f"  water red-cell failed-test map: {len(water_redcells)} samples")

    for sheet_name, ym, rows in iter_data_sheets(src_path, year, schema):
        audit["sheets_processed"] += 1
        s_year, s_month = ym

        # Pesticides long-format forward-fill: lab writes sample metadata only on
        # the first detected-pesticide row per sample; continuation rows are blank.
        if section == "pesticides":
            ffill_keys = ("sample_id", "sample_name", "sample_category",
                          "sample_code", "facility_name", "district_name",
                          "street_name", "municipality", "license_number",
                          "analysis_section", "receiving_date", "sampling_date",
                          "testing_date", "validity_raw", "invalid_test",
                          "num_pesticides", "sample_notes")
            last_vals = {k: None for k in ffill_keys}
            for raw_row in rows:
                if raw_row.get("sample_id") or raw_row.get("sample_name"):
                    for k in ffill_keys:
                        if raw_row.get(k) is not None:
                            last_vals[k] = raw_row[k]
                else:
                    for k in ffill_keys:
                        if raw_row.get(k) is None and last_vals[k] is not None:
                            raw_row[k] = last_vals[k]

        for raw_row in rows:
            audit["rows_in"] += 1
            recv = parse_date(raw_row.get("receiving_date"))
            samp = parse_date(raw_row.get("sampling_date"))
            test = parse_date(raw_row.get("testing_date"))
            sampling = samp or recv or date(s_year, s_month, 1)

            # Year purity by date: drop rows where the resolved sampling date
            # is in a FUTURE year vs the file's declared year. Catches stray
            # 2026 rows in 2025 single-sheet sections (honey) where the
            # sheet-name year filter can't help. Past-year typos (e.g. an
            # 0204-03-18 entry that's clearly meant to be 2024) are kept.
            if sampling.year > year:
                audit["rows_dropped_empty"] += 1
                audit["flags"][f"date_after_file_year:{sampling.year}"] += 1
                continue

            sample_id_raw = raw_row.get("sample_id")
            sample_id_norm = normalise_sample_id(sample_id_raw) if sample_id_raw else None
            sample_name = normalise_text(raw_row.get("sample_name"))
            facility = normalise_text(raw_row.get("facility_name"))
            municipality_raw = normalise_text(raw_row.get("municipality"))

            # Strict row-drop — no sample identity = not a real sample row.
            if sample_id_norm is None and sample_name is None:
                audit["rows_dropped_empty"] += 1
                continue

            # Header-leak drop: lab sometimes copy-pasted the header row into
            # the data range (water_analysis June-25 had 10 such rows). Drop
            # any row where the sample_id literally contains the header text.
            sid_lc = (sample_id_norm or "").lower()
            if "sample id" in sid_lc or "رمز العينة" in sid_lc:
                audit["rows_dropped_empty"] += 1
                audit["flags"]["header_leak_dropped"] += 1
                continue

            v_bool, v_flag = map_validity(raw_row.get("validity_raw"))
            if v_flag:
                audit["flags"][v_flag] += 1

            rec = {
                "source_file":      src_filename,
                "year":             year,
                "sheet_name":       sheet_name,
                # For single-sheet sections, use the per-row sampling date
                # to build the month bucket. For monthly sheets, use the sheet's name.
                "sheet_year_month": (f"{sampling.year:04d}-{sampling.month:02d}"
                                       if is_single_sheet
                                       else f"{s_year:04d}-{s_month:02d}"),
                "sampling_date":    sampling,
                "receiving_date":   recv,
                "testing_date":     test,
                "sample_id_raw":    str(sample_id_raw) if sample_id_raw else None,
                "sample_id":        sample_id_norm,
                "sample_name":      sample_name,
                "sample_category":  normalise_text(raw_row.get("sample_category")),
                "facility_name":    facility,
                "district_name":    normalise_text(raw_row.get("district_name")),
                "street_name":      normalise_text(raw_row.get("street_name")),
                "municipality":     municipality_raw,
                "license_number":   normalise_text(raw_row.get("license_number")),
                "analysis_section": normalise_text(raw_row.get("analysis_section")),
                "validity_raw":     normalise_text(raw_row.get("validity_raw")),
                "is_valid":         v_bool,
                "invalid_test":     normalise_text(raw_row.get("invalid_test")),
                "sample_notes":     normalise_text(raw_row.get("sample_notes")),
                "testing_notes":    normalise_text(raw_row.get("testing_notes")),
            }

            # Per-test result columns.
            n_failed = 0
            failed_tests: list[str] = []
            for tdef in tests_def:
                r_key = tdef["result"]
                l_key = tdef.get("limit")
                if tdef.get("kind") == "string":
                    rec[f"{r_key}_text"] = normalise_text(raw_row.get(r_key))
                    continue
                derived = derive_test_result(
                    raw_row.get(r_key),
                    raw_row.get(l_key) if l_key else None,
                    direction=tdef.get("direction", "max"),
                )
                rec[f"{r_key}_value"] = derived["numeric"]
                rec[f"{r_key}_nd"]    = derived["is_nd"]
                if l_key:
                    rec[f"{l_key}_value"] = derived["limit_numeric"]
                if derived["exceeds_limit"] is True:
                    n_failed += 1
                    failed_tests.append(tdef["name"])
            if tests_def:
                rec["n_failed_tests_derived"] = n_failed
                rec["failed_tests_derived"]   = "|".join(failed_tests) if failed_tests else None

                # Fallback validity: when the lab forgot to fill validity_raw
                # (e.g. aflatoxins Jan/Feb 2024 — 219 rows blank), derive from
                # exceeded-limit count. We require that at least one test on
                # the row had both a result AND a limit, so we don't mark
                # rows with no numeric basis. Lab's explicit verdict wins
                # whenever they entered one.
                if rec["is_valid"] is None:
                    has_basis = any(
                        rec.get(f"{t['result']}_value") is not None
                        and t.get("limit")
                        and rec.get(f"{t['limit']}_value") is not None
                        for t in tests_def if t.get("kind") != "string"
                    )
                    if has_basis:
                        rec["is_valid"] = (n_failed == 0)
                        audit["flags"]["is_valid_derived_from_limits"] += 1

                # Second fallback: per-test Arabic verdicts in string-kind
                # columns (e.g. food_chemistry sensory_*_text say "مطابق" /
                # "غير مطابق"). Catches rows like food_chemistry Aug-2024 &
                # Feb-2025 where only sensory tests were run and no numeric
                # limits exist on the row.
                if rec["is_valid"] is None:
                    text_vals = [
                        str(rec.get(f"{t['result']}_text"))
                        for t in tests_def if t.get("kind") == "string"
                        and rec.get(f"{t['result']}_text") is not None
                    ]
                    if text_vals:
                        # "غير" is the Arabic negation in "غير مطابق" (non-
                        # compliant). "مطايق" is a real-data typo of مطابق.
                        if any("غير" in v for v in text_vals):
                            rec["is_valid"] = False
                            audit["flags"]["is_valid_derived_from_verdict_text"] += 1
                        elif any(("مطابق" in v) or ("مطايق" in v) for v in text_vals):
                            rec["is_valid"] = True
                            audit["flags"]["is_valid_derived_from_verdict_text"] += 1

            # Pesticides long-format pass-through.
            if section == "pesticides":
                p_conc, p_nd = parse_nd_to_float(raw_row.get("concentration_ppm"))
                p_lim = parse_limit_to_float(raw_row.get("limit_ppm"))
                rec["pesticide_name"]      = normalise_text(raw_row.get("pesticide_name"))
                rec["concentration_ppm"]   = p_conc
                rec["concentration_is_nd"] = p_nd
                rec["limit_ppm"]           = p_lim
                rec["exceeds_limit"] = (
                    bool(p_conc > p_lim) if (p_conc is not None and p_lim is not None) else None
                )
                rec["specification"]  = normalise_text(raw_row.get("specification"))
                rec["analyzer"]       = normalise_text(raw_row.get("analyzer"))
                rec["num_pesticides"] = parse_limit_to_float(raw_row.get("num_pesticides"))
                # Pesticide row-level fallback: long format has one row per
                # pesticide. When lab left validity blank but we have conc &
                # limit, derive: exceeds → invalid, else → valid.
                if rec["is_valid"] is None and p_conc is not None and p_lim is not None:
                    rec["is_valid"] = not (p_conc > p_lim)
                    audit["flags"]["is_valid_derived_from_limits"] += 1

            # Classify *why* a row's verdict is what it is, so the dashboard
            # can split the "unknown" bucket into actionable sub-categories
            # ("no limit available", "lab rejected sample") instead of
            # lumping them with truly empty rows.
            if rec["is_valid"] is True:
                rec["validity_status"] = "valid"
            elif rec["is_valid"] is False:
                rec["validity_status"] = "invalid"
            else:
                vr = (rec.get("validity_raw") or "").strip()
                vr_norm = " ".join(vr.split()).lower()
                # Arabic rejection tokens: مرفوض / مرفوضة → lab rejected
                # English markers we've seen: 'no specifications', 'no specs'
                if any(tok in vr_norm for tok in ("مرفوض", "rejected", "raw_meat")):
                    rec["validity_status"] = "rejected"
                elif any(tok in vr_norm for tok in ("no specification", "no spec")):
                    rec["validity_status"] = "no_limit"
                else:
                    # Check if any test result was recorded without a limit
                    # (heavy_metals Aug 2024 fish samples, pesticides with
                    # no limit_ppm).
                    has_value_no_limit = False
                    if section == "pesticides":
                        has_value_no_limit = (rec.get("concentration_ppm") is not None
                                              and rec.get("limit_ppm") is None)
                    else:
                        has_value_no_limit = any(
                            rec.get(f"{t['result']}_value") is not None
                            and (not t.get("limit")
                                 or rec.get(f"{t['limit']}_value") is None)
                            for t in (tests_def or []) if t.get("kind") != "string"
                        )
                    rec["validity_status"] = "no_limit" if has_value_no_limit else "unknown"
                rec["sample_code"]    = normalise_text(raw_row.get("sample_code"))

            # Canonical sample-category + name grouping (2026-07-01 spec).
            canon, cat_flag = categories.classify(
                section, rec.get("sample_category"), rec.get("sample_name"))
            rec["sample_category_canonical"] = canon
            rec["category_flag"] = cat_flag
            grp = categories.name_group(rec.get("sample_name"))
            rec["sample_name_group"] = grp or rec.get("sample_name")

            # M1: surface water failed-tests that have no numeric basis.
            if section == "water_analysis" and not rec.get("failed_tests_derived"):
                recovered = None
                if int(year) == 2024:
                    sid = rec.get("sample_id")
                    recovered = water_redcells.get(sid.lower()) if sid else None
                else:  # 2025 — lab-entered invalid_test column
                    it = rec.get("invalid_test")
                    if it and str(it).strip():
                        recovered = "|".join(p.strip() for p in re.split(r"[,،]", str(it)) if p.strip())
                if recovered:
                    rec["failed_tests_derived"] = recovered
                    rec["n_failed_tests_derived"] = len(recovered.split("|"))

            records.append(rec)
            audit["rows_out"] += 1

    # Intra-sheet deduplication — drop rows that repeat the same logical
    # sample within the same sheet. Source xlsx have known copy-paste issues
    # (e.g. water_analysis June-25 had 234 duplicate rows). The dedup key is
    # section-specific:
    #   - pesticides (long-format): (sheet, sample_id, pesticide_name, concentration_ppm)
    #   - everything else: (sheet, sample_id)
    # Rows without a sample_id (rare after the strict drop above) are kept.
    if records:
        if section == "pesticides":
            key_fields = ("sheet_name", "sample_id", "pesticide_name", "concentration_ppm")
        else:
            key_fields = ("sheet_name", "sample_id")
        seen: set[tuple] = set()
        kept: list[dict] = []
        n_dropped = 0
        for rec in records:
            # Keep rows that have no sample_id (shouldn't happen after strict drop)
            if rec.get("sample_id") is None:
                kept.append(rec); continue
            key = tuple(rec.get(k) for k in key_fields)
            if key in seen:
                n_dropped += 1
                continue
            seen.add(key)
            kept.append(rec)
        if n_dropped:
            audit["flags"][f"intra_sheet_duplicates_dropped"] += n_dropped
            audit["rows_out"] = len(kept)
        records = kept

    out_path = CLEAN_DIR / f"chem_{section}_{year}.parquet"
    string_cols = [
        "source_file", "sheet_name", "sheet_year_month",
        "sample_id_raw", "sample_id", "sample_name", "sample_category",
        "sample_category_canonical", "category_flag", "sample_name_group",
        "facility_name", "district_name", "street_name", "municipality",
        "license_number", "analysis_section", "validity_raw", "invalid_test",
        "sample_notes", "testing_notes", "failed_tests_derived",
        "pesticide_name", "specification", "analyzer", "sample_code",
    ]
    for tdef in tests_def:
        if tdef.get("kind") == "string":
            string_cols.append(f"{tdef['result']}_text")
    string_cols.extend(passthrough_keys)
    n = write_parquet(records, out_path, string_cols=string_cols)
    print(f"  wrote {out_path}  ({n} rows; sheets={audit['sheets_processed']}, dropped={audit['rows_dropped_empty']})")
    if audit["flags"]:
        print(f"  flags: {dict(audit['flags'].most_common(5))}")

    report_path = REPORT_DIR / f"chem_{section}_{year}.md"
    report_path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        f"# {section} · {year}",
        f"- source: `{src_filename}`",
        f"- sheets processed: {audit['sheets_processed']}",
        f"- rows in: {audit['rows_in']}",
        f"- rows dropped (no sample identity): {audit['rows_dropped_empty']}",
        f"- rows out: {audit['rows_out']}",
        f"- tests in panel: {len(audit['tests_in_panel'])}",
    ]
    if audit["flags"]:
        lines.append("\n## Flags raised")
        for flag, cnt in audit["flags"].most_common():
            lines.append(f"- `{flag}`: {cnt}")
    report_path.write_text("\n".join(lines), encoding="utf-8")
    return n, audit


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--section", required=True,
                    choices=["aflatoxins", "food_chemistry", "heavy_metals", "honey",
                             "hormones_antibiotics", "pesticides", "water_analysis",
                             "all"])
    ap.add_argument("--year", type=int, default=None)
    args = ap.parse_args()

    sections = [args.section] if args.section != "all" else [
        "aflatoxins", "food_chemistry", "heavy_metals", "honey",
        "hormones_antibiotics", "pesticides", "water_analysis",
    ]
    grand_total = 0
    for sec in sections:
        try:
            schema = load_schema(sec)
        except FileNotFoundError:
            print(f"[skip] no schema for section={sec}", file=sys.stderr)
            continue
        years = [args.year] if args.year else list(schema.get("applies_to", {}).keys())
        years = [int(y) for y in years]
        for y in sorted(years):
            n, _ = clean_section(sec, y)
            grand_total += n
    print(f"\n=== grand total: {grand_total} rows across all (section, year) pairs ===")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
