"""Build an Excel verification workbook to check classification + find gaps vs
the official annual numbers. Sheets:
  • Numbers        — our totals vs official, Δ (numbers only)
  • Classification — distinct (year · raw category · bucket · GSO · source) + counts
  • Samples        — every sample with its classification + missing-info flags
  • Missing_info   — only the samples with a flag (the ones to check)

Run:  microbiology/.venv/bin/python microbiology/scripts/build_verification_workbook.py
Out:  microbiology/reports/annual_verification_check.xlsx
"""
from __future__ import annotations
from collections import defaultdict
from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

from build_classification_table import classify, _val

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "reports" / "annual_verification_check.xlsx"

# Official MICRO totals (samples, compliant) per year — from Muhannad.
OFF = {2024: (9108, 6399), 2025: (11404, 8345)}
OFF_TESTS = {2024: (None, None), 2025: (46309, 4211)}
MNUM = {f"{i:02d}": m for i, m in enumerate(
    ["", "Jan", "Feb", "Mar", "Apr", "May", "June", "July", "Aug", "Sep", "Oct", "Nov", "Dec"])}


def flags_for(r, gso, src):
    f = []
    if _val(r.get("sample_id")) is None:       f.append("no_sample_id")
    if _val(r.get("sample_name")) is None:     f.append("no_sample_name")
    if _val(r.get("category_canonical")) is None and _val(r.get("gso_category_name_en")) is None:
        f.append("no_category")
    if pd.isna(r.get("is_valid")):             f.append("unknown_validity")
    if gso == "Miscellaneous Foods":           f.append("miscellaneous")
    if src == "name-keyword":                  f.append("name_guess")
    if r.get("gso_panel_complete") is False:   f.append("incomplete_panel")
    return f


def build():
    # ---- load + per-sample classification + flags ----
    sample_rows = []
    groups = defaultdict(lambda: {"n": 0, "ex": set()})
    for y in (2024, 2025):
        d = pd.read_parquet(ROOT / "cleaned" / f"data{y}.parquet")
        d["__month"] = d["year_month"].astype(str).str[-2:].map(MNUM)
        for r in d.to_dict("records"):
            gso, src = classify(r)
            raw = _val(r.get("category_canonical")) or _val(r.get("gso_category_name_en")) or "—"
            fl = flags_for(r, gso, src)
            valid = r.get("is_valid")
            sample_rows.append({
                "sample_id": _val(r.get("sample_id")) or "",
                "year": y, "month": r["__month"],
                "raw_category": str(raw),
                "sample_name": str(_val(r.get("sample_name")) or ""),
                "sample_type_bucket": str(_val(r.get("sample_type")) or "—"),
                "gso_category": gso, "source": src,
                "valid": ("" if pd.isna(valid) else bool(valid)),
                "flags": ", ".join(fl),
            })
            g = groups[(y, str(raw), str(_val(r.get("sample_type")) or "—"), gso, src)]
            g["n"] += 1
            nm = _val(r.get("sample_name"))
            if nm and len(g["ex"]) < 4:
                g["ex"].add(str(nm))

    samples = pd.DataFrame(sample_rows)
    missing = samples[samples["flags"] != ""].copy()

    cls = pd.DataFrame(
        [{"year": k[0], "raw_category": k[1], "sample_type_bucket": k[2],
          "gso_category": k[3], "source": k[4], "N": v["n"],
          "examples": " · ".join(sorted(v["ex"]))}
         for k, v in groups.items()]
    ).sort_values(["gso_category", "N"], ascending=[True, False])

    # ---- Numbers sheet ----
    num = []
    for y in (2024, 2025):
        d = samples[samples["year"] == y]
        os_, oc = OFF[y]
        rs = len(d); rc = int((d["valid"] == True).sum())  # noqa: E712
        num += [
            {"year": y, "metric": "Samples", "official": os_, "our": rs, "delta": rs - os_},
            {"year": y, "metric": "Compliant", "official": oc, "our": rc, "delta": rc - oc},
            {"year": y, "metric": "Non-compliant", "official": os_ - oc, "our": rs - rc, "delta": (rs - rc) - (os_ - oc)},
            {"year": y, "metric": "Compliance %", "official": round(100 * oc / os_, 2),
             "our": round(100 * rc / rs, 2), "delta": round(100 * rc / rs - 100 * oc / os_, 2)},
            {"year": y, "metric": "Total tests", "official": OFF_TESTS[y][0], "our": None, "delta": None},
            {"year": y, "metric": "Non-comp tests", "official": OFF_TESTS[y][1], "our": None, "delta": None},
        ]
    numbers = pd.DataFrame(num)

    # ---- flag summary appended to Numbers ----
    flag_counts = defaultdict(int)
    for fl in samples["flags"]:
        for f in (fl.split(", ") if fl else []):
            flag_counts[f] += 1

    # ---- write ----
    with pd.ExcelWriter(OUT, engine="openpyxl") as xl:
        numbers.to_excel(xl, sheet_name="Numbers", index=False, startrow=0)
        # flag summary below the numbers table
        fs = pd.DataFrame(sorted(flag_counts.items(), key=lambda x: -x[1]),
                          columns=["missing/flag", "samples"])
        fs.to_excel(xl, sheet_name="Numbers", index=False, startrow=len(numbers) + 3)
        cls.to_excel(xl, sheet_name="Classification", index=False)
        samples.to_excel(xl, sheet_name="Samples", index=False)
        missing.to_excel(xl, sheet_name="Missing_info", index=False)

        wb = xl.book
        hdr = PatternFill("solid", fgColor="1C2742")
        soft = PatternFill("solid", fgColor="FFF3C7")
        for name in wb.sheetnames:
            ws = wb[name]
            for c in ws[1]:
                if c.value is not None:
                    c.fill = hdr; c.font = Font(color="FFFFFF", bold=True)
                    c.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "A2"
            # widen a few columns
            for col, w in {"A": 16, "B": 30, "C": 18, "D": 34, "E": 20, "F": 10, "G": 40, "H": 40, "I": 12, "J": 26}.items():
                ws.column_dimensions[col].width = w
        # highlight the flag summary header on Numbers
        ws = wb["Numbers"]
        r0 = len(numbers) + 4
        ws.cell(row=r0, column=1).fill = soft
        ws.cell(row=r0, column=2).fill = soft

    print(f"wrote {OUT}")
    print(f"  samples={len(samples)}  classification-groups={len(cls)}  flagged(missing-info)={len(missing)}")
    print("  flag summary:", dict(sorted(flag_counts.items(), key=lambda x: -x[1])))


if __name__ == "__main__":
    build()
