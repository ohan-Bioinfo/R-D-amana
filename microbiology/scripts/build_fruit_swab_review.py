"""Fillable verification workbook scoped to two categories Muhannad wants to
check: Fruit & Vegetables and Environmental Swabs. Each distinct sample_name is
listed once with its current GSO classification, how it was derived, per-year
counts and a compliance snapshot, plus a yellow corrected_gso_category (dropdown)
+ notes column. Send it back and I apply the corrections.

Run:  microbiology/.venv/bin/python microbiology/scripts/build_fruit_swab_review.py
Out:  microbiology/reports/fruit_swab_review_to_fill.xlsx
"""
from __future__ import annotations
from collections import defaultdict
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

from build_classification_table import classify, _val

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "reports" / "fruit_swab_review_to_fill.xlsx"

TARGETS = {
    "Fruit_and_Veg": "Fruit and Vegetables",
    "Swabs":         "Environmental Swabs",
}

HDR = PatternFill("solid", fgColor="1C2742")
YEL = PatternFill("solid", fgColor="F59E0B")
YELCELL = PatternFill("solid", fgColor="FFF3C7")
WBF = Font(color="FFFFFF", bold=True)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build():
    valid_gso = set()
    # per target category: name -> {src, n, years, nc, cats}
    buckets = {k: defaultdict(lambda: {"src": None, "n": 0, "years": set(),
                                       "nc": 0, "cats": set()}) for k in TARGETS}
    cat_to_key = {v: k for k, v in TARGETS.items()}
    for y in (2024, 2025):
        d = pd.read_parquet(ROOT / "cleaned" / f"data{y}.parquet")
        for r in d.to_dict("records"):
            gso, src = classify(r)
            valid_gso.add(gso)
            if gso not in cat_to_key:
                continue
            g = buckets[cat_to_key[gso]][str(_val(r.get("sample_name")) or "(no name)")]
            g["src"] = src
            g["n"] += 1
            g["years"].add(y)
            if r.get("is_failure") is True:
                g["nc"] += 1
            raw = _val(r.get("category_canonical")) or _val(r.get("gso_category_name_en"))
            if raw:
                g["cats"].add(str(raw))

    wb = Workbook()
    first = True
    dv_source = f"Valid_GSO_categories!$A$2:$A${len(valid_gso) + 1}"

    for key, cat in TARGETS.items():
        ws = wb.active if first else wb.create_sheet(key)
        ws.title = key
        first = False
        ws.append(["current_gso", "sample_name", "raw_category", "source",
                   "samples", "non_compliant", "nc_rate_%", "year(s)",
                   "corrected_gso_category", "notes"])
        rows = sorted(buckets[key].items(), key=lambda kv: -kv[1]["n"])
        for nm, g in rows:
            rate = round(100 * g["nc"] / g["n"], 1) if g["n"] else 0
            ws.append([cat, nm, " · ".join(sorted(g["cats"])[:2]) or "—", g["src"],
                       g["n"], g["nc"], rate,
                       "/".join(str(x) for x in sorted(g["years"])), "", ""])
        # styling + dropdown on corrected column (I = 9th, 1-indexed)
        for c in ws[1]:
            c.fill = HDR; c.font = WBF; c.alignment = CEN
        ws.cell(row=1, column=9).fill = YEL
        ws.cell(row=1, column=10).fill = YEL
        for i in range(2, len(rows) + 2):
            ws.cell(row=i, column=9).fill = YELCELL
            ws.cell(row=i, column=10).fill = YELCELL
        dv = DataValidation(type="list", formula1=dv_source, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"I2:I{len(rows) + 1}")
        ws.freeze_panes = "A2"
        for col, w in {"A": 22, "B": 32, "C": 26, "D": 20, "E": 9, "F": 12,
                       "G": 9, "H": 9, "I": 34, "J": 26}.items():
            ws.column_dimensions[col].width = w

    vg = wb.create_sheet("Valid_GSO_categories")
    vg.append(["valid_gso_category"])
    for g in sorted(valid_gso):
        vg.append([g])
    vg.cell(row=1, column=1).fill = HDR
    vg.cell(row=1, column=1).font = WBF
    vg.column_dimensions["A"].width = 40

    wb.save(OUT)
    print(f"wrote {OUT}")
    for key, cat in TARGETS.items():
        b = buckets[key]
        print(f"  {key}: {len(b)} distinct names, {sum(g['n'] for g in b.values())} samples")


if __name__ == "__main__":
    build()
