"""Phase-4 cleaner — applies schemas/lab_data_2024_v2.yaml to every .xlsx file
under a yearly source dir and emits:

    cleaned/data<YEAR>_long.parquet   one row per (sample, test)
    cleaned/data<YEAR>.parquet        one row per sample (aligned with 2025)
    reports/data<YEAR>_clean_report.md run summary + per-file audit

Run:
    .venv/bin/python scripts/clean_2024.py              # default: year=2024
    .venv/bin/python scripts/clean_2024.py --year 2023  # cleans 2023
"""
from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

import pandas as pd
import yaml
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SCHEMA_PATH = ROOT / "schemas" / "lab_data_2024_v2.yaml"

# Per-year config. INPUT_ROOT picks up the variant directory layout the user
# happened to organise each year's files under.
YEAR_CONFIG: dict[int, dict] = {
    2024: {"input_root": ROOT / "2024-original" / "2024"},
    2023: {"input_root": ROOT / "2023"},
}

# Module-level state set by main() from CLI args. The helpers (clean_one_file,
# write_report, etc.) read these via globals so we don't need to thread `year`
# through every function signature.
YEAR = 2024
INPUT_ROOT = YEAR_CONFIG[YEAR]["input_root"]
OUT_LONG = ROOT / "cleaned" / f"data{YEAR}_long.parquet"
OUT_WIDE = ROOT / "cleaned" / f"data{YEAR}.parquet"
REPORT_PATH = ROOT / "reports" / f"data{YEAR}_clean_report.md"

ZERO_WIDTH = "​‌‍﻿"
RTL_MARKS = "‎‏‪‫‬‭‮⁦⁧⁨⁩"
NBSP = " "
ARABIC_INDIC_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")
EASTERN_ARABIC_DIGITS = str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789")

DATE_REGEX = re.compile(r"يوم\s+\S+\s*(\d{1,2})/(\d{1,2})/(\d{4})")

FILENAME_PATTERNS = [
    re.compile(r"^(?P<dd>\d{2})(?P<mm>\d{2})(?P<yy>\d{2})[a-z]?\.xlsx$"),
    re.compile(r"^(?P<dd>\d{2})(?P<mm>\d{2})(?P<yyyy>\d{4})\.xlsx$"),
    re.compile(r"^(?P<dd>\d{2})(?P<mm>\d{2})(?P<yy>\d{2})\s*-+\s*\d*\s*-?\.xlsx$"),
    re.compile(r"^Results\s+(?P<dd>\d{2})(?P<mm>\d{2})(?P<yy>\d{2})(\s*\(\d+\))?\.xlsx$"),  # 2023 Aug uses 'Results NNNNNN.xlsx' (+ optional " (N)" suffix)
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def load_schema() -> dict:
    with SCHEMA_PATH.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def normalise_text(v):
    if v is None:
        return None
    if not isinstance(v, str):
        return v
    s = unicodedata.normalize("NFKC", v)
    for ch in ZERO_WIDTH + RTL_MARKS:
        s = s.replace(ch, "")
    s = s.replace(NBSP, " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s if s != "" else None


def col_letter_to_idx(letter: str) -> int:
    out = 0
    for ch in letter.upper():
        out = out * 26 + (ord(ch) - ord("A") + 1)
    return out


def parse_date_from_filename(name: str) -> date | None:
    for pat in FILENAME_PATTERNS:
        m = pat.match(name)
        if not m:
            continue
        g = m.groupdict()
        try:
            if "yyyy" in g and g.get("yyyy"):
                return date(int(g["yyyy"]), int(g["mm"]), int(g["dd"]))
            return date(2000 + int(g["yy"]), int(g["mm"]), int(g["dd"]))
        except ValueError:
            return None
    return None


def parse_date_from_cell(cell_value):
    if cell_value is None:
        return None
    if isinstance(cell_value, datetime):
        return cell_value.date()
    if isinstance(cell_value, date):
        return cell_value
    if not isinstance(cell_value, str):
        return None
    m = DATE_REGEX.search(cell_value)
    if not m:
        m2 = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", cell_value)
        if not m2:
            return None
        dd, mm, yyyy = m2.groups()
    else:
        dd, mm, yyyy = m.groups()
    try:
        return date(int(yyyy), int(mm), int(dd))
    except ValueError:
        return None


def parse_numeric_best_effort(v) -> tuple[float | None, list[str]]:
    flags: list[str] = []
    if v is None:
        return None, flags
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v), flags
    if not isinstance(v, str):
        return None, flags
    s = v.strip()
    if s == "":
        return None, flags
    s = s.translate(ARABIC_INDIC_DIGITS).translate(EASTERN_ARABIC_DIGITS)
    s = s.replace("٬", "").replace("٫", ".").replace(",", "")
    for prefix in ("<=", ">=", "≤", "≥", "<", ">"):
        if s.startswith(prefix):
            flags.append(f"comparison_prefix:{prefix}")
            s = s[len(prefix):].strip()
            break
    s = re.sub(r"[^\d.\-eE]", "", s)
    if s in ("", "-", ".", "-."):
        return None, flags
    try:
        return float(s), flags
    except ValueError:
        return None, flags


_FR_NORMALISE = [
    # taa marbouta ↔ ha — common interchange in the source text
    ("ة", "ه"),
    # double-letter doubling typos seen in 2024 reports
    ("العدد", "العد"),
    ("الانتيروباكتيرييسي", "انتيروباكتريسي"),
    ("الانتيروباكتيريسي", "انتيروباكتريسي"),
    ("انتيروباكتيرييسي", "انتيروباكتريسي"),
    ("انتيروباكتيريسي", "انتيروباكتريسي"),
    ("الانتيرو", "انتيرو"),
    # spaced compound test names
    ("استافيلو كوكس", "استافيلوكوكس"),
    ("الاستافيللوكوكس", "استافيلوكوكس"),
    ("الاستافيلوكوكس", "استافيلوكوكس"),
    ("الايشيريشيا", "ايشيريشيا"),
    ("الايشيريا", "ايشيريشيا"),
    ("الكوليفورم", "كوليفورم"),
    ("الباسيللس سيرس", "باصلص سيرز"),
    ("الباسيللس سيريس", "باصلص سيرز"),
    ("الباسيللس سيرز", "باصلص سيرز"),
    ("الباسيلس", "باصلص"),
    ("الأعفان", "اعفان"),
    ("الاعفان", "اعفان"),
    ("الخمائر", "خمائر"),
    ("والاعفان", " اعفان"),
    ("والأعفان", " اعفان"),
    ("بكتريا", "بكتيريا"),
    # water variants
    ("المياة", "المياه"),
]


def _norm_for_match(s: str) -> str:
    for a, b in _FR_NORMALISE:
        s = s.replace(a, b)
    return s


def match_final_report_code(text: str | None, code_map: dict) -> str | None:
    if text is None:
        return None
    norm = _norm_for_match(text)
    for code, rule in code_map.items():
        mt = rule.get("match_type")
        if mt == "fallback":
            continue
        needles = [_norm_for_match(n) for n in rule.get("needles", [])]
        not_contains = [_norm_for_match(n) for n in rule.get("not_contains", [])]
        if mt == "contains":
            if any(n in norm for n in needles) and not any(n in norm for n in not_contains):
                return code
        elif mt == "contains_all":
            if all(n in norm for n in needles) and not any(n in norm for n in not_contains):
                return code
    return next((c for c, r in code_map.items() if r.get("match_type") == "fallback"), None)


# ---------------------------------------------------------------------------
# Header & column detection
# ---------------------------------------------------------------------------
def detect_header_row(ws, marker: str, scan_range: tuple[int, int]) -> int | None:
    lo, hi = scan_range
    hi = min(hi, ws.max_row or hi)
    for r in range(lo, hi + 1):
        for c in range(1, (ws.max_column or 0) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and marker in v:
                return r
    return None


def build_column_map(grid: list[list], header_row: int, schema: dict) -> tuple[dict[str, int], list[str]]:
    """Return {logical_name: col_idx_1based}. Subheader-driven confirmatory
    columns are added with logical names confirm_vitek / confirm_vidas / confirm_rt_pcr.
    Also returns a list of warnings for missing optional columns / sub-headers.
    """
    warnings: list[str] = []
    headers = grid[header_row]
    sub_row = grid[header_row + 1] if header_row + 1 < len(grid) else []
    norm_headers = [normalise_text(h) if isinstance(h, str) else h for h in headers]
    norm_sub = [normalise_text(h) if isinstance(h, str) else h for h in sub_row]

    col_map: dict[str, int] = {}
    for spec in schema["column_map"]:
        logical = spec["logical"]
        aliases = spec.get("aliases", [])
        if logical == "confirmatory_block":
            found_block_col = None
            for i, h in enumerate(norm_headers):
                if h in aliases:
                    found_block_col = i
                    break
            if found_block_col is None:
                continue
            for sub in spec.get("sub_headers", []):
                sub_logical = sub["logical"]
                sub_aliases = sub["aliases"]
                # Sub-headers span the block; search the sub_row near the block.
                for j in range(max(1, found_block_col - 1), min(len(norm_sub), found_block_col + 4)):
                    if norm_sub[j] in sub_aliases:
                        col_map[sub_logical] = j
                        break
                else:
                    warnings.append(f"confirmatory_sub_header_not_found:{sub_logical}")
            continue
        for i, h in enumerate(norm_headers):
            if h in aliases:
                col_map[logical] = i
                break
        else:
            if spec.get("required", False):
                warnings.append(f"required_column_missing:{logical}")
    return col_map, warnings


# ---------------------------------------------------------------------------
# Workbook reading with merge-fill
# ---------------------------------------------------------------------------
def read_grid_with_merge_fill(path: Path, header_row: int):
    """Returns (grid, ws_title). grid is 1-indexed list[row][col]."""
    wb = load_workbook(path, data_only=True, read_only=False, keep_links=False)
    ws = wb.active
    title = ws.title
    n_rows = ws.max_row or 0
    n_cols = ws.max_column or 0
    if n_rows == 0 or n_cols == 0:
        wb.close()
        return None, title

    grid: list[list] = [[None] * (n_cols + 1) for _ in range(n_rows + 1)]
    for row in ws.iter_rows(min_row=1, max_row=n_rows, max_col=n_cols, values_only=False):
        for cell in row:
            grid[cell.row][cell.column] = cell.value

    # Materialise merges strictly below the header row.
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row <= header_row:
            continue
        v = grid[mr.min_row][mr.min_col]
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                grid[r][c] = v
    wb.close()
    return grid, title


def find_date_in_banner(grid: list[list], header_row: int) -> date | None:
    e_col = col_letter_to_idx("E")
    for r in range(1, header_row):
        v = grid[r][e_col] if e_col < len(grid[r]) else None
        d = parse_date_from_cell(v)
        if d is not None:
            return d
    # Fallback: search all cells above header for a date string.
    for r in range(1, header_row):
        for c in range(1, len(grid[r])):
            d = parse_date_from_cell(grid[r][c])
            if d is not None:
                return d
    return None


# ---------------------------------------------------------------------------
# Per-file cleaner — returns long-format rows + audit
# ---------------------------------------------------------------------------
def clean_one_file(path: Path, schema: dict) -> tuple[pd.DataFrame | None, dict]:
    audit = {
        "source_file": path.name,
        "month_folder": path.parent.name,
        "header_row": None,
        "sheet_date": None,
        "sheet_name": None,
        "rows_in": 0,
        "rows_out": 0,
        "flags_count": 0,
        "changes_count": 0,
        "drops_count": 0,
        "skip_reason": None,
        "warnings": [],
    }

    try:
        # Probe workbook for emptiness first. NOTE: opened WITHOUT read_only —
        # ~20 Apr/Aug files omit the workbook <dimension> tag, so read_only
        # reports ws.max_row = None, which the old check mis-read as "empty" and
        # silently skipped ~1,287 real samples. Normal mode reports the true
        # dimension. (2026-07-16) Genuinely-empty files still fail header
        # detection below and skip with header_row_not_found.
        wb = load_workbook(path, data_only=True, keep_links=False)
        ws = wb.active
        if not ws.max_row or not ws.max_column:
            wb.close()
            audit["skip_reason"] = "empty_workbook"
            return None, audit
        audit["sheet_name"] = ws.title
        # Detect header row using the read-only handle.
        header_row = detect_header_row(
            ws,
            schema["workbook"]["header_detection"]["marker_text"],
            tuple(schema["workbook"]["header_detection"]["scan_rows"]),
        )
        wb.close()
        if header_row is None:
            audit["skip_reason"] = "header_row_not_found"
            return None, audit
        audit["header_row"] = header_row
    except Exception as e:
        audit["skip_reason"] = f"open_failed:{type(e).__name__}:{e}"
        return None, audit

    # Re-open in write mode for merge-fill.
    grid, title = read_grid_with_merge_fill(path, header_row)
    if grid is None:
        audit["skip_reason"] = "empty_workbook"
        return None, audit

    banner_date = find_date_in_banner(grid, header_row)
    filename_date = parse_date_from_filename(path.name)
    sheet_date = banner_date or filename_date
    if sheet_date is None:
        audit["skip_reason"] = "sheet_date_unparseable"
        return None, audit
    # Banner-year-typo correction: if the banner year doesn't match the
    # expected YEAR but the filename does, trust the filename. Catches the
    # 18122024.xlsx case where the banner shows 2022-12-18.
    if (banner_date is not None and filename_date is not None
            and banner_date.year != YEAR and filename_date.year == YEAR):
        sheet_date = filename_date
        audit.setdefault("warnings", []).append(
            f"banner_year_typo: banner={banner_date.isoformat()} → filename={filename_date.isoformat()}"
        )
    audit["sheet_date"] = sheet_date.isoformat()

    col_map, warnings = build_column_map(grid, header_row, schema)
    audit["warnings"].extend(warnings)
    if any(w.startswith("required_column_missing") for w in warnings):
        audit["skip_reason"] = "required_columns_missing"
        return None, audit

    data_start = header_row + 1
    # If a confirmatory sub-header was detected, the actual data starts one row below.
    if any(k.startswith("confirm_") for k in col_map):
        data_start = header_row + 2

    # Build rows.
    flags_per_row: dict[int, list[str]] = defaultdict(list)
    rows: list[dict] = []
    test_mappings = schema["mappings"]["test"]["aliases"]
    test_canonical = set(schema["mappings"]["test"]["canonical_values"])
    non_test_values = set(schema["mappings"]["test"].get("non_test_values", []))
    validity_true = set(schema["mappings"]["validity"]["map_to_true"])
    validity_false = set(schema["mappings"]["validity"]["map_to_false"])
    code_map = schema["final_report_code_map"]
    gso_pattern = re.compile(schema["gso_code"]["pattern"])

    def get_cell(r: int, logical: str):
        idx = col_map.get(logical)
        if idx is None:
            return None
        if idx >= len(grid[r]):
            return None
        return grid[r][idx]

    critical_logical = ["m_s_no", "s_no", "gso_code", "sample_name", "test", "result_raw"]
    n_raw = 0
    for r in range(data_start, len(grid)):
        n_raw += 1
        # Read all logical columns.
        raw: dict[str, object] = {}
        for logical in col_map:
            raw[logical] = get_cell(r, logical)
        # Skip if all critical empty.
        crit_vals = [raw.get(k) for k in critical_logical]
        if all((v is None) or (isinstance(v, str) and v.strip() == "") for v in crit_vals):
            audit["drops_count"] += 1
            continue
        # Normalise strings.
        for k, v in list(raw.items()):
            if isinstance(v, str):
                nv = normalise_text(v)
                if nv != v:
                    audit["changes_count"] += 1
                raw[k] = nv
        # GSO ISO placeholder. "ISO" means the sample was tested under an ISO
        # method (all such samples are environmental swabs / equipment
        # surfaces) — correctly outside GSO 1016 scope, so this is an
        # informational tag, not an error (reclassified 2026-08-08).
        if raw.get("gso_code") == "ISO":
            raw["gso_code"] = None
            flags_per_row[r].append("iso_method_outside_gso1016")
            audit["changes_count"] += 1
        # 'H' is not a GSO 1016 category letter — the lab used it as an
        # internal code. Map by product name (user-approved 2026-08-08):
        # cheddar → A-13 (hard/semi-hard cheese), ketchup → G-2 (tomato
        # products), all other sauces → G-3 (mayonnaise/sauces).
        if isinstance(raw.get("gso_code"), str) and raw["gso_code"].strip() == "H":
            nm = str(raw.get("sample_name") or "")
            if "شيدر" in nm:
                raw["gso_code"] = "A-13"
            elif "كاتشب" in nm:
                raw["gso_code"] = "G-2"
            else:
                raw["gso_code"] = "G-3"
            flags_per_row[r].append("gso_code_h_mapped_by_name")
            audit["changes_count"] += 1
        # GSO pattern violation.
        gso = raw.get("gso_code")
        if isinstance(gso, str) and not gso_pattern.match(gso):
            flags_per_row[r].append("gso_code_pattern_violation")
        # Test mapping.
        t = raw.get("test")
        if isinstance(t, str):
            # Non-test verdicts (e.g. "العينة مرفوضة", "موجودة") drop the row
            # with a separate flag instead of becoming an unrecognised test.
            if t in non_test_values:
                audit["drops_count"] += 1
                continue
            if t in test_mappings:
                new_t = test_mappings[t]
                audit["changes_count"] += 1
                raw["test"] = new_t
                t = new_t
            if t not in test_canonical:
                flags_per_row[r].append("test_value_unrecognised")
        # Validity mapping.
        v_raw = raw.get("validity")
        v_norm = v_raw.strip() if isinstance(v_raw, str) else v_raw
        if v_norm is None:
            v_bool = None
        elif v_norm in validity_true:
            v_bool = True
        elif v_norm in validity_false:
            v_bool = False
        else:
            v_bool = None
            flags_per_row[r].append("validity_value_unrecognised")
        # Final-report code.
        fr = raw.get("final_report_raw")
        code = match_final_report_code(fr, code_map)
        if fr is not None and code == "UNKNOWN":
            flags_per_row[r].append("final_report_unmapped")
        # M.S.No 4223 typo flag.
        m_s_no = raw.get("m_s_no")
        try:
            if m_s_no is not None and not (isinstance(m_s_no, float) and pd.isna(m_s_no)):
                if int(m_s_no) == 4223:
                    flags_per_row[r].append("msno_value_4223_suspect_typo")
        except (TypeError, ValueError):
            pass
        # Numeric parse.
        l_num, l_flags = parse_numeric_best_effort(raw.get("limit_raw"))
        for fl in l_flags:
            flags_per_row[r].append(f"limit_numeric_{fl}")
        r_num, r_flags = parse_numeric_best_effort(raw.get("result_raw"))
        for fl in r_flags:
            flags_per_row[r].append(f"result_numeric_{fl}")
        # Date sanity: flag if outside the expected year (could indicate
        # banner-from-merged-cell leakage or filename/banner mismatch).
        if sheet_date.year != YEAR:
            flags_per_row[r].append("date_year_mismatch")
        # Year coercion mismatch with filename — flagged at file level (skip per-row).
        # Pseudomonas spelling flag.
        if raw.get("test") == "سيدومومناس":
            flags_per_row[r].append("pseudomonas_spelling_unverified")

        rows.append({
            "source_file": path.name,
            "month_folder": path.parent.name,
            "sheet_date": sheet_date,
            "row_excel": r,
            "barcode": raw.get("barcode"),
            "m_s_no": m_s_no,
            "s_no": raw.get("s_no"),
            "gso_code": raw.get("gso_code"),
            "sample_name": raw.get("sample_name"),
            "restaurant_name": raw.get("restaurant_name"),
            "test": raw.get("test"),
            "limit_raw": raw.get("limit_raw"),
            "limit_numeric": l_num,
            "result_raw": raw.get("result_raw"),
            "result_numeric": r_num,
            "validity": v_bool,
            "final_report_raw": fr,
            "final_report_code": code,
            "conc_10_1": raw.get("conc_10_1"),
            "conc_10_2": raw.get("conc_10_2"),
            "conc_10_3": raw.get("conc_10_3"),
            "conc_10_4": raw.get("conc_10_4"),
            "confirm_vitek": raw.get("confirm_vitek"),
            "confirm_vidas": raw.get("confirm_vidas"),
            "confirm_rt_pcr": raw.get("confirm_rt_pcr"),
            "data_quality_flags": "|".join(sorted(set(flags_per_row.get(r, [])))) or None,
        })

    audit["rows_in"] = n_raw
    audit["rows_out"] = len(rows)
    audit["flags_count"] = sum(len(v) for v in flags_per_row.values())
    if not rows:
        return None, audit
    df = pd.DataFrame(rows)
    return df, audit


# ---------------------------------------------------------------------------
# Long → wide aggregation
# ---------------------------------------------------------------------------
def aggregate_to_wide(long_df: pd.DataFrame) -> pd.DataFrame:
    def first_non_null(s: pd.Series):
        for v in s:
            if v is None:
                continue
            try:
                if pd.isna(v):
                    continue
            except (TypeError, ValueError):
                pass
            return v
        return None

    def all_valid(s: pd.Series):
        if s.isna().all():
            return pd.NA
        # True iff all non-null entries are True. If any False -> False.
        if (s == False).any():
            return False
        if (s == True).all():
            return True
        # Mix of True and NA -> treat as True (only failures are decisive)
        return True

    def joined_failed_tests(group: pd.DataFrame) -> tuple[list, str]:
        failed = [t for t, v in zip(group["test"], group["validity"]) if v is False]
        failed_unique = sorted({f for f in failed if f is not None})
        return failed_unique

    grouped = long_df.groupby(["source_file", "m_s_no"], dropna=False, sort=True)
    out_rows = []
    for (sf, msno), g in grouped:
        invalid_tests = []
        for t, v in zip(g["test"], g["validity"]):
            # v comes from a pandas BooleanArray — use equality + NA-aware check
            # instead of `is False`, which doesn't match numpy.bool_(False).
            if pd.notna(v) and not bool(v) and t is not None and t not in invalid_tests:
                invalid_tests.append(t)
        is_valid_val = all_valid(g["validity"])
        flags_union = set()
        for fs in g["data_quality_flags"].dropna():
            for f in str(fs).split("|"):
                if f:
                    flags_union.add(f)
        sheet_date = first_non_null(g["sheet_date"])
        ts = pd.Timestamp(sheet_date) if sheet_date is not None else pd.NaT
        out_rows.append({
            "source_file": sf,
            "row_excel": int(g["row_excel"].min()),
            "sampling_date": ts,
            "year_month": ts.strftime("%Y-%m") if pd.notna(ts) else None,
            "iso_year_week": ts.strftime("%G-W%V") if pd.notna(ts) else None,
            "quarter": int(ts.quarter) if pd.notna(ts) else None,
            "day_of_week": int(ts.dayofweek) if pd.notna(ts) else None,
            # Use m_s_no (scoped by source file) as the sample identifier instead
            # of gso_code, which is a product standard, not a sample ID.
            # Fall back to gso_code when m_s_no is missing, and to source_file+row
            # only when both are missing, so every sample has a non-null ID.
            "sample_id": (f"{sf}_{int(msno)}" if pd.notna(msno)
                          else first_non_null(g["gso_code"])
                          if first_non_null(g["gso_code"]) is not None
                          else f"{sf}_row{int(g['row_excel'].min())}"),
            "sample_id_raw": (str(int(msno)) if pd.notna(msno)
                              else first_non_null(g["gso_code"])
                              if first_non_null(g["gso_code"]) is not None
                              else f"row{int(g['row_excel'].min())}"),
            "sample_name": first_non_null(g["sample_name"]),
            "category_canonical": None,
            "category_en": None,
            "category_raw": None,
            "sample_type": None,
            "facility_name": first_non_null(g["restaurant_name"]),
            "facility_chain": None,
            "facility_branch": None,
            "facility_name_raw": first_non_null(g["restaurant_name"]),
            "municipality": None,
            "municipality_type": None,
            "municipality_raw": None,
            "is_valid": is_valid_val,
            "is_failure": (not is_valid_val) if isinstance(is_valid_val, bool) else pd.NA,
            "invalid_tests_raw": "|".join(invalid_tests) if invalid_tests else "لا يوجد",
            "invalid_tests": invalid_tests,
            "n_failed_tests": len(invalid_tests),
            "m_s_no": None if pd.isna(msno) else int(msno),
            "s_no_first": int(g["s_no"].dropna().min()) if g["s_no"].notna().any() else None,
            "gso_code": first_non_null(g["gso_code"]),
            "barcode": first_non_null(g["barcode"]),
            "final_report_raw": first_non_null(g["final_report_raw"]),
            "final_report_code": first_non_null(g["final_report_code"]),
            "data_quality_flags": "|".join(sorted(flags_union)) or None,
            "year": YEAR,
        })
    wide = pd.DataFrame(out_rows)
    # De-duplicate cross-file re-exports: the same real barcode appearing in >1
    # source file (a workbook re-saved under a second name). Keep the first;
    # never dedup null barcodes — each is a distinct un-barcoded sample. (2026-07-16)
    _dup = wide["barcode"].notna() & wide["barcode"].duplicated(keep="first")
    if int(_dup.sum()):
        print(f"  de-duplicated {int(_dup.sum())} rows with a repeated barcode")
        wide = wide[~_dup].reset_index(drop=True)
    # De-duplicate BARCODE-LESS rows that are identical *including the verdict*
    # (same file/date/name/sample_id/is_valid/is_failure/invalid_tests). Barcode
    # dedup can't catch these (no barcode), yet a row repeated with the same
    # pass/fail result is a true double-count (e.g. موية فلتر N-1, a report read
    # twice). Rows that share an id/position but DIFFER in verdict are kept — they
    # are distinct samples (e.g. فلافل P-6/1: one compliant, one not). (2026-07-18)
    _idcols = ["source_file", "sampling_date", "sample_name", "sample_id",
               "is_valid", "is_failure", "invalid_tests_raw"]
    _nobar = wide["barcode"].isna()
    _bdup = pd.Series(False, index=wide.index)
    _bdup.loc[_nobar] = wide.loc[_nobar, _idcols].astype("string").duplicated(keep="first")
    if int(_bdup.sum()):
        print(f"  de-duplicated {int(_bdup.sum())} barcode-less rows identical incl. verdict")
        wide = wide[~_bdup].reset_index(drop=True)
    # Cast types to align with 2025.
    wide["is_valid"] = pd.array(wide["is_valid"].tolist(), dtype="boolean")
    wide["is_failure"] = pd.array(wide["is_failure"].tolist(), dtype="boolean")
    wide["n_failed_tests"] = wide["n_failed_tests"].astype("int32")
    wide["row_excel"] = wide["row_excel"].astype("int32")
    wide["m_s_no"] = pd.array(wide["m_s_no"].tolist(), dtype="Int64")
    wide["s_no_first"] = pd.array(wide["s_no_first"].tolist(), dtype="Int64")
    wide["quarter"] = pd.array(wide["quarter"].tolist(), dtype="Int8")
    wide["day_of_week"] = pd.array(wide["day_of_week"].tolist(), dtype="Int8")
    string_cols = [
        "source_file", "year_month", "iso_year_week", "sample_id", "sample_id_raw",
        "sample_name", "category_canonical", "category_en", "category_raw",
        "sample_type", "facility_name", "facility_chain", "facility_branch",
        "facility_name_raw", "municipality", "municipality_type", "municipality_raw",
        "invalid_tests_raw", "gso_code", "barcode",
        "final_report_raw", "final_report_code", "data_quality_flags",
    ]
    def _safe_str(v):
        if v is None:
            return None
        try:
            if pd.isna(v):
                return None
        except (TypeError, ValueError):
            pass
        return str(v)
    for c in string_cols:
        wide[c] = pd.array([_safe_str(v) for v in wide[c]], dtype="string")
    wide = wide.sort_values(["sampling_date", "source_file", "m_s_no"], kind="mergesort", ignore_index=True)
    return wide


# ---------------------------------------------------------------------------
# Run summary
# ---------------------------------------------------------------------------
def write_report(audits: list[dict], long_df: pd.DataFrame, wide_df: pd.DataFrame, path: Path) -> None:
    lines: list[str] = []
    total_files = len(audits)
    skipped = [a for a in audits if a["skip_reason"]]
    succeeded = [a for a in audits if not a["skip_reason"]]
    lines.append("# 2024 lab-data cleaning report\n")
    lines.append(f"- **Schema:** `lab_data_2024_v2`")
    lines.append(f"- **Source:** `2024-original/2024/**/*.xlsx`")
    lines.append(f"- **Total files seen:** {total_files}")
    lines.append(f"- **Files processed:** {len(succeeded)}")
    lines.append(f"- **Files skipped:** {len(skipped)}")
    lines.append(f"- **Long rows out:** {len(long_df)}")
    lines.append(f"- **Wide rows out (samples):** {len(wide_df)}")
    lines.append("")

    # Skipped files breakdown
    lines.append("## Skipped files")
    lines.append("")
    if skipped:
        by_reason: dict[str, list[str]] = defaultdict(list)
        for a in skipped:
            by_reason[a["skip_reason"]].append(f"{a['month_folder']}/{a['source_file']}")
        lines.append("| reason | count | files |")
        lines.append("|---|---:|---|")
        for reason, files in sorted(by_reason.items()):
            sample = ", ".join(files[:10])
            if len(files) > 10:
                sample += f", … (+{len(files) - 10})"
            lines.append(f"| `{reason}` | {len(files)} | {sample} |")
    else:
        lines.append("_none_")
    lines.append("")

    # Per-month summary
    lines.append("## Per-month summary")
    lines.append("")
    by_month: dict[str, dict] = defaultdict(lambda: {"files": 0, "rows_in": 0, "rows_out": 0, "skipped": 0, "flags": 0})
    for a in audits:
        m = a["month_folder"]
        by_month[m]["files"] += 1
        by_month[m]["rows_in"] += a["rows_in"]
        by_month[m]["rows_out"] += a["rows_out"]
        by_month[m]["skipped"] += 1 if a["skip_reason"] else 0
        by_month[m]["flags"] += a["flags_count"]
    lines.append("| month | files | skipped | rows_in | long_rows_out | flags |")
    lines.append("|---|---:|---:|---:|---:|---:|")
    for m in sorted(by_month):
        s = by_month[m]
        lines.append(f"| `{m}` | {s['files']} | {s['skipped']} | {s['rows_in']} | {s['rows_out']} | {s['flags']} |")
    lines.append("")

    # Flag distribution
    lines.append("## Flag distribution (long rows)")
    lines.append("")
    flag_counter: Counter = Counter()
    for fs in long_df["data_quality_flags"].dropna():
        for f in str(fs).split("|"):
            if f:
                flag_counter[f] += 1
    if flag_counter:
        lines.append("| flag | count |")
        lines.append("|---|---:|")
        for f, n in sorted(flag_counter.items(), key=lambda kv: (-kv[1], kv[0])):
            lines.append(f"| `{f}` | {n} |")
    else:
        lines.append("_no flags raised_")
    lines.append("")

    # Validity summary
    lines.append("## Wide-output validity summary")
    lines.append("")
    if len(wide_df):
        n = len(wide_df)
        n_valid = int((wide_df["is_valid"] == True).sum())
        n_invalid = int((wide_df["is_valid"] == False).sum())
        n_unknown = n - n_valid - n_invalid
        lines.append(f"- Samples: **{n}**")
        lines.append(f"- Valid: **{n_valid}** ({100*n_valid/n:.1f}%)")
        lines.append(f"- Invalid: **{n_invalid}** ({100*n_invalid/n:.1f}%)")
        lines.append(f"- Unknown: **{n_unknown}**")
    lines.append("")

    # Per-file audit (compact table; top 50 by rows_out + first 50 skipped)
    lines.append("## Per-file audit (all files)")
    lines.append("")
    lines.append("| month | file | header_row | sheet_date | rows_in | rows_out | flags | drops | skip_reason |")
    lines.append("|---|---|---:|---|---:|---:|---:|---:|---|")
    for a in sorted(audits, key=lambda x: (x["month_folder"], x["source_file"])):
        lines.append(
            f"| `{a['month_folder']}` | `{a['source_file']}` | {a['header_row'] or '-'} | "
            f"{a['sheet_date'] or '-'} | {a['rows_in']} | {a['rows_out']} | "
            f"{a['flags_count']} | {a['drops_count']} | {a['skip_reason'] or ''} |"
        )

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def _parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Lab-data cleaner (year-parameterised).")
    ap.add_argument("--year", type=int, default=2024,
                    help="Year to clean (default 2024). Must have an entry in YEAR_CONFIG.")
    return ap.parse_args()


def main() -> int:
    global YEAR, INPUT_ROOT, OUT_LONG, OUT_WIDE, REPORT_PATH
    args = _parse_args()
    YEAR = args.year
    if YEAR not in YEAR_CONFIG:
        print(f"year {YEAR} has no entry in YEAR_CONFIG", file=sys.stderr)
        return 2
    INPUT_ROOT = YEAR_CONFIG[YEAR]["input_root"]
    OUT_LONG    = ROOT / "cleaned" / f"data{YEAR}_long.parquet"
    OUT_WIDE    = ROOT / "cleaned" / f"data{YEAR}.parquet"
    REPORT_PATH = ROOT / "reports" / f"data{YEAR}_clean_report.md"
    print(f"YEAR={YEAR}  INPUT_ROOT={INPUT_ROOT}")

    schema = load_schema()
    files = sorted(INPUT_ROOT.rglob("*.xlsx"))
    if not files:
        print(f"no .xlsx files under {INPUT_ROOT}", file=sys.stderr)
        return 2

    all_audits: list[dict] = []
    long_frames: list[pd.DataFrame] = []
    for i, f in enumerate(files, 1):
        df, audit = clean_one_file(f, schema)
        all_audits.append(audit)
        if df is not None:
            long_frames.append(df)
        if i % 25 == 0 or i == len(files):
            print(f"  [{i}/{len(files)}] {f.parent.name}/{f.name}: "
                  f"{'skip='+audit['skip_reason'] if audit['skip_reason'] else f'rows={audit['rows_out']}'}")

    if not long_frames:
        print("no data produced — aborting", file=sys.stderr)
        # still write a report so the user sees what was skipped
        empty_long = pd.DataFrame()
        empty_wide = pd.DataFrame()
        write_report(all_audits, empty_long, empty_wide, REPORT_PATH)
        return 1

    long_df = pd.concat(long_frames, ignore_index=True, sort=False)
    long_df["year"] = YEAR

    # Cast long columns for parquet FIRST, then sort. Pandas can't lexsort
    # object-dtype columns with mixed int/None/NaN values; converting to Int64
    # first ensures consistent ordering with NA-last semantics.
    long_df["row_excel"] = long_df["row_excel"].astype("int32")
    long_df["m_s_no"] = pd.array(
        [int(v) if isinstance(v, (int, float)) and not isinstance(v, bool) and not pd.isna(v) else pd.NA for v in long_df["m_s_no"]],
        dtype="Int64",
    )
    long_df["s_no"] = pd.array(
        [int(v) if isinstance(v, (int, float)) and not isinstance(v, bool) and not pd.isna(v) else pd.NA for v in long_df["s_no"]],
        dtype="Int64",
    )
    long_df["limit_numeric"] = long_df["limit_numeric"].astype("Float64")
    long_df["result_numeric"] = long_df["result_numeric"].astype("Float64")
    long_df["validity"] = pd.array(long_df["validity"].tolist(), dtype="boolean")
    string_cols = [
        "source_file", "month_folder", "barcode", "gso_code", "sample_name",
        "restaurant_name", "test", "limit_raw", "result_raw",
        "final_report_raw", "final_report_code",
        "conc_10_1", "conc_10_2", "conc_10_3", "conc_10_4",
        "confirm_vitek", "confirm_vidas", "confirm_rt_pcr",
        "data_quality_flags",
    ]
    def _safe_str(v):
        if v is None:
            return None
        try:
            if pd.isna(v):
                return None
        except (TypeError, ValueError):
            pass
        return str(v)
    for c in string_cols:
        long_df[c] = pd.array([_safe_str(v) for v in long_df[c]], dtype="string")
    long_df["sheet_date"] = pd.to_datetime(long_df["sheet_date"]).astype("datetime64[ms]")

    # NOW sort, after all dtypes are settled.
    long_df = long_df.sort_values(
        ["sheet_date", "source_file", "m_s_no", "s_no", "test"],
        kind="mergesort",
        na_position="last",
        ignore_index=True,
    )

    OUT_LONG.parent.mkdir(parents=True, exist_ok=True)
    long_df.to_parquet(OUT_LONG, compression="zstd", index=False)
    print(f"wrote {OUT_LONG} ({len(long_df)} rows × {long_df.shape[1]} cols)")

    wide_df = aggregate_to_wide(long_df)
    wide_df["sampling_date"] = pd.to_datetime(wide_df["sampling_date"]).astype("datetime64[ms]")
    OUT_WIDE.parent.mkdir(parents=True, exist_ok=True)
    wide_df.to_parquet(OUT_WIDE, compression="zstd", index=False)
    print(f"wrote {OUT_WIDE} ({len(wide_df)} rows × {wide_df.shape[1]} cols)")

    write_report(all_audits, long_df, wide_df, REPORT_PATH)
    print(f"wrote {REPORT_PATH}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
