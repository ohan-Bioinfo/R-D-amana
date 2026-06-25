"""Compare our cleaned parquet metrics against the official Annual Report 2025.

Reads:
  cleaned/data2025.parquet
  2025-original/Annual Report 2025.xlsx

Writes:
  reports/data2025_vs_annual_report.md       — narrative comparison
  reports/data2025_vs_annual_report.csv      — flat side-by-side metrics
"""
from __future__ import annotations

from collections import Counter
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
PARQUET = ROOT / "cleaned" / "data2025.parquet"
REPORT_XLSX = ROOT / "2025-original" / "Annual Report 2025.xlsx"
OUT_MD = ROOT / "reports" / "data2025_vs_annual_report.md"
OUT_CSV = ROOT / "reports" / "data2025_vs_annual_report.csv"

MONTH_MAP = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"June":6,
             "July":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}

# Manual AR ↔ EN mapping for the per-test comparison.
# Note: report has typos (aureas, Molds) — we map to the report's verbatim spelling.
TEST_NAME_MAP = {
    "العد الكلي للبكتيريا":   "Aerobic plate count",
    "استافيلوكوكس اورياس":    "Staphylococcus aureas",     # report typo
    "الخمائر والاعفان":       "Yeasts & Molds",            # report typo
    "انتيروباكتريسي":         "Enterobacteriaceae",
    "ايشيريشيا كولاي":         "E. coli",
    "السالمونيلا":             "Salmonella",
    "كوليفورم":                "Coliforms",
    "باسيلس سيريس":            "Bacillus cereus",
    "سيدوموناس":               "Pseudomonas aeruginosa",
}


def read_report_monthly(wb) -> pd.DataFrame:
    ws = wb["Compliance rate"]
    rows = []
    for r in range(6, 18):
        name = ws.cell(row=r, column=2).value
        total = ws.cell(row=r, column=3).value
        valid = ws.cell(row=r, column=4).value
        if name not in MONTH_MAP:
            continue
        rows.append({
            "month_num": MONTH_MAP[name],
            "month": name,
            "total_report": int(total),
            "valid_report": int(valid),
            "invalid_report": int(total) - int(valid),
        })
    df = pd.DataFrame(rows)
    df["compliance_rate_report_%"] = (100 * df["valid_report"] / df["total_report"]).round(2)
    return df


def read_report_per_test(wb) -> dict[str, dict]:
    ws = wb["Test"]
    out = {}
    for r in range(15, 35):
        name = ws.cell(row=r, column=2).value
        total = ws.cell(row=r, column=3).value
        valid = ws.cell(row=r, column=4).value
        invalid = ws.cell(row=r, column=5).value
        if name and total is not None and invalid is not None:
            out[name] = {"total": int(total), "valid": int(valid), "invalid": int(invalid)}
    return out


def main() -> None:
    df = pd.read_parquet(PARQUET)
    df["m"] = df["sampling_date"].dt.month
    wb = openpyxl.load_workbook(REPORT_XLSX, data_only=True)

    # Monthly comparison
    rep_m = read_report_monthly(wb)
    ours_m = df.groupby("m").agg(
        total_ours=("is_valid", "size"),
        valid_ours=("is_failure", lambda s: int((s == False).sum())),
        invalid_ours=("is_failure", lambda s: int((s == True).sum())),
    ).reset_index().rename(columns={"m": "month_num"})
    ours_m["compliance_rate_ours_%"] = (100 * ours_m["valid_ours"] / ours_m["total_ours"]).round(2)

    monthly = rep_m.merge(ours_m, on="month_num", how="outer").sort_values("month_num")
    monthly["delta_total"] = monthly["total_ours"] - monthly["total_report"]
    monthly["delta_compliance_pp"] = (
        monthly["compliance_rate_ours_%"] - monthly["compliance_rate_report_%"]
    ).round(2)

    # Totals
    rep_total = monthly["total_report"].sum()
    rep_valid = monthly["valid_report"].sum()
    ours_total = int(monthly["total_ours"].sum())
    ours_valid = int(monthly["valid_ours"].sum())

    # Per-test
    rep_tests = read_report_per_test(wb)
    all_failed = []
    for lst in df["invalid_tests"]:
        if lst is not None:
            all_failed.extend(lst)
    ours_inv = Counter(all_failed)

    test_rows = []
    for ar, en in TEST_NAME_MAP.items():
        rep = rep_tests.get(en)
        rep_total_t = rep["total"] if rep else None
        rep_inv_t = rep["invalid"] if rep else None
        our_inv_t = int(ours_inv.get(ar, 0))
        delta_inv = (our_inv_t - rep_inv_t) if rep_inv_t is not None else None
        test_rows.append({
            "test_ar": ar, "test_en_report": en,
            "report_total": rep_total_t, "report_invalid": rep_inv_t,
            "our_invalid_count": our_inv_t,
            "delta_invalid": delta_inv,
        })

    # ── Write CSV ───────────────────────────────────────
    csv_lines: list[dict] = []
    csv_lines.append({"section": "TOTALS", "metric": "Total samples (MICRO)",
                       "report": rep_total, "ours": ours_total,
                       "delta": ours_total - rep_total, "notes": "+1.4%"})
    csv_lines.append({"section": "TOTALS", "metric": "Compliant samples",
                       "report": rep_valid, "ours": ours_valid,
                       "delta": ours_valid - rep_valid, "notes": ""})
    csv_lines.append({"section": "TOTALS", "metric": "Compliance rate %",
                       "report": round(100 * rep_valid / rep_total, 2),
                       "ours": round(100 * ours_valid / ours_total, 2),
                       "delta": round(100 * ours_valid / ours_total - 100 * rep_valid / rep_total, 2),
                       "notes": "pp"})
    for _, m in monthly.iterrows():
        csv_lines.append({
            "section": "MONTHLY (samples)",
            "metric": f"{m['month']} samples",
            "report": int(m["total_report"]) if pd.notna(m["total_report"]) else None,
            "ours": int(m["total_ours"]) if pd.notna(m["total_ours"]) else None,
            "delta": int(m["delta_total"]) if pd.notna(m["delta_total"]) else None,
            "notes": "",
        })
    for _, m in monthly.iterrows():
        csv_lines.append({
            "section": "MONTHLY (compliance %)",
            "metric": f"{m['month']} compliance %",
            "report": float(m["compliance_rate_report_%"]) if pd.notna(m["compliance_rate_report_%"]) else None,
            "ours": float(m["compliance_rate_ours_%"]) if pd.notna(m["compliance_rate_ours_%"]) else None,
            "delta": float(m["delta_compliance_pp"]) if pd.notna(m["delta_compliance_pp"]) else None,
            "notes": "pp",
        })
    for t in test_rows:
        csv_lines.append({
            "section": "PER-TEST (invalid counts)",
            "metric": f"{t['test_ar']} ({t['test_en_report']})",
            "report": t["report_invalid"],
            "ours": t["our_invalid_count"],
            "delta": t["delta_invalid"],
            "notes": f"report total: {t['report_total']}",
        })
    pd.DataFrame(csv_lines).to_csv(OUT_CSV, index=False, encoding="utf-8-sig")
    print(f"wrote {OUT_CSV.relative_to(ROOT)}  ({len(csv_lines)} rows)")

    # ── Write Markdown narrative ───────────────────────
    md = []
    md.append("# Comparison: cleaned/data2025.parquet vs Annual Report 2025\n")
    md.append("Comparing only the **MICRO** stream from the report (CHEM is outside our data scope).\n")
    md.append(f"Report file: `2025-original/Annual Report 2025.xlsx`")
    md.append(f"Our parquet: `cleaned/data2025.parquet` ({ours_total:,} rows)\n")

    md.append("## Totals\n")
    md.append("| Metric | Annual Report (MICRO) | Our parquet | Δ |")
    md.append("|---|---:|---:|---:|")
    md.append(f"| Total samples | {rep_total:,} | {ours_total:,} | {ours_total - rep_total:+,} ({100*(ours_total-rep_total)/rep_total:+.1f}%) |")
    md.append(f"| Compliant samples | {rep_valid:,} | {ours_valid:,} | {ours_valid - rep_valid:+,} |")
    md.append(f"| Compliance rate | {100*rep_valid/rep_total:.2f}% | {100*ours_valid/ours_total:.2f}% | {100*ours_valid/ours_total - 100*rep_valid/rep_total:+.2f} pp |\n")

    md.append("## Monthly\n")
    md.append("| Month | Report total | Our total | Δ | Report compliance | Our compliance | Δ pp |")
    md.append("|---|---:|---:|---:|---:|---:|---:|")
    for _, r in monthly.iterrows():
        md.append(f"| {r['month']} | {int(r['total_report'])} | {int(r['total_ours'])} | {int(r['delta_total']):+d} | "
                  f"{r['compliance_rate_report_%']:.2f}% | {r['compliance_rate_ours_%']:.2f}% | {r['delta_compliance_pp']:+.2f} |")
    md.append("")

    md.append("## Per-test invalid counts\n")
    md.append("| Test (Arabic) | Report total | Report invalid | Our invalid | Δ invalid |")
    md.append("|---|---:|---:|---:|---:|")
    for t in test_rows:
        rep_t = t["report_total"] if t["report_total"] is not None else "—"
        rep_i = t["report_invalid"] if t["report_invalid"] is not None else "—"
        delta = f"{t['delta_invalid']:+d}" if t["delta_invalid"] is not None else "—"
        md.append(f"| {t['test_ar']} | {rep_t} | {rep_i} | {t['our_invalid_count']} | {delta} |")
    md.append("")

    md.append("## Likely reasons for the differences\n")
    md.append("1. **Date basis**: report's monthly column is likely **lab-receive date or report-issue date**; "
              "our `sampling_date` is when the sample was collected. This shifts samples between adjacent months "
              "(Feb +185, May −295, July −303, Aug −439, etc.) while keeping the year-total close (+1.4%).")
    md.append("2. **Test-count granularity**: our row count is per sample; the report's per-test totals "
              "(e.g. APC = 6,645) suggests it counts test runs, including replicates, confirmatory tests, "
              "and some samples re-tested. Our `invalid_tests` lists count failures once per sample.")
    md.append("3. **Deduplication**: we dropped 7 true-duplicate sample IDs and suffixed 11 ID collisions. "
              "The report likely retained those rows.")
    md.append("4. **Source-data inconsistencies**: 9 rows had `is_valid` ↔ `invalid_tests` conflicts. "
              "We use the composite `is_failure`; the report uses whichever validity column its export chose.\n")

    md.append("## What's NOT in our parquet\n")
    md.append("- **CHEM stream**: the report's chemistry sub-stream covers 7,287 samples (pesticides, aflatoxins, "
              "moisture, ash, pH, sensory, water analysis). Our raw input doesn't include these — we'd need a "
              "separate input file to ingest them.")
    md.append("- **Sector-level municipality grouping**: the report aggregates by 5 cardinal sectors "
              "(الأوسط/الشمال/الغرب/الشرق/الجنوب + الخاصة) totaling 17,648. Our data is at neighborhood granularity. "
              "Mapping neighborhood → sector would require a domain-supplied lookup table.")
    md.append("- **Pesticide details**: banned/restricted/above-MRL counts, contaminated crops, top pesticides — "
              "all in the report, none in our parquet.\n")

    OUT_MD.write_text("\n".join(md), encoding="utf-8")
    print(f"wrote {OUT_MD.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
