"""2025 audit workbook — numbers + GSO/name classification, for Muhannad's sign-off.

Because 2025 has NO native lab GSO (every sample is derived from its sample_type
bucket), the audit hunts for likely mis-classifications by flagging samples where
a HIGH-CONFIDENCE category keyword in the name disagrees with the assigned GSO —
excluding environmental swabs and names he has already explicitly corrected.

Sheets:
  • Numbers        — our 2025 totals vs official, Δ
  • Dedup_findings — exact-id / full-row / re-test / identity duplicate checks
  • Likely_errors  — curated suspects (fillable: corrected_gso dropdown + notes)
  • All_conflicts  — every bucket-vs-name conflict, reference only
  • Valid_GSO_categories — dropdown source

Run:  microbiology/.venv/bin/python microbiology/scripts/build_audit_2025.py
Out:  microbiology/reports/audit_2025_to_fill.xlsx
"""
from __future__ import annotations
from collections import defaultdict
from pathlib import Path
import re
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

from build_classification_table import classify, _val
from build_dashboard_combined import (
    NAME_CORRECTIONS, classify_sample_name, classify_name_override,
)

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "reports" / "audit_2025_to_fill.xlsx"
OFF = (11404, 8345)   # official 2025 (samples, compliant)

# High-confidence indicators: if the name contains one of these AND the current
# GSO is not the implied category, it's a suspect. Kept deliberately tight to
# avoid the name-keyword classifier's false matches (ثوم→produce, رز→egg, …).
INDICATORS = [
    (["سلطة", "ايدام", "محشي", "معجنات", "فطيرة", "وجبة", "بيتزا", "برجر",
      "ساندويش", "شاورما", "صيادية", "مقلوبة", "كبسة", "منسف"], "Ready to Eat Foods"),
    (["كيك", "حلا", "بسكوت", "شوكولاتة", "شوكولا", "حلوى", "بسبوسة", "كنافة",
      "بقلاوة", "دونت", "كروسان", "معمول", "غريبة"], "Chocolate, Sweets and their Ingredients"),
    (["جبنة", "حليب", "زبادي", "قشطة", "لبنة", "حلوم"], "Dairy Products"),
    (["سمك", "تونة", "روبيان", "جمبري", "سلمون", "حبار"], "Fish and Shellfish their Products"),
]

HDR = PatternFill("solid", fgColor="1C2742")
YEL = PatternFill("solid", fgColor="F59E0B")
YELCELL = PatternFill("solid", fgColor="FFF3C7")
WBF = Font(color="FFFFFF", bold=True)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _retest_base(x):
    return re.sub(r'[-_ ]*[Rr]0*\d+$', '', str(x).strip())


def build():
    d = pd.read_parquet(ROOT / "cleaned" / "data2025.parquet")
    recs = d.to_dict("records")
    valid_gso = set()

    # ---- suspects + all-conflicts ----
    suspects = defaultdict(lambda: {"cur": None, "bucket": None, "suggest": None,
                                    "n": 0, "nc": 0})
    conflicts = defaultdict(lambda: {"cur": None, "bucket": None, "suggest": None, "n": 0})
    for r in recs:
        cur, src = classify(r)
        valid_gso.add(cur)
        nm = str(_val(r.get("sample_name")) or "").strip()
        bucket = str(_val(r.get("sample_type")) or "—")
        # curated suspects (exclude swabs + already-ruled names)
        if cur != "Environmental Swabs" and nm not in NAME_CORRECTIONS:
            sug = None
            if cur == "Miscellaneous Foods":
                sug = "(needs a category)"
            else:
                for kws, cat in INDICATORS:
                    if cur != cat and any(k in nm for k in kws):
                        sug = cat
                        break
            if sug:
                s = suspects[nm]
                s["cur"] = cur; s["bucket"] = bucket; s["suggest"] = sug
                s["n"] += 1
                if r.get("is_failure") is True:
                    s["nc"] += 1
        # all bucket-vs-name conflicts (reference)
        if src == "sample_type-bucket":
            kw = classify_name_override(nm) or classify_sample_name(nm)
            if kw and kw != cur:
                c = conflicts[nm]
                c["cur"] = cur; c["bucket"] = bucket; c["suggest"] = kw; c["n"] += 1

    # ---- numbers ----
    rs = len(d)
    rc = int((d["is_failure"] != True).sum())  # noqa: E712
    numbers = [
        ("Samples", OFF[0], rs, rs - OFF[0]),
        ("Compliant", OFF[1], rc, rc - OFF[1]),
        ("Non-compliant", OFF[0] - OFF[1], rs - rc, (rs - rc) - (OFF[0] - OFF[1])),
        ("Compliance %", round(100 * OFF[1] / OFF[0], 2), round(100 * rc / rs, 2),
         round(100 * rc / rs - 100 * OFF[1] / OFF[0], 2)),
    ]

    # ---- dedup findings ----
    key = d["sample_id"].astype("string").str.strip()
    exact = int(key.dropna().duplicated().sum())
    hashable = [c for c in d.columns
                if not d[c].map(lambda v: isinstance(v, (list, dict, set, np.ndarray))).any()]
    fullrow = int(d[hashable].duplicated().sum())
    bases = key.dropna().map(_retest_base)
    shared = int(len(bases) - bases.nunique())
    idcols = ["sampling_date", "sample_name", "facility_name", "sample_id", "source_file"]
    idcols = [c for c in idcols if c in d.columns]
    ident = int(d[idcols].astype("string").duplicated().sum())
    dedup = [
        ("Exact sample_id duplicates", exact),
        ("Full-row duplicates (all columns)", fullrow),
        ("Re-test rows sharing a base id", shared),
        ("Identity dups (date+name+facility+id+file)", ident),
    ]

    # ---- write ----
    wb = Workbook()

    def style_header(ws, yellow=()):
        for j, c in enumerate(ws[1], 1):
            c.fill = YEL if j in yellow else HDR
            c.font = WBF; c.alignment = CEN
        ws.freeze_panes = "A2"

    # Numbers
    ws = wb.active; ws.title = "Numbers"
    ws.append(["metric", "official_2025", "our_2025", "delta"])
    for row in numbers:
        ws.append(list(row))
    ws.append([])
    ws.append(["note", "2025 has no native GSO; +160 vs official is a sampling-date basis shift (see dashboard)."])
    style_header(ws)
    for col, w in {"A": 22, "B": 14, "C": 12, "D": 10}.items():
        ws.column_dimensions[col].width = w

    # Dedup
    ws = wb.create_sheet("Dedup_findings")
    ws.append(["check", "count", "action"])
    for name, cnt in dedup:
        ws.append([name, cnt, "none" if cnt == 0 else "REVIEW"])
    style_header(ws)
    for col, w in {"A": 42, "B": 8, "C": 12}.items():
        ws.column_dimensions[col].width = w

    # Likely_errors (fillable)
    ws = wb.create_sheet("Likely_errors")
    ws.append(["current_gso", "sample_name", "bucket", "suggested_gso",
               "samples", "non_compliant", "nc_rate_%", "corrected_gso", "notes"])
    rows = sorted(suspects.items(), key=lambda kv: -kv[1]["n"])
    for nm, s in rows:
        rate = round(100 * s["nc"] / s["n"], 1) if s["n"] else 0
        ws.append([s["cur"], nm, s["bucket"], s["suggest"],
                   s["n"], s["nc"], rate, "", ""])
    style_header(ws, yellow=(8, 9))
    for i in range(2, len(rows) + 2):
        ws.cell(row=i, column=8).fill = YELCELL
        ws.cell(row=i, column=9).fill = YELCELL
    dv = DataValidation(type="list",
                        formula1=f"Valid_GSO_categories!$A$2:$A${len(valid_gso) + 1}",
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"H2:H{len(rows) + 1}")
    for col, w in {"A": 30, "B": 32, "C": 16, "D": 30, "E": 9, "F": 12, "G": 9,
                   "H": 30, "I": 26}.items():
        ws.column_dimensions[col].width = w

    # All_conflicts (reference)
    ws = wb.create_sheet("All_conflicts")
    ws.append(["current_gso", "sample_name", "bucket", "name_keyword_suggests", "samples"])
    for nm, c in sorted(conflicts.items(), key=lambda kv: -kv[1]["n"]):
        ws.append([c["cur"], nm, c["bucket"], c["suggest"], c["n"]])
    style_header(ws)
    for col, w in {"A": 30, "B": 32, "C": 16, "D": 30, "E": 9}.items():
        ws.column_dimensions[col].width = w

    # Valid GSO dropdown source
    vg = wb.create_sheet("Valid_GSO_categories")
    vg.append(["valid_gso_category"])
    for g in sorted(valid_gso):
        vg.append([g])
    vg.cell(row=1, column=1).fill = HDR
    vg.cell(row=1, column=1).font = WBF
    vg.column_dimensions["A"].width = 40

    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  Numbers: our {rs} vs official {OFF[0]} (Δ {rs - OFF[0]})")
    print(f"  Dedup: {dedup}")
    print(f"  Likely_errors: {len(rows)} names ({sum(s['n'] for s in suspects.values())} samples)")
    print(f"  All_conflicts: {len(conflicts)} names")


if __name__ == "__main__":
    build()
