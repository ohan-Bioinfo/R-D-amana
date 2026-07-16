"""Build a fillable Excel to (a) validate numbers vs the official annual figures
and (b) let Muhannad correct the samples that are unclassified / classification-
doubtful. Doubtful = classified only by name keyword or fell to Miscellaneous.

Rows are grouped by distinct sample_name so each product is classified once; the
YELLOW 'corrected_gso_category' column has a dropdown of the valid GSO categories.
Send it back and I apply the corrections (name rule / bucket / per-sample).

Run:  microbiology/.venv/bin/python microbiology/scripts/build_correction_workbook.py
Out:  microbiology/reports/classification_corrections_to_fill.xlsx
"""
from __future__ import annotations
from collections import defaultdict
from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

from build_classification_table import classify, _val

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "reports" / "classification_corrections_to_fill.xlsx"
OFF = {2024: (9108, 6399), 2025: (11404, 8345)}     # (samples, compliant)
DOUBTFUL = {"name-keyword", "Miscellaneous-fallback"}


def build():
    rows = []
    valid_gso = set()
    review = defaultdict(lambda: {"n": 0, "years": set(), "cats": set(), "gso": None, "src": None})
    counts = {}
    for y in (2024, 2025):
        d = pd.read_parquet(ROOT / "cleaned" / f"data{y}.parquet")
        counts[y] = (len(d), int((d["is_failure"] != True).sum()))  # noqa: E712
        for r in d.to_dict("records"):
            gso, src = classify(r)
            valid_gso.add(gso)
            if src in DOUBTFUL:
                nm = str(_val(r.get("sample_name")) or "(no name)")
                g = review[nm]
                g["n"] += 1; g["years"].add(y); g["gso"] = gso; g["src"] = src
                cat = _val(r.get("category_canonical")) or _val(r.get("gso_category_name_en"))
                if cat:
                    g["cats"].add(str(cat))

    review_rows = sorted(
        [(nm, v["n"], "/".join(str(x) for x in sorted(v["years"])), v["gso"], v["src"],
          " · ".join(sorted(v["cats"])[:3]))
         for nm, v in review.items()],
        key=lambda t: -t[1])

    # ---- write ----
    HDR = PatternFill("solid", fgColor="1C2742")
    YEL = PatternFill("solid", fgColor="F59E0B")
    YELCELL = PatternFill("solid", fgColor="FFF3C7")
    WB = Font(color="FFFFFF", bold=True)
    CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)

    numbers = pd.DataFrame([
        {"year": y, "metric": m, "official": o, "our": r, "delta": r - o}
        for y in (2024, 2025)
        for m, o, r in [
            ("Samples", OFF[y][0], counts[y][0]),
            ("Compliant", OFF[y][1], counts[y][1]),
            ("Non-compliant", OFF[y][0] - OFF[y][1], counts[y][0] - counts[y][1]),
            ("Compliance %", round(100 * OFF[y][1] / OFF[y][0], 2), round(100 * counts[y][1] / counts[y][0], 2)),
        ]])
    review_df = pd.DataFrame(review_rows, columns=[
        "sample_name", "samples", "year(s)", "current_gso", "source", "example_raw_category"])
    review_df["corrected_gso_category"] = ""
    review_df["notes"] = ""
    valid_df = pd.DataFrame(sorted(valid_gso), columns=["valid_gso_category"])

    # Pure openpyxl write (avoids the pandas ExcelWriter interaction bug).
    from openpyxl import Workbook
    wb = Workbook()

    def _clean(v):
        return None if (isinstance(v, float) and pd.isna(v)) else v

    def write_sheet(title, df, yellow_cols=()):
        ws = wb.active if title == "Numbers" else wb.create_sheet(title)
        ws.title = title
        for j, col in enumerate(df.columns, 1):
            c = ws.cell(row=1, column=j, value=col)
            c.fill = YEL if (j - 1) in yellow_cols else HDR
            c.font = WB; c.alignment = CEN
        for i, (_, row) in enumerate(df.iterrows(), 2):
            for j, val in enumerate(row.tolist(), 1):
                cc = ws.cell(row=i, column=j, value=_clean(val))
                if (j - 1) in yellow_cols:
                    cc.fill = YELCELL
        ws.freeze_panes = "A2"
        return ws

    write_sheet("Numbers", numbers)
    rv = write_sheet("To_review", review_df, yellow_cols=(6, 7))
    write_sheet("Valid_GSO_categories", valid_df)

    gso_list = sorted(valid_gso)
    dv = DataValidation(type="list",
                        formula1=f"Valid_GSO_categories!$A$2:$A${len(gso_list) + 1}",
                        allow_blank=True)
    rv.add_data_validation(dv)
    dv.add(f"G2:G{len(review_df) + 1}")
    for col, w in {"A": 34, "B": 9, "C": 9, "D": 34, "E": 22, "F": 34, "G": 34, "H": 40}.items():
        rv.column_dimensions[col].width = w
    for col, w in {"A": 8, "B": 16, "C": 12, "D": 22, "E": 10}.items():
        wb["Numbers"].column_dimensions[col].width = w
    wb["Valid_GSO_categories"].column_dimensions["A"].width = 40
    wb.save(OUT)

    print(f"wrote {OUT}")
    print(f"  distinct doubtful sample_names to review: {len(review_rows)} "
          f"(covering {sum(r[1] for r in review_rows)} samples)")


if __name__ == "__main__":
    build()
