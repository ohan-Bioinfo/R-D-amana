"""Phase 1 read-only exploration of a single .xlsx file.

Usage: python explore.py <path-to-xlsx> <path-to-output-md>

Does NOT modify the source file. Produces a Markdown exploration report and
also prints it to stdout. The report is intentionally exhaustive — it will be
the input for the Phase 2 schema proposal.
"""
from __future__ import annotations

import io
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ARABIC_DIACRITICS = "ًٌٍَُِّْٰٕٓٔ"
TATWEEL = "ـ"
RTL_MARKS = "‎‏‪‫‬‭‮⁦⁧⁨⁩"
ZERO_WIDTH = "​‌‍﻿"
NBSP = " "

MISSING_TOKENS = {
    "": "empty_string",
    "NA": "literal_NA",
    "N/A": "literal_N_slash_A",
    "n/a": "literal_n_slash_a",
    "na": "literal_na",
    "None": "literal_None",
    "none": "literal_none",
    "NULL": "literal_NULL",
    "null": "literal_null",
    "-": "ascii_dash",
    "--": "ascii_double_dash",
    "–": "en_dash",
    "—": "em_dash",
    "−": "minus_sign",
    "?": "question_mark",
    ".": "single_dot",
}


def normalize_arabic(s: str) -> str:
    """Aggressive normalization for *comparison only* — never written back."""
    if not isinstance(s, str):
        return s
    s = unicodedata.normalize("NFKC", s)
    for ch in ARABIC_DIACRITICS + TATWEEL + ZERO_WIDTH:
        s = s.replace(ch, "")
    for ch in RTL_MARKS:
        s = s.replace(ch, "")
    s = s.replace(NBSP, " ")
    s = s.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    s = s.replace("ى", "ي")
    s = s.replace("ة", "ه")
    s = s.replace("ک", "ك")
    s = s.replace("ی", "ي")
    s = re.sub(r"\s+", " ", s).strip()
    return s.casefold()


def char_inventory(s: str) -> dict[str, int]:
    """Count suspicious characters in a string."""
    if not isinstance(s, str):
        return {}
    counts: dict[str, int] = {}
    for ch in s:
        cp = ord(ch)
        name = None
        if ch in RTL_MARKS:
            name = f"RTL-mark U+{cp:04X}"
        elif ch in ZERO_WIDTH:
            name = f"zero-width U+{cp:04X}"
        elif ch == NBSP:
            name = "NBSP U+00A0"
        elif ch == TATWEEL:
            name = "TATWEEL U+0640"
        elif ch in ARABIC_DIACRITICS:
            name = f"Arabic-diacritic U+{cp:04X}"
        elif ch == "\t":
            name = "TAB"
        elif ch == "\n":
            name = "NEWLINE"
        elif cp > 0x10000:
            name = f"supplementary U+{cp:06X}"
        if name:
            counts[name] = counts.get(name, 0) + 1
    return counts


def looks_numeric(v) -> bool:
    if isinstance(v, (int, float)):
        return True
    if not isinstance(v, str):
        return False
    s = v.strip().replace(",", "").replace("٬", "").replace("٫", ".")
    s = s.replace("٠", "0").replace("١", "1").replace("٢", "2").replace("٣", "3").replace("٤", "4")
    s = s.replace("٥", "5").replace("٦", "6").replace("٧", "7").replace("٨", "8").replace("٩", "9")
    if s == "":
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


def is_missing(v) -> tuple[bool, str | None]:
    if v is None:
        return True, "None_object"
    if isinstance(v, float) and pd.isna(v):
        return True, "NaN"
    if isinstance(v, str):
        if v in MISSING_TOKENS:
            return True, MISSING_TOKENS[v]
        if v.strip() == "" and v != "":
            return True, "whitespace_only"
        norm = unicodedata.normalize("NFKC", v).strip()
        if norm in MISSING_TOKENS:
            return True, f"normalized_{MISSING_TOKENS[norm]}"
    return False, None


def md_escape(s: str) -> str:
    if not isinstance(s, str):
        return repr(s)
    return s.replace("|", "\\|").replace("\n", "\\n")


def explore_workbook_with_openpyxl(path: Path) -> dict:
    """Pull structural info openpyxl exposes that pandas hides."""
    wb = load_workbook(path, data_only=False, read_only=False, keep_links=False)
    info: dict = {"sheets": {}, "all_sheet_names": wb.sheetnames}
    for name in wb.sheetnames:
        ws = wb[name]
        merged = [str(r) for r in ws.merged_cells.ranges]
        hidden_cols = []
        for col_letter, col_dim in ws.column_dimensions.items():
            if getattr(col_dim, "hidden", False):
                hidden_cols.append(col_letter)
        hidden_rows = []
        for row_idx, row_dim in ws.row_dimensions.items():
            if getattr(row_dim, "hidden", False):
                hidden_rows.append(row_idx)
        formulas = []
        colored_cells = []
        max_scan_rows = min(ws.max_row or 0, 200)
        max_scan_cols = min(ws.max_column or 0, 60)
        for r in range(1, max_scan_rows + 1):
            for c in range(1, max_scan_cols + 1):
                cell = ws.cell(row=r, column=c)
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    formulas.append((cell.coordinate, v[:80]))
                fill = cell.fill
                if fill is not None and fill.fgColor is not None:
                    rgb = getattr(fill.fgColor, "rgb", None)
                    if rgb and isinstance(rgb, str) and rgb not in ("00000000",) and not rgb.endswith("FFFFFF"):
                        if fill.patternType not in (None, "none"):
                            colored_cells.append((cell.coordinate, rgb, fill.patternType))
        sheet_state = ws.sheet_state
        info["sheets"][name] = {
            "dims": (ws.max_row, ws.max_column),
            "merged_ranges": merged,
            "hidden_cols": hidden_cols,
            "hidden_rows": hidden_rows,
            "formulas_sample": formulas[:20],
            "formulas_total": len(formulas),
            "colored_cells_sample": colored_cells[:30],
            "colored_cells_total": len(colored_cells),
            "sheet_state": sheet_state,
        }
    wb.close()
    return info


def detect_header_row(df_raw: pd.DataFrame, max_scan: int = 10) -> int:
    """Heuristic: header row is the first row where most cells are non-null
    short-ish strings AND distinct from each other."""
    best_row = 0
    best_score = -1.0
    n_rows = min(len(df_raw), max_scan)
    for r in range(n_rows):
        row = df_raw.iloc[r].tolist()
        non_null = [v for v in row if v is not None and not (isinstance(v, float) and pd.isna(v))]
        if not non_null:
            continue
        str_vals = [str(v) for v in non_null]
        density = len(non_null) / max(len(row), 1)
        uniq = len(set(str_vals)) / max(len(str_vals), 1)
        avg_len = sum(len(s) for s in str_vals) / max(len(str_vals), 1)
        len_penalty = 1.0 if avg_len <= 40 else 40.0 / avg_len
        numeric_frac = sum(looks_numeric(v) for v in non_null) / max(len(non_null), 1)
        score = density * uniq * len_penalty * (1 - numeric_frac)
        if score > best_score:
            best_score = score
            best_row = r
    return best_row


def explore_sheet(path: Path, sheet_name: str, structural: dict) -> str:
    out: list[str] = []
    out.append(f"## Sheet: `{sheet_name}`")
    out.append("")

    df_raw = pd.read_excel(path, sheet_name=sheet_name, header=None, dtype=object)
    n_rows_raw, n_cols_raw = df_raw.shape
    out.append(f"- Raw dimensions (header=None): **{n_rows_raw} rows × {n_cols_raw} cols**")
    sd = structural["sheets"].get(sheet_name, {})
    out.append(f"- openpyxl dimensions (max_row × max_col): **{sd.get('dims')}**")
    out.append(f"- Sheet visibility state: `{sd.get('sheet_state')}`")
    if sd.get("merged_ranges"):
        out.append(f"- Merged cell ranges ({len(sd['merged_ranges'])}): {sd['merged_ranges'][:20]}")
    else:
        out.append("- Merged cell ranges: none")
    if sd.get("hidden_cols"):
        out.append(f"- Hidden columns: {sd['hidden_cols']}")
    if sd.get("hidden_rows"):
        out.append(f"- Hidden rows: {sd['hidden_rows']}")
    if sd.get("formulas_total"):
        out.append(f"- Formulas: {sd['formulas_total']} (sample: {sd['formulas_sample'][:5]})")
    if sd.get("colored_cells_total"):
        out.append(f"- Cells with non-default fill colour: {sd['colored_cells_total']} (sample: {sd['colored_cells_sample'][:10]})")
    out.append("")

    out.append("### First 12 raw rows (header=None, repr() to expose hidden chars)")
    out.append("")
    out.append("```")
    preview_rows = min(12, n_rows_raw)
    for r in range(preview_rows):
        row_vals = [repr(v) for v in df_raw.iloc[r].tolist()]
        out.append(f"row {r}: " + " | ".join(row_vals))
    out.append("```")
    out.append("")

    header_row = detect_header_row(df_raw)
    out.append(f"### Auto-detected header row index: **{header_row}**")
    out.append("")

    if header_row >= n_rows_raw:
        out.append("> Header detection failed (file too small). Skipping per-column analysis.")
        return "\n".join(out)

    raw_headers = df_raw.iloc[header_row].tolist()
    out.append("### Raw column headers (repr — shows hidden whitespace, RTL marks, etc.)")
    out.append("")
    out.append("| col # | repr(header) | char inventory of suspicious chars |")
    out.append("|---:|---|---|")
    for i, h in enumerate(raw_headers):
        inv = char_inventory(h) if isinstance(h, str) else {}
        inv_s = ", ".join(f"{k}×{v}" for k, v in inv.items()) or "—"
        out.append(f"| {i} | `{md_escape(repr(h))}` | {md_escape(inv_s)} |")
    out.append("")

    df = df_raw.iloc[header_row + 1:].reset_index(drop=True)
    df.columns = [
        f"__col_{i}__" if (h is None or (isinstance(h, float) and pd.isna(h))) else str(h)
        for i, h in enumerate(raw_headers)
    ]
    df = df.dropna(how="all").reset_index(drop=True)
    out.append(f"### Data block (after header row, dropped all-empty rows): **{len(df)} rows**")
    out.append("")

    n_dup_full = df.duplicated().sum()
    out.append(f"- Full-row duplicates: **{n_dup_full}**")
    out.append("")

    out.append("### Per-column analysis")
    out.append("")
    for col in df.columns:
        s = df[col]
        out.append(f"#### Column `{col}`")
        out.append("")

        miss_counter: Counter[str] = Counter()
        non_missing = []
        for v in s:
            mflag, kind = is_missing(v)
            if mflag:
                miss_counter[kind or "unknown"] += 1
            else:
                non_missing.append(v)
        miss_total = sum(miss_counter.values())
        out.append(f"- Inferred pandas dtype: `{s.dtype}` | Python type histogram: " + repr(dict(Counter(type(v).__name__ for v in s))))
        out.append(f"- Missing breakdown ({miss_total}/{len(s)}): " + (", ".join(f"{k}={v}" for k, v in miss_counter.most_common()) or "no missing"))
        out.append(f"- Non-missing count: {len(non_missing)} | unique non-missing: {len(set(map(repr, non_missing)))}")

        sample_vals = non_missing[:5]
        out.append(f"- First 5 non-missing values (repr): " + ", ".join(repr(v) for v in sample_vals))

        try:
            extreme: list = []
            lengths = [(len(v) if isinstance(v, str) else -1, v) for v in non_missing]
            lengths.sort(reverse=True)
            extreme.extend(v for _, v in lengths[:3])
            numeric_vals = [float(str(v).replace(",", "")) for v in non_missing if looks_numeric(v)]
            if numeric_vals:
                extreme.append(("min_numeric", min(numeric_vals)))
                extreme.append(("max_numeric", max(numeric_vals)))
            out.append(f"- Extreme/longest values (repr): " + ", ".join(repr(v) for v in extreme[:8]))
        except Exception as e:
            out.append(f"- Extreme value scan errored: {e}")

        n_numeric_looking = sum(looks_numeric(v) for v in non_missing)
        n_text = len(non_missing) - n_numeric_looking
        out.append(f"- Looks-numeric: {n_numeric_looking} | non-numeric: {n_text}")
        if n_numeric_looking and n_numeric_looking == len(non_missing):
            try:
                nums = [float(str(v).replace(",", "")) for v in non_missing]
                out.append(f"  - numeric stats: min={min(nums)}, max={max(nums)}, mean={sum(nums)/len(nums):.4g}")
                stored_as_text = sum(1 for v in non_missing if isinstance(v, str))
                out.append(f"  - numeric values stored as TEXT: {stored_as_text}")
            except Exception:
                pass

        text_vals = [v for v in non_missing if isinstance(v, str)]
        if text_vals:
            freq = Counter(text_vals).most_common(20)
            out.append(f"- Unique text values: {len(set(text_vals))}")
            out.append(f"- Top 20 by frequency:")
            for v, c in freq:
                inv = char_inventory(v)
                inv_s = (" — chars: " + ", ".join(f"{k}×{n}" for k, n in inv.items())) if inv else ""
                out.append(f"    - `{md_escape(repr(v))}` × {c}{md_escape(inv_s)}")
            buckets: dict[str, list[str]] = defaultdict(list)
            for v in set(text_vals):
                buckets[normalize_arabic(v)].append(v)
            near_dups = {k: vs for k, vs in buckets.items() if len(vs) > 1}
            if near_dups:
                out.append(f"- **Near-duplicates** (same after NFKC + Arabic-norm + casefold + whitespace-collapse): {len(near_dups)} groups")
                for k, vs in list(near_dups.items())[:30]:
                    out.append(f"    - canonical `{md_escape(repr(k))}` ← " + ", ".join(f"`{md_escape(repr(v))}`" for v in vs))
            else:
                out.append("- Near-duplicates: none")
        out.append("")

    return "\n".join(out)


def main() -> None:
    if len(sys.argv) != 3:
        print("Usage: python explore.py <input.xlsx> <output.md>", file=sys.stderr)
        sys.exit(2)
    in_path = Path(sys.argv[1]).resolve()
    out_path = Path(sys.argv[2]).resolve()

    structural = explore_workbook_with_openpyxl(in_path)

    buf = io.StringIO()
    buf.write(f"# Exploration report: `{in_path}`\n\n")
    buf.write(f"Generated by `scripts/explore.py` — read-only.\n\n")
    buf.write(f"**File size:** {in_path.stat().st_size} bytes\n\n")
    buf.write(f"**Sheets ({len(structural['all_sheet_names'])}):** " + ", ".join(f"`{s}`" for s in structural["all_sheet_names"]) + "\n\n")
    buf.write("## Sheet summary table\n\n")
    buf.write("| sheet | dims (row×col) | state | merged | hidden cols | hidden rows | formulas | colored cells |\n")
    buf.write("|---|---|---|---:|---:|---:|---:|---:|\n")
    for name in structural["all_sheet_names"]:
        sd = structural["sheets"][name]
        buf.write(f"| `{name}` | {sd['dims']} | {sd['sheet_state']} | {len(sd['merged_ranges'])} | {len(sd['hidden_cols'])} | {len(sd['hidden_rows'])} | {sd['formulas_total']} | {sd['colored_cells_total']} |\n")
    buf.write("\n")

    for name in structural["all_sheet_names"]:
        buf.write(explore_sheet(in_path, name, structural))
        buf.write("\n\n")

    text = buf.getvalue()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(text, encoding="utf-8")
    sys.stdout.write(text)


if __name__ == "__main__":
    main()
