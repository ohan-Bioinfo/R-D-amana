"""Build an Excel reconciliation workbook: OUR cleaned numbers vs the OFFICIAL
Annual Report (MICRO), with blank YELLOW columns for Muhannad to fill in his
target / guidance so we can converge the pipeline to the official annual numbers.

Run:  microbiology/.venv/bin/python microbiology/scripts/build_reconciliation_workbook.py
Out:  microbiology/reports/annual_reconciliation_to_fill.xlsx
"""
from __future__ import annotations
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from build_dashboard_combined import derive_sector_5

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "reports" / "annual_reconciliation_to_fill.xlsx"

# ── OFFICIAL figures (MICRO), from Muhannad's Annual Report ──────────────────
# month → (samples, non-compliant, compliant)
OFF_2025 = {"Jan":(859,256,603),"Feb":(782,216,566),"Mar":(761,143,618),"Apr":(1380,421,959),
            "May":(1269,363,906),"June":(704,186,518),"July":(1130,315,815),"Aug":(1107,315,792),
            "Sep":(806,205,601),"Oct":(803,227,576),"Nov":(935,227,708),"Dec":(868,185,683)}
OFF_2024 = {"Jan":(653,220,433),"Feb":(641,260,381),"Mar":(501,142,359),"Apr":(348,128,220),
            "May":(574,150,424),"June":(402,157,245),"July":(1483,407,1076),"Aug":(1425,409,1016),
            "Sep":(504,134,370),"Oct":(966,271,695),"Nov":(777,232,545),"Dec":(834,199,635)}
OFF_TESTS = {2024:(None,None), 2025:(46309,4211)}   # (total tests, non-compliant tests) — 2024 not supplied
# per-test 2025 (test_en, total, invalid) ; compliant = total-invalid
OFF_PERTEST_2025 = [
    ("Aerobic plate count","العد الكلي للبكتيريا",6645,1514),
    ("Yeasts & Molds","الخمائر والاعفان",4561,736),
    ("Enterobacteriaceae","انتيروباكتريسي",3784,556),
    ("Staphylococcus aureus","استافيلوكوكس اورياس",7250,862),
    ("Coliforms","كوليفورم",778,86),
    ("Pseudomonas aeruginosa","سيدوموناس",332,20),
    ("E. coli","ايشيريشيا كولاي",7342,264),
    ("Bacillus cereus","باسيلس سيريس",1340,33),
    ("Salmonella","السالمونيلا",8305,140),
    ("Campylobacter jejuni","كامبيلوباكتر",365,0),
    ("Clostridium perfringens","كلوستريديوم بيرفرنجنز",1356,0),
    ("Clostridium botulinum","كلوستريديوم بوتولينوم",44,0),
    ("E. coli O157","ايشيريشيا كولاي O157",1816,0),
    ("Listeria monocytogenes","الليستيريا",2241,0),
    ("Vibrio parahaemolyticus","فيبريو",150,0),
]
# 2025 sectors (Municipalities sheet — COLLECTION basis, incl. chem; 17,648 total)
OFF_SECTORS_2025 = [("القطاع الأوسط",6790),("قطاع الشرق",5405),("قطاع الشمال",3133),
                    ("قطاع الغرب",1230),("قطاع الجنوب",995),("العينات الخاصة",95)]

MONTHS = ["Jan","Feb","Mar","Apr","May","June","July","Aug","Sep","Oct","Nov","Dec"]
MNUM = {"01":"Jan","02":"Feb","03":"Mar","04":"Apr","05":"May","06":"June",
        "07":"July","08":"Aug","09":"Sep","10":"Oct","11":"Nov","12":"Dec"}

# ── styles ──────────────────────────────────────────────────────────────────
HDR = PatternFill("solid", fgColor="1C2742")
FILLCOL = PatternFill("solid", fgColor="FFF3C7")     # yellow = you fill
FILLHDR = PatternFill("solid", fgColor="F59E0B")
TOTROW = PatternFill("solid", fgColor="EEF2FF")
WHITEB = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
THIN = Side(style="thin", color="D0D5DD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load():
    out = {}
    for y in (2024, 2025):
        d = pd.read_parquet(ROOT / "cleaned" / f"data{y}.parquet")
        d["mm"] = d["year_month"].astype(str).str[-2:].map(MNUM)
        out[y] = d
    return out


def our_month(d, m):
    s = d[d["mm"] == m]
    n = len(s)
    nc = int((s["is_failure"] == True).sum())  # noqa: E712
    return n, nc, n - nc


def our_pertest(d2025, ar):
    n = 0
    for lst in d2025.get("invalid_tests", pd.Series([], dtype=object)):
        if lst is None:
            continue
        if ar in set(lst):
            n += 1
    return n


def our_sectors(d2025):
    # Use the dashboard's exact derivation (municipality → 5 sectors + Special).
    from collections import Counter
    c = Counter()
    for r in d2025[["municipality", "sector"]].itertuples(index=False):
        s = derive_sector_5(r.municipality, r.sector)
        if s:
            c[s] += 1
    return dict(c)   # keys: East / North / West / Central / South / Special


def style_header(ws, row, fills):
    for c, fill in zip(ws[row], fills):
        c.fill = fill
        c.font = WHITEB if fill is HDR else (Font(bold=True) if fill is FILLHDR else BOLD)
        c.alignment = CEN
        c.border = BORDER


def put(ws, r, values, fillcols=(), border=True, totrow=False):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=r, column=i, value=v)
        if border:
            c.border = BORDER
        if totrow:
            c.fill = TOTROW
            c.font = BOLD
        if (i - 1) in fillcols and not totrow:
            c.fill = FILLCOL


def build():
    data = load()
    wb = Workbook()

    # ── README ──
    ws = wb.active; ws.title = "README"
    ws["A1"] = "Annual reconciliation — reach the official MICRO numbers"
    ws["A1"].font = Font(bold=True, size=14)
    notes = [
        "",
        "Purpose: converge our cleaned pipeline to your official Annual Report (MICRO) numbers.",
        "",
        "Each sheet compares OUR current numbers vs your OFFICIAL numbers, with the gap (Δ).",
        "YELLOW columns are for YOU — fill in your target and how we should reconcile.",
        "",
        "Sheets:",
        "  • Samples_2024 / Samples_2025 — monthly samples / non-compliant / compliant.",
        "  • Totals — year totals + compliance % + tests.",
        "  • PerTest_2025 — invalid per organism (official has the denominator; we only have invalid).",
        "  • Sectors_2025 — sample split by sector (report is COLLECTION basis = 17,648, incl. chem).",
        "",
        "In the yellow 'Guidance' columns, tell me things like:",
        "  – 'these N are duplicates, keep them'   – 'we dropped month X rows, recover them'",
        "  – 'date basis is receive-date, re-key'  – 'exclude swabs'   – 'this test total is right'",
        "",
        "Known gaps to explain: 2024 samples 8,094 (ours) vs 9,108 (official) = -1,014;",
        "2025 samples 11,564 (ours) vs 11,404 (official) = +160.",
    ]
    for i, t in enumerate(notes, 3):
        ws.cell(row=i, column=1, value=t)
    ws.column_dimensions["A"].width = 100

    # ── Samples_YYYY ──
    for y, off in ((2024, OFF_2024), (2025, OFF_2025)):
        ws = wb.create_sheet(f"Samples_{y}")
        headers = ["Month",
                   "Official samples", "Our samples", "Δ samples",
                   "Official non-comp", "Our non-comp", "Δ non-comp",
                   "Official compliant", "Our compliant", "Δ compliant",
                   "YOUR target samples", "Guidance / how to reconcile"]
        ws.append(headers)
        fills = [HDR]*10 + [FILLHDR]*2
        style_header(ws, 1, fills)
        d = data[y]
        tos = ton = toc = ros = ron = roc = 0
        for m in MONTHS:
            osamp, onc, ocomp = off[m]
            rsamp, rnc, rcomp = our_month(d, m)
            put(ws, ws.max_row+1,
                [m, osamp, rsamp, rsamp-osamp, onc, rnc, rnc-onc, ocomp, rcomp, rcomp-ocomp, None, None],
                fillcols=(10, 11))
            tos += osamp; ton += onc; toc += ocomp; ros += rsamp; ron += rnc; roc += rcomp
        put(ws, ws.max_row+1,
            ["TOTAL", tos, ros, ros-tos, ton, ron, ron-ton, toc, roc, roc-toc, None, None], totrow=True)
        for col, w in zip("ABCDEFGHIJKL", [8,15,12,10,15,12,11,15,13,11,18,55]):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"

    # ── Totals ──
    ws = wb.create_sheet("Totals")
    ws.append(["Year", "Metric", "Official", "Our", "Δ", "YOUR target", "Guidance"])
    style_header(ws, 1, [HDR]*5 + [FILLHDR]*2)
    o24s = sum(v[0] for v in OFF_2024.values()); o24c = sum(v[2] for v in OFF_2024.values())
    o25s = sum(v[0] for v in OFF_2025.values()); o25c = sum(v[2] for v in OFF_2025.values())
    for y, osamp, ocomp in ((2024, o24s, o24c), (2025, o25s, o25c)):
        d = data[y]; rs = len(d); rc = int((d["is_failure"] != True).sum())  # noqa: E712
        put(ws, ws.max_row+1, [y, "Samples", osamp, rs, rs-osamp, None, None], fillcols=(5, 6))
        put(ws, ws.max_row+1, [y, "Compliant", ocomp, rc, rc-ocomp, None, None], fillcols=(5, 6))
        put(ws, ws.max_row+1, [y, "Non-compliant", osamp-ocomp, rs-rc, (rs-rc)-(osamp-ocomp), None, None], fillcols=(5, 6))
        put(ws, ws.max_row+1, [y, "Compliance %", round(100*ocomp/osamp, 2), round(100*rc/rs, 2),
                               round(100*rc/rs-100*ocomp/osamp, 2), None, None], fillcols=(5, 6))
        ot, onct = OFF_TESTS[y]
        put(ws, ws.max_row+1, [y, "Total tests", ot, "(no per-test source)" if y == 2025 else None, None, None, None], fillcols=(5, 6))
        put(ws, ws.max_row+1, [y, "Non-comp tests", onct, None, None, None, None], fillcols=(5, 6))
    for col, w in zip("ABCDEFG", [7, 16, 12, 22, 10, 14, 55]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    # ── PerTest_2025 ──
    ws = wb.create_sheet("PerTest_2025")
    ws.append(["Test (EN)", "Test (AR)", "Official total", "Official invalid", "Official rate %",
               "Our invalid (samples)", "YOUR target invalid", "Guidance"])
    style_header(ws, 1, [HDR]*6 + [FILLHDR]*2)
    d25 = data[2025]
    for en, ar, tot, inv in OFF_PERTEST_2025:
        ours = our_pertest(d25, ar)
        put(ws, ws.max_row+1, [en, ar, tot, inv, round(100*inv/tot, 1) if tot else 0, ours, None, None],
            fillcols=(6, 7))
    put(ws, ws.max_row+1, ["TOTAL", "", 46309, 4211, round(100*4211/46309, 1), None, None, None], totrow=True)
    for col, w in zip("ABCDEFGH", [24, 22, 13, 14, 13, 18, 18, 45]):
        ws.column_dimensions[col].width = w
    ws["B1"].alignment = CEN
    ws.freeze_panes = "A2"

    # ── Sectors_2025 ──
    ws = wb.create_sheet("Sectors_2025")
    ws.append(["Sector (AR, report)", "Official samples (collection)", "Our samples (micro)", "Guidance"])
    style_header(ws, 1, [HDR]*3 + [FILLHDR])
    osec = {"القطاع الأوسط":"Central","قطاع الشرق":"East","قطاع الشمال":"North",
            "قطاع الغرب":"West","قطاع الجنوب":"South","العينات الخاصة":"Special"}
    oursec = our_sectors(d25)
    for ar, n in OFF_SECTORS_2025:
        en = osec.get(ar)
        put(ws, ws.max_row+1, [ar, n, int(oursec.get(en, 0)), None], fillcols=(3,))
    put(ws, ws.max_row+1, ["TOTAL", 17648, int(sum(oursec.values())), None], totrow=True)
    ws.cell(row=ws.max_row+1, column=1,
            value="Note: the report's sector split is COLLECTION basis (17,648 incl. chem), "
                  "not the 11,404 micro-tested total — so it will not sum to the micro samples.")
    for col, w in zip("ABCD", [22, 28, 20, 55]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    build()
