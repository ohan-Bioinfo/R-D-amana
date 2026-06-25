"""Build a packaged summary folder containing:

  summary/
    data2025_vs_annual_report.md      — narrative comparison
    data2025_vs_annual_report.csv     — flat metric table
    figures/
      01_totals_comparison.png        — bar chart of report vs ours (totals)
      02_monthly_samples.png          — grouped bar chart by month
      03_monthly_compliance.png       — line chart of compliance % by month
      04_per_test_invalid.png         — grouped horizontal bars per test
      00_summary_dashboard.png        — multi-panel composite

Reads:
  cleaned/data2025.parquet
  2025-original/Annual Report 2025.xlsx
"""
from __future__ import annotations

import shutil
from collections import Counter
from pathlib import Path

import matplotlib.pyplot as plt
import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
PARQUET = ROOT / "cleaned" / "data2025.parquet"
REPORT_XLSX = ROOT / "2025-original" / "Annual Report 2025.xlsx"
OUT_ROOT = ROOT / "summary"
FIG_DIR = OUT_ROOT / "figures"

EXISTING_MD = ROOT / "reports" / "data2025_vs_annual_report.md"
EXISTING_CSV = ROOT / "reports" / "data2025_vs_annual_report.csv"

MONTH_MAP = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"June":6,
             "July":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}
MONTHS_SHORT = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

TEST_NAME_MAP = {
    "العد الكلي للبكتيريا":   "Aerobic plate count",
    "استافيلوكوكس اورياس":    "Staphylococcus aureas",
    "الخمائر والاعفان":       "Yeasts & Molds",
    "انتيروباكتريسي":         "Enterobacteriaceae",
    "ايشيريشيا كولاي":         "E. coli",
    "السالمونيلا":             "Salmonella",
    "كوليفورم":                "Coliforms",
    "باسيلس سيريس":            "Bacillus cereus",
    "سيدوموناس":               "Pseudomonas aeruginosa",
}

# ── colours ──────────────────────────────────────────
COL_REPORT = "#1f6feb"
COL_OURS   = "#f59e0b"
COL_ACCENT = "#9a6700"
plt.rcParams.update({
    "font.family": ["DejaVu Sans"],
    "axes.titlesize": 12,
    "axes.titleweight": "bold",
    "axes.spines.top": False,
    "axes.spines.right": False,
    "axes.grid": True,
    "grid.alpha": 0.25,
    "grid.linestyle": "--",
    "figure.facecolor": "white",
})


def collect_data():
    df = pd.read_parquet(PARQUET)
    df["m"] = df["sampling_date"].dt.month
    wb = openpyxl.load_workbook(REPORT_XLSX, data_only=True)

    # Monthly
    ws = wb["Compliance rate"]
    rep = []
    for r in range(6, 18):
        name = ws.cell(row=r, column=2).value
        if name not in MONTH_MAP:
            continue
        rep.append({
            "month_num": MONTH_MAP[name],
            "report_total": int(ws.cell(row=r, column=3).value),
            "report_valid": int(ws.cell(row=r, column=4).value),
        })
    rep_df = pd.DataFrame(rep)
    rep_df["report_invalid"] = rep_df["report_total"] - rep_df["report_valid"]
    rep_df["report_compliance"] = 100 * rep_df["report_valid"] / rep_df["report_total"]

    ours = df.groupby("m").agg(
        ours_total=("is_valid", "size"),
        ours_valid=("is_failure", lambda s: int((s == False).sum())),
    ).reset_index().rename(columns={"m": "month_num"})
    ours["ours_invalid"] = ours["ours_total"] - ours["ours_valid"]
    ours["ours_compliance"] = 100 * ours["ours_valid"] / ours["ours_total"]

    monthly = rep_df.merge(ours, on="month_num", how="outer").sort_values("month_num")

    # Per-test
    ws_t = wb["Test"]
    rep_tests = {}
    for r in range(15, 35):
        name = ws_t.cell(row=r, column=2).value
        total = ws_t.cell(row=r, column=3).value
        invalid = ws_t.cell(row=r, column=5).value
        if name and invalid is not None:
            rep_tests[name] = {"total": int(total), "invalid": int(invalid)}

    failed = []
    for lst in df["invalid_tests"]:
        if lst is not None:
            failed.extend(lst)
    ours_inv = Counter(failed)

    test_rows = []
    for ar, en in TEST_NAME_MAP.items():
        rep = rep_tests.get(en, {"total": None, "invalid": None})
        test_rows.append({
            "test": en, "test_ar": ar,
            "report_invalid": rep["invalid"], "ours_invalid": int(ours_inv.get(ar, 0)),
        })
    test_df = pd.DataFrame(test_rows)

    totals = {
        "report_total": int(rep_df["report_total"].sum()),
        "report_valid": int(rep_df["report_valid"].sum()),
        "ours_total":   int(ours["ours_total"].sum()),
        "ours_valid":   int(ours["ours_valid"].sum()),
    }
    return monthly, test_df, totals


# ── individual figures ──────────────────────────────
def fig_totals(totals, path):
    fig, ax = plt.subplots(figsize=(7, 4.2))
    cats = ["Total samples", "Compliant", "Non-compliant"]
    rep_vals = [totals["report_total"], totals["report_valid"],
                totals["report_total"] - totals["report_valid"]]
    our_vals = [totals["ours_total"], totals["ours_valid"],
                totals["ours_total"] - totals["ours_valid"]]
    x = range(len(cats))
    w = 0.38
    ax.bar([i - w/2 for i in x], rep_vals, w, label="Annual Report", color=COL_REPORT)
    ax.bar([i + w/2 for i in x], our_vals, w, label="Our parquet", color=COL_OURS)
    for i, (a, b) in enumerate(zip(rep_vals, our_vals)):
        ax.text(i - w/2, a, f"{a:,}", ha="center", va="bottom", fontsize=9)
        ax.text(i + w/2, b, f"{b:,}", ha="center", va="bottom", fontsize=9)
    ax.set_xticks(list(x))
    ax.set_xticklabels(cats)
    ax.set_ylabel("samples")
    ax.set_title("Totals comparison — Annual Report (MICRO) vs our parquet")
    ax.legend(loc="upper right", frameon=False)
    fig.tight_layout()
    fig.savefig(path, dpi=130)
    plt.close(fig)


def fig_monthly_samples(monthly, path):
    fig, ax = plt.subplots(figsize=(11, 4.5))
    x = range(len(monthly))
    w = 0.4
    ax.bar([i - w/2 for i in x], monthly["report_total"], w, label="Annual Report", color=COL_REPORT)
    ax.bar([i + w/2 for i in x], monthly["ours_total"],   w, label="Our parquet",   color=COL_OURS)
    ax.set_xticks(list(x))
    ax.set_xticklabels(MONTHS_SHORT)
    ax.set_ylabel("samples")
    ax.set_title("Monthly sample volume — by sampling date (ours) vs report month (report)")
    ax.legend(loc="upper right", frameon=False)
    # delta annotations
    for i, (a, b) in enumerate(zip(monthly["report_total"], monthly["ours_total"])):
        d = int(b - a)
        ax.text(i, max(a, b) + 30, f"{d:+d}", ha="center", fontsize=8,
                color="#cf222e" if abs(d) > 100 else "#57606a")
    fig.tight_layout()
    fig.savefig(path, dpi=130)
    plt.close(fig)


def fig_monthly_compliance(monthly, path):
    fig, ax = plt.subplots(figsize=(11, 4.5))
    x = range(len(monthly))
    ax.plot(x, monthly["report_compliance"], marker="o", linewidth=2.2,
            label="Annual Report", color=COL_REPORT)
    ax.plot(x, monthly["ours_compliance"], marker="s", linewidth=2.2,
            label="Our parquet", color=COL_OURS)
    ax.set_xticks(list(x))
    ax.set_xticklabels(MONTHS_SHORT)
    ax.set_ylabel("compliance %")
    ax.set_ylim(60, 90)
    ax.set_title("Monthly compliance rate — by month")
    ax.legend(loc="lower right", frameon=False)
    for i, (a, b) in enumerate(zip(monthly["report_compliance"], monthly["ours_compliance"])):
        ax.annotate(f"{a:.1f}", (i, a), textcoords="offset points", xytext=(0, 10),
                    ha="center", fontsize=8, color=COL_REPORT)
        ax.annotate(f"{b:.1f}", (i, b), textcoords="offset points", xytext=(0, -14),
                    ha="center", fontsize=8, color=COL_ACCENT)
    fig.tight_layout()
    fig.savefig(path, dpi=130)
    plt.close(fig)


def fig_per_test(test_df, path):
    fig, ax = plt.subplots(figsize=(10, 5.2))
    df = test_df.dropna(subset=["report_invalid"]).copy()
    df = df.sort_values("report_invalid")
    y = range(len(df))
    h = 0.4
    ax.barh([i - h/2 for i in y], df["report_invalid"], h, label="Annual Report", color=COL_REPORT)
    ax.barh([i + h/2 for i in y], df["ours_invalid"],   h, label="Our parquet",   color=COL_OURS)
    ax.set_yticks(list(y))
    ax.set_yticklabels(df["test"])
    ax.set_xlabel("non-compliant test count")
    ax.set_title("Non-compliant test counts — Annual Report vs ours")
    ax.legend(loc="lower right", frameon=False)
    for i, (a, b) in enumerate(zip(df["report_invalid"], df["ours_invalid"])):
        ax.text(a, i - h/2, f" {int(a)}", va="center", fontsize=8)
        ax.text(b, i + h/2, f" {int(b)}", va="center", fontsize=8)
    fig.tight_layout()
    fig.savefig(path, dpi=130)
    plt.close(fig)


def fig_dashboard(monthly, test_df, totals, path):
    fig = plt.figure(figsize=(15, 11))
    gs = fig.add_gridspec(3, 2, hspace=0.45, wspace=0.25)

    # Panel 1 — totals (top-left)
    ax = fig.add_subplot(gs[0, 0])
    cats = ["Total", "Compliant", "Non-compliant"]
    rep_vals = [totals["report_total"], totals["report_valid"],
                totals["report_total"] - totals["report_valid"]]
    our_vals = [totals["ours_total"], totals["ours_valid"],
                totals["ours_total"] - totals["ours_valid"]]
    x = range(len(cats))
    w = 0.38
    ax.bar([i - w/2 for i in x], rep_vals, w, label="Report", color=COL_REPORT)
    ax.bar([i + w/2 for i in x], our_vals, w, label="Ours",   color=COL_OURS)
    for i, (a, b) in enumerate(zip(rep_vals, our_vals)):
        ax.text(i - w/2, a, f"{a:,}", ha="center", va="bottom", fontsize=9)
        ax.text(i + w/2, b, f"{b:,}", ha="center", va="bottom", fontsize=9)
    ax.set_xticks(list(x)); ax.set_xticklabels(cats)
    ax.set_title("Totals (MICRO)")
    ax.legend(frameon=False)

    # Panel 2 — overall compliance gauge style (top-right)
    ax = fig.add_subplot(gs[0, 1])
    rep_rate = 100 * totals["report_valid"] / totals["report_total"]
    our_rate = 100 * totals["ours_valid"] / totals["ours_total"]
    bars = ax.barh(["Report", "Ours"], [rep_rate, our_rate],
                    color=[COL_REPORT, COL_OURS])
    for bar, v in zip(bars, [rep_rate, our_rate]):
        ax.text(v + 0.4, bar.get_y() + bar.get_height()/2, f"{v:.2f}%",
                va="center", fontsize=11, fontweight="bold")
    ax.set_xlim(0, 100)
    ax.set_xlabel("compliance %")
    ax.set_title("Overall compliance rate")
    ax.axvline(rep_rate, color=COL_REPORT, linestyle=":", alpha=0.4)

    # Panel 3 — monthly sample volume (middle row, full width)
    ax = fig.add_subplot(gs[1, :])
    x = range(len(monthly))
    w = 0.4
    ax.bar([i - w/2 for i in x], monthly["report_total"], w, label="Report", color=COL_REPORT)
    ax.bar([i + w/2 for i in x], monthly["ours_total"],   w, label="Ours",   color=COL_OURS)
    ax.set_xticks(list(x)); ax.set_xticklabels(MONTHS_SHORT)
    ax.set_ylabel("samples")
    ax.set_title("Monthly sample volume")
    ax.legend(frameon=False)
    for i, (a, b) in enumerate(zip(monthly["report_total"], monthly["ours_total"])):
        d = int(b - a)
        ax.text(i, max(a, b) + 30, f"{d:+d}", ha="center", fontsize=8,
                color="#cf222e" if abs(d) > 100 else "#57606a")

    # Panel 4 — monthly compliance line (bottom-left)
    ax = fig.add_subplot(gs[2, 0])
    x = range(len(monthly))
    ax.plot(x, monthly["report_compliance"], marker="o", linewidth=2,
            label="Report", color=COL_REPORT)
    ax.plot(x, monthly["ours_compliance"], marker="s", linewidth=2,
            label="Ours", color=COL_OURS)
    ax.set_xticks(list(x)); ax.set_xticklabels(MONTHS_SHORT)
    ax.set_ylim(60, 90)
    ax.set_title("Monthly compliance rate")
    ax.set_ylabel("%")
    ax.legend(frameon=False)

    # Panel 5 — per-test invalid (bottom-right)
    ax = fig.add_subplot(gs[2, 1])
    sub = test_df.dropna(subset=["report_invalid"]).sort_values("report_invalid")
    y = range(len(sub))
    h = 0.4
    ax.barh([i - h/2 for i in y], sub["report_invalid"], h, label="Report", color=COL_REPORT)
    ax.barh([i + h/2 for i in y], sub["ours_invalid"],   h, label="Ours",   color=COL_OURS)
    ax.set_yticks(list(y)); ax.set_yticklabels(sub["test"], fontsize=9)
    ax.set_xlabel("non-compliant test count")
    ax.set_title("Non-compliant tests")
    ax.legend(frameon=False, loc="lower right")

    fig.suptitle("Annual Report 2025 vs cleaned/data2025.parquet  ·  MICRO stream comparison",
                  fontsize=14, fontweight="bold", y=0.995)
    fig.savefig(path, dpi=130, bbox_inches="tight")
    plt.close(fig)


def main():
    OUT_ROOT.mkdir(parents=True, exist_ok=True)
    FIG_DIR.mkdir(parents=True, exist_ok=True)

    # Copy text reports into summary/
    if EXISTING_MD.exists():
        shutil.copy2(EXISTING_MD, OUT_ROOT / EXISTING_MD.name)
    if EXISTING_CSV.exists():
        shutil.copy2(EXISTING_CSV, OUT_ROOT / EXISTING_CSV.name)

    monthly, test_df, totals = collect_data()

    fig_totals(totals,                          FIG_DIR / "01_totals_comparison.png")
    fig_monthly_samples(monthly,                FIG_DIR / "02_monthly_samples.png")
    fig_monthly_compliance(monthly,             FIG_DIR / "03_monthly_compliance.png")
    fig_per_test(test_df,                       FIG_DIR / "04_per_test_invalid.png")
    fig_dashboard(monthly, test_df, totals,     FIG_DIR / "00_summary_dashboard.png")

    print(f"summary folder: {OUT_ROOT.relative_to(ROOT)}/")
    for p in sorted(OUT_ROOT.rglob("*")):
        if p.is_file():
            kb = p.stat().st_size / 1024
            print(f"  {p.relative_to(OUT_ROOT)}   ({kb:.0f} KB)")


if __name__ == "__main__":
    main()
