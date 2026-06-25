"""Phase 3 cleaner — applies schemas/lab_data_v1.yaml to a single .xlsx file.

Public API:
    clean_file(path) -> pd.DataFrame
        Returns the canonical long-format DataFrame.

    clean_file_with_audit(path) -> tuple[pd.DataFrame, dict]
        Same plus an audit dict (changes log, drops, flags) for diff reports.

CLI:
    python clean_file.py <input.xlsx> <output.parquet> <diff.md>

Determinism: same input bytes -> same output (sort order fixed, no set/dict
iteration order in the output, all transforms are pure).

Idempotency: running the cleaner twice on the same input produces identical
output. The cleaner never writes back to the source xlsx.

Failure mode: raises SchemaValidationError on any condition that the schema
marks fail_on. Conditions in warn_on become flags in data_quality_flags but do
NOT raise.
"""
from __future__ import annotations

import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

import pandas as pd
import yaml
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Constants & schema location
# ---------------------------------------------------------------------------
SCHEMA_PATH = Path(__file__).resolve().parent.parent / "schemas" / "lab_data_v1.yaml"

ZERO_WIDTH = "​‌‍﻿"
RTL_MARKS = "‎‏‪‫‬‭‮⁦⁧⁨⁩"
NBSP = " "

ARABIC_INDIC_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")
EASTERN_ARABIC_DIGITS = str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789")


class SchemaValidationError(Exception):
    pass


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def load_schema() -> dict:
    with SCHEMA_PATH.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def normalise_text(v):
    """NFKC + remove invisible marks + collapse whitespace.
    Preserves Arabic diacritics and letter forms (those carry meaning).
    """
    if v is None:
        return None
    if not isinstance(v, str):
        return v
    s = unicodedata.normalize("NFKC", v)
    for ch in ZERO_WIDTH + RTL_MARKS:
        s = s.replace(ch, "")
    s = s.replace(NBSP, " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s if s != "" else None


def excel_col_letter_to_index(letter: str) -> int:
    """A -> 1, B -> 2, ..., Z -> 26, AA -> 27. 1-indexed."""
    out = 0
    for ch in letter.upper():
        out = out * 26 + (ord(ch) - ord("A") + 1)
    return out


def parse_date_from_cell(cell_value) -> date | None:
    """Date is stored in cell E4 as: 'نتائج يوم الاثنين 01/01/2024'.
    Capture the dd/mm/yyyy.
    """
    if cell_value is None:
        return None
    if isinstance(cell_value, datetime):
        return cell_value.date()
    if isinstance(cell_value, date):
        return cell_value
    if not isinstance(cell_value, str):
        return None
    m = re.search(r"(\d{2})/(\d{2})/(\d{4})", cell_value)
    if not m:
        return None
    dd, mm, yyyy = m.groups()
    return date(int(yyyy), int(mm), int(dd))


def parse_date_from_filename(name: str) -> date | None:
    m = re.match(r"^(\d{2})(\d{2})(\d{2})[a-z]?\.xlsx$", name)
    if not m:
        return None
    dd, mm, yy = m.groups()
    return date(2000 + int(yy), int(mm), int(dd))


def parse_numeric_best_effort(v) -> tuple[float | None, list[str]]:
    """Try to coerce a value to float. Returns (value, prefix_flags).
    Prefix flags note things like '<', '>'.
    """
    flags: list[str] = []
    if v is None:
        return None, flags
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v), flags
    if not isinstance(v, str):
        return None, flags
    s = v.strip()
    if s == "":
        return None, flags
    s = s.translate(ARABIC_INDIC_DIGITS).translate(EASTERN_ARABIC_DIGITS)
    s = s.replace("٬", "").replace("٫", ".").replace(",", "")
    for prefix in ("<=", ">=", "≤", "≥", "<", ">"):
        if s.startswith(prefix):
            flags.append(f"comparison_prefix:{prefix}")
            s = s[len(prefix):].strip()
            break
    s = re.sub(r"[^\d.\-eE]", "", s)
    if s in ("", "-", ".", "-.", "."):
        return None, flags
    try:
        return float(s), flags
    except ValueError:
        return None, flags


def match_final_report_code(text: str | None, code_map: dict) -> str | None:
    if text is None:
        return None
    for code, rule in code_map.items():
        mt = rule.get("match_type")
        if mt == "fallback":
            continue
        needles = rule.get("needles", [])
        not_contains = rule.get("not_contains", [])
        if mt == "contains":
            if any(n in text for n in needles) and not any(n in text for n in not_contains):
                return code
        elif mt == "contains_all":
            if all(n in text for n in needles) and not any(n in text for n in not_contains):
                return code
    fb = next((c for c, r in code_map.items() if r.get("match_type") == "fallback"), None)
    return fb


# ---------------------------------------------------------------------------
# Workbook reading with merge-fill
# ---------------------------------------------------------------------------
def read_with_merge_fill(path: Path, header_row_excel: int) -> tuple[list[list], date | None, str]:
    """Read the worksheet into a 2-D list of values, after materialising
    every merged-cell range below the header into its top-left value.

    Merges that overlap the header row are skipped (they would otherwise
    propagate header text into the first data row).
    """
    wb = load_workbook(path, data_only=True, read_only=False, keep_links=False)
    ws = wb.active
    sheet_name = ws.title

    # Date extraction from cell E4 (column 5, row 4).
    date_cell = ws.cell(row=4, column=excel_col_letter_to_index("E")).value
    sheet_date = parse_date_from_cell(date_cell)
    if sheet_date is None:
        sheet_date = parse_date_from_filename(path.name)

    n_rows = ws.max_row or 0
    n_cols = ws.max_column or 0

    # Build a 2-D list (1-indexed rows/cols match Excel) of cell values.
    grid: list[list] = [[None] * (n_cols + 1) for _ in range(n_rows + 1)]
    for row in ws.iter_rows(min_row=1, max_row=n_rows, max_col=n_cols, values_only=False):
        for cell in row:
            grid[cell.row][cell.column] = cell.value

    # Apply merge-fill: for each merged range strictly below the header row,
    # write the top-left value into every covered cell.
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row <= header_row_excel:
            continue
        v = grid[mr.min_row][mr.min_col]
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                grid[r][c] = v

    wb.close()
    return grid, sheet_date, sheet_name


# ---------------------------------------------------------------------------
# Main cleaner
# ---------------------------------------------------------------------------
def clean_file_with_audit(path: Path) -> tuple[pd.DataFrame, dict]:
    schema = load_schema()
    audit: dict = {
        "source_file": str(path),
        "schema_name": schema["schema_name"],
        "schema_version": schema["schema_version"],
        "rows_in_raw": 0,
        "rows_after_dropna": 0,
        "rows_out": 0,
        "drops": [],          # list of {row_excel, reason}
        "changes": [],        # list of {row_excel, column, before, after, reason}
        "flags": [],          # list of {row_excel, flag, value}
        "warnings": [],
        "sequence_gaps_s_no": [],
        "sheet_date": None,
        "sheet_name": None,
    }

    header_row_excel = schema["workbook"]["header_row_excel"]
    data_starts_row_excel = schema["workbook"]["data_starts_row_excel"]
    use_cols = [excel_col_letter_to_index(c) for c in
                ["E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R"]]

    grid, sheet_date, sheet_name = read_with_merge_fill(path, header_row_excel)
    audit["sheet_date"] = sheet_date.isoformat() if sheet_date else None
    audit["sheet_name"] = sheet_name

    if sheet_date is None:
        raise SchemaValidationError(
            f"sheet_date_missing: could not extract date from cell E4 or filename {path.name}"
        )
    if sheet_name != schema["workbook"]["expected_sheet_name"]:
        audit["warnings"].append(f"unexpected_sheet_name: {sheet_name!r} (expected {schema['workbook']['expected_sheet_name']!r})")

    # Pull out the headers from header_row_excel and build a per-column index.
    raw_headers = [grid[header_row_excel][c] for c in use_cols]
    expected = [c["raw_name"] for c in schema["raw_columns"]]
    if raw_headers != expected:
        diffs = [(i, expected[i], raw_headers[i]) for i in range(len(expected)) if expected[i] != raw_headers[i]]
        raise SchemaValidationError(
            f"header_mismatch: column headers do not match schema. Diffs (idx, expected, got): {diffs}"
        )

    # Build the data frame from rows below the header.
    data_rows = []
    for r in range(data_starts_row_excel, len(grid)):
        row_vals = [grid[r][c] for c in use_cols]
        data_rows.append((r, row_vals))

    audit["rows_in_raw"] = len(data_rows)

    # Drop rows where all of the critical columns are empty.
    critical_idx = {expected.index(name) for name in
                    ["M.S.No", "S.No", "GSO code", "Sample Name", "Test", "Results"]}

    kept_rows = []
    for r_excel, vals in data_rows:
        all_critical_empty = all(
            (vals[i] is None) or (isinstance(vals[i], str) and vals[i].strip() == "")
            for i in critical_idx
        )
        if all_critical_empty:
            audit["drops"].append({"row_excel": r_excel, "reason": "all_critical_columns_empty"})
        else:
            kept_rows.append((r_excel, vals))
    audit["rows_after_dropna"] = len(kept_rows)

    # Materialise into a DataFrame with the raw column names, plus row_excel.
    df = pd.DataFrame([vals for _, vals in kept_rows], columns=expected)
    df.insert(0, "row_excel", [r for r, _ in kept_rows])

    # ---- Step 06: text normalisation (track changes per cell) ----
    text_columns = ["GSO code", "Sample Name", "Restaurant name",
                    "10 1", "10 2", "10 3", "Test",
                    "GSO 1016 Limits", "Results", "Validity", "Final Report",
                    "Barcodes"]
    for col in text_columns:
        for i in df.index:
            before = df.at[i, col]
            after = normalise_text(before) if isinstance(before, str) else before
            if before != after and not (before is None and after is None):
                df.at[i, col] = after
                audit["changes"].append({
                    "row_excel": int(df.at[i, "row_excel"]),
                    "column": col,
                    "before": before,
                    "after": after,
                    "reason": "text_normalise",
                })

    # ---- Step 07: value mappings ----
    test_col_def = next(c for c in schema["output_columns"] if c["name"] == "test")
    test_mapping: dict = test_col_def["mapping"]
    test_canonical: set = set(test_col_def["canonical_values"])
    for i in df.index:
        before = df.at[i, "Test"]
        if isinstance(before, str) and before in test_mapping:
            after = test_mapping[before]
            df.at[i, "Test"] = after
            audit["changes"].append({
                "row_excel": int(df.at[i, "row_excel"]),
                "column": "Test",
                "before": before,
                "after": after,
                "reason": "test_canonicalisation",
            })

    validity_col_def = next(c for c in schema["output_columns"] if c["name"] == "validity")
    validity_mapping: dict = validity_col_def["mapping"]

    # ---- Step 08: GSO 'ISO' special value ----
    for i in df.index:
        before = df.at[i, "GSO code"]
        if isinstance(before, str) and before.strip() == "ISO":
            df.at[i, "GSO code"] = None
            audit["changes"].append({
                "row_excel": int(df.at[i, "row_excel"]),
                "column": "GSO code",
                "before": before,
                "after": None,
                "reason": "gso_iso_placeholder_to_null",
            })
            audit["flags"].append({
                "row_excel": int(df.at[i, "row_excel"]),
                "flag": "gso_code_was_iso_placeholder",
                "value": before,
            })

    # ---- Step 09: M.S.No 4223 typo flag ----
    # Use Series iteration (yields Python ints) rather than df.at (yields np.int64
    # which fails isinstance checks in Python 3).
    for r_excel, v in zip(df["row_excel"], df["M.S.No"]):
        try:
            ival = int(v) if v is not None and not (isinstance(v, float) and pd.isna(v)) else None
        except (TypeError, ValueError):
            ival = None
        if ival == 4223:
            audit["flags"].append({
                "row_excel": int(r_excel),
                "flag": "msno_value_4223_suspect_typo",
                "value": ival,
            })

    # ---- Step 10/11: Final Report truncation fix + collapse double-space ----
    for i in df.index:
        v = df.at[i, "Final Report"]
        if isinstance(v, str):
            if v.lstrip().startswith("م تحليل"):
                fixed = "ت" + v.lstrip()
                df.at[i, "Final Report"] = fixed
                audit["changes"].append({
                    "row_excel": int(df.at[i, "row_excel"]),
                    "column": "Final Report",
                    "before": v,
                    "after": fixed,
                    "reason": "final_report_truncation_fix_prepend_taa",
                })
                audit["flags"].append({
                    "row_excel": int(df.at[i, "row_excel"]),
                    "flag": "final_report_truncation",
                    "value": v,
                })

    # ---- Step 12: Final Report code map ----
    code_map = schema["final_report_code_map"]
    final_report_code = []
    for i in df.index:
        v = df.at[i, "Final Report"]
        code = match_final_report_code(v, code_map)
        final_report_code.append(code)
        if v is not None and code == "UNKNOWN":
            audit["flags"].append({
                "row_excel": int(df.at[i, "row_excel"]),
                "flag": "final_report_unmapped",
                "value": v,
            })

    # ---- Step 13: Best-effort numeric parsing of Limits / Results ----
    limit_numeric = []
    result_numeric = []
    for i in df.index:
        lv, lflags = parse_numeric_best_effort(df.at[i, "GSO 1016 Limits"])
        limit_numeric.append(lv)
        for fl in lflags:
            audit["flags"].append({
                "row_excel": int(df.at[i, "row_excel"]),
                "flag": f"limit_numeric_{fl}",
                "value": df.at[i, "GSO 1016 Limits"],
            })
        rv, rflags = parse_numeric_best_effort(df.at[i, "Results"])
        result_numeric.append(rv)
        for fl in rflags:
            audit["flags"].append({
                "row_excel": int(df.at[i, "row_excel"]),
                "flag": f"result_numeric_{fl}",
                "value": df.at[i, "Results"],
            })

    # ---- Step 14: S.No sequence gaps (file-level summary) ----
    s_no_vals = sorted({int(v) for v in df["S.No"].dropna().tolist() if isinstance(v, (int, float)) and not isinstance(v, bool)})
    if s_no_vals:
        full = set(range(min(s_no_vals), max(s_no_vals) + 1))
        gaps = sorted(full - set(s_no_vals))
        audit["sequence_gaps_s_no"] = gaps
        audit["s_no_range"] = [min(s_no_vals), max(s_no_vals)]
    else:
        audit["s_no_range"] = None

    # ---- Pseudomonas spelling flag (residual ambiguity) ----
    for i in df.index:
        v = df.at[i, "Test"]
        if v == "سيدومومناس":
            audit["flags"].append({
                "row_excel": int(df.at[i, "row_excel"]),
                "flag": "pseudomonas_spelling_unverified",
                "value": v,
            })

    # ---- Dilution column nonstandard flag ----
    for i in df.index:
        for col in ["10 1", "10 2", "10 3"]:
            v = df.at[i, col]
            if isinstance(v, str) and v in {"الربع300<", "الخامس 248"}:
                audit["flags"].append({
                    "row_excel": int(df.at[i, "row_excel"]),
                    "flag": "dilution_value_nonstandard",
                    "value": v,
                })

    # ---- GSO code pattern violations (after ISO->null) ----
    gso_pattern = re.compile(r"^[A-Z]-\d+(/\d+)?$")
    for i in df.index:
        v = df.at[i, "GSO code"]
        if isinstance(v, str) and not gso_pattern.match(v):
            audit["flags"].append({
                "row_excel": int(df.at[i, "row_excel"]),
                "flag": "gso_code_pattern_violation",
                "value": v,
            })

    # ---- Validity mapping (FAIL on unknown) ----
    validity_out = []
    for i in df.index:
        v = df.at[i, "Validity"]
        if v is None or (isinstance(v, float) and pd.isna(v)):
            validity_out.append(None)
        elif v in validity_mapping:
            validity_out.append(validity_mapping[v])
        else:
            raise SchemaValidationError(
                f"validity_value_not_in_mapping: row {df.at[i, 'row_excel']} has Validity={v!r}"
            )

    # ---- Test mapping (FAIL on unknown after canonicalisation) ----
    for i in df.index:
        v = df.at[i, "Test"]
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        if v not in test_canonical:
            raise SchemaValidationError(
                f"test_value_not_in_canonical: row {df.at[i, 'row_excel']} has Test={v!r}"
            )

    # ---- Build the canonical output frame ----
    out = pd.DataFrame({
        "source_file": path.name,
        "sheet_date": sheet_date,
        "row_excel": df["row_excel"].astype("int32"),
        "barcode": df["Barcodes"],
        "m_s_no": pd.array(
            [int(v) if isinstance(v, (int, float)) and not isinstance(v, bool) and not pd.isna(v) else pd.NA for v in df["M.S.No"]],
            dtype="Int64",
        ),
        "s_no": pd.array(
            [int(v) if isinstance(v, (int, float)) and not isinstance(v, bool) and not pd.isna(v) else pd.NA for v in df["S.No"]],
            dtype="Int64",
        ),
        "gso_code": df["GSO code"],
        "sample_name": df["Sample Name"],
        "restaurant_name": df["Restaurant name"],
        "test": df["Test"],
        "limit_raw": df["GSO 1016 Limits"],
        "limit_numeric": limit_numeric,
        "result_raw": df["Results"],
        "result_numeric": result_numeric,
        "validity": pd.array(validity_out, dtype="boolean"),
        "final_report_raw": df["Final Report"],
        "final_report_code": final_report_code,
        "conc_10_1": df["10 1"],
        "conc_10_2": df["10 2"],
        "conc_10_3": df["10 3"],
    })

    # data_quality_flags column: pipe-separated per row.
    flags_per_row: dict[int, list[str]] = defaultdict(list)
    for f in audit["flags"]:
        flags_per_row[int(f["row_excel"])].append(f["flag"])
    out["data_quality_flags"] = [
        "|".join(sorted(set(flags_per_row.get(int(r), [])))) or None
        for r in out["row_excel"]
    ]

    # Deterministic sort: sheet_date, m_s_no (NA last), s_no (NA last), test (NA last).
    out = out.sort_values(
        by=["sheet_date", "m_s_no", "s_no", "test"],
        kind="mergesort",
        na_position="last",
        ignore_index=True,
    )

    # Cast string-declared columns to a uniform string dtype so parquet doesn't
    # choke on mixed object columns (some cells were int/float, some str).
    string_cols = [
        "source_file", "barcode", "gso_code", "sample_name", "restaurant_name",
        "test", "limit_raw", "result_raw",
        "final_report_raw", "final_report_code",
        "conc_10_1", "conc_10_2", "conc_10_3",
        "data_quality_flags",
    ]
    for col in string_cols:
        out[col] = pd.array(
            [None if (v is None or (isinstance(v, float) and pd.isna(v))) else str(v) for v in out[col]],
            dtype="string",
        )

    audit["rows_out"] = len(out)
    return out, audit


def clean_file(path: Path) -> pd.DataFrame:
    df, _ = clean_file_with_audit(Path(path))
    return df


# ---------------------------------------------------------------------------
# Diff report
# ---------------------------------------------------------------------------
def write_diff_report(audit: dict, df: pd.DataFrame, out_path: Path) -> None:
    lines: list[str] = []
    lines.append(f"# Cleaning diff report: {Path(audit['source_file']).name}\n")
    lines.append(f"- **Schema:** `{audit['schema_name']}` v{audit['schema_version']}")
    lines.append(f"- **Sheet:** `{audit['sheet_name']}`  ·  **Date:** `{audit['sheet_date']}`")
    lines.append(f"- **Rows in (raw, below header):** {audit['rows_in_raw']}")
    lines.append(f"- **Rows after empty-row drop:** {audit['rows_after_dropna']}")
    lines.append(f"- **Rows out (canonical):** {audit['rows_out']}")
    lines.append(f"- **Cells changed (total):** {len(audit['changes'])}")
    lines.append(f"- **Rows dropped:** {len(audit['drops'])}")
    lines.append(f"- **Flags raised (total):** {len(audit['flags'])}")
    if audit["warnings"]:
        lines.append(f"- **Warnings:** {audit['warnings']}")
    lines.append("")

    if audit["sequence_gaps_s_no"]:
        lines.append("## S.No sequence gaps")
        lines.append("")
        rng = audit.get("s_no_range") or [None, None]
        lines.append(f"S.No range observed: **{rng[0]}..{rng[1]}** ({len(audit['sequence_gaps_s_no'])} missing values)")
        lines.append("")
        lines.append("Missing: " + ", ".join(str(g) for g in audit["sequence_gaps_s_no"]))
        lines.append("")

    lines.append("## Cells changed (per column)")
    lines.append("")
    by_col: dict[str, int] = Counter(c["column"] for c in audit["changes"])
    if by_col:
        lines.append("| column | count |")
        lines.append("|---|---:|")
        for col, n in sorted(by_col.items(), key=lambda kv: (-kv[1], kv[0])):
            lines.append(f"| `{col}` | {n} |")
    else:
        lines.append("_no changes_")
    lines.append("")

    lines.append("## Every value remapped")
    lines.append("")
    lines.append("| row (Excel) | column | before | after | reason |")
    lines.append("|---:|---|---|---|---|")
    for c in audit["changes"]:
        before = repr(c["before"]) if c["before"] is not None else "—"
        after = repr(c["after"]) if c["after"] is not None else "—"
        before = before.replace("|", "\\|")
        after = after.replace("|", "\\|")
        lines.append(f"| {c['row_excel']} | `{c['column']}` | {before} | {after} | {c['reason']} |")
    lines.append("")

    lines.append("## Rows dropped")
    lines.append("")
    if audit["drops"]:
        lines.append("| row (Excel) | reason |")
        lines.append("|---:|---|")
        for d in audit["drops"]:
            lines.append(f"| {d['row_excel']} | {d['reason']} |")
    else:
        lines.append("_no drops_")
    lines.append("")

    lines.append("## Flags raised (per flag)")
    lines.append("")
    flags_by_name: dict[str, list[dict]] = defaultdict(list)
    for f in audit["flags"]:
        flags_by_name[f["flag"]].append(f)
    if flags_by_name:
        lines.append("| flag | count | rows (Excel) | sample value |")
        lines.append("|---|---:|---|---|")
        for flag in sorted(flags_by_name):
            entries = flags_by_name[flag]
            rows = ", ".join(str(e["row_excel"]) for e in entries[:15])
            if len(entries) > 15:
                rows += f", … (+{len(entries) - 15})"
            sample = repr(entries[0]["value"]) if entries[0]["value"] is not None else "—"
            sample = sample.replace("|", "\\|")
            lines.append(f"| `{flag}` | {len(entries)} | {rows} | {sample} |")
    else:
        lines.append("_no flags_")
    lines.append("")

    lines.append("## Output column summary")
    lines.append("")
    lines.append("| column | dtype | non-null | nulls | sample |")
    lines.append("|---|---|---:|---:|---|")
    for col in df.columns:
        s = df[col]
        non_null = int(s.notna().sum())
        nulls = int(s.isna().sum())
        sample_val = next((v for v in s if v is not None and not (isinstance(v, float) and pd.isna(v))), None)
        sample_repr = repr(sample_val) if sample_val is not None else "—"
        sample_repr = sample_repr[:80].replace("|", "\\|")
        lines.append(f"| `{col}` | `{s.dtype}` | {non_null} | {nulls} | {sample_repr} |")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(lines), encoding="utf-8")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main() -> None:
    if len(sys.argv) != 4:
        print("Usage: python clean_file.py <input.xlsx> <output.parquet> <diff.md>", file=sys.stderr)
        sys.exit(2)
    in_path = Path(sys.argv[1]).resolve()
    parquet_path = Path(sys.argv[2]).resolve()
    diff_path = Path(sys.argv[3]).resolve()

    df, audit = clean_file_with_audit(in_path)
    parquet_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(parquet_path, compression="zstd", index=False)
    write_diff_report(audit, df, diff_path)

    print(f"wrote {parquet_path} ({len(df)} rows × {df.shape[1]} cols)")
    print(f"wrote {diff_path}")
    print(f"flags raised: {len(audit['flags'])}; cells changed: {len(audit['changes'])}; drops: {len(audit['drops'])}")


if __name__ == "__main__":
    main()
