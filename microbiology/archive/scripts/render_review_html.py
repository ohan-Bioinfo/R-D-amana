"""Render an Excel sheet as an HTML review page that preserves the original
layout (merged cells, RTL Arabic text, exact values) and colour-codes the
issues found in Phase 1, alongside recommendations and open questions.

Usage: python render_review_html.py <input.xlsx> <output.html>
"""
from __future__ import annotations

import html
import sys
from pathlib import Path

from openpyxl import load_workbook


CANONICAL_TESTS = {
    # legend spellings (canonical) → expanded set of variants seen in data
    "السالمونيلا": {"السالمونيلا"},
    "ايشيريشيا كولاي": {"ايشيريشيا كولاي"},
    "استافيلوكوكس اورياس": {"استافيلوكوكس اورياس"},
    "انتيروباكتريسي": {"انتيروباكتريسي"},
    "العدد الكلي للبكتيريا": {"العدد الكلي للبكتيريا"},
    "كوليفورم": {"كوليفورم", "كولي فورم"},
    "باصلص سيرز": {"باصلص سيرز", "باصلص سيرس"},
    "الخمائر والاعفان": {"الخمائر والاعفان", "الخماير والاعفان", "خمائر", "اعفان"},
    "سيدومومناس": {"سيدومومناس"},  # spelling tbd
}

VARIANT_TO_CANONICAL = {v: k for k, vs in CANONICAL_TESTS.items() for v in vs}


def cell_issues(row_idx: int, col_idx: int, val, header_row: int):
    """Return list of issue tags for a single cell.

    row_idx, col_idx are 1-indexed (Excel-style)."""
    issues = []
    if val is None:
        return issues

    if isinstance(val, str):
        if val and val != val.strip():
            issues.append("trailing-ws")
        if "  " in val.strip():
            issues.append("internal-double-space")

    # Column-specific (only valid for rows after the header)
    if row_idx > header_row:
        # F (col 6) = M.S.No: 4223 looks like a typo for 4523
        if col_idx == 6 and isinstance(val, (int, float)):
            try:
                if int(val) == 4223:
                    issues.append("msno-typo")
            except Exception:
                pass

        # N (col 14) = Test: near-dupes against canonical legend
        if col_idx == 14 and isinstance(val, str):
            stripped = val.strip()
            canon = VARIANT_TO_CANONICAL.get(stripped)
            if canon and canon != stripped:
                issues.append("test-near-dup")

        # R (col 18) = Final Report: truncation marker
        if col_idx == 18 and isinstance(val, str):
            if val.lstrip().startswith("م تحليل"):
                issues.append("final-report-truncation")

        # K, L, M (cols 11/12/13) = "10 1", "10 2", "10 3" — mixed-type cells
        if col_idx in (11, 12, 13) and isinstance(val, str):
            stripped = val.strip()
            if stripped in {"الربع300<", "الخامس 248"}:
                issues.append("dilution-garbled")

        # I (col 9) = Sample Name: trailing-ws already caught above; nothing extra
        # E (col 5) = Barcodes (often empty); J (col 10) = Restaurant name (often empty)

    # The two columns that are 100% empty in this file
    if col_idx in (5, 10) and val is None and row_idx > header_row:
        issues.append("col-empty-in-file")

    return issues


def render(xlsx_path: Path, html_path: Path) -> None:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    sheet_name = ws.title
    n_rows = ws.max_row
    n_cols = ws.max_column

    HEADER_ROW = 6  # 1-indexed (== row index 5 zero-indexed). The auto-detector found this in Phase 1.

    # Merge map: top-left → (rowspan, colspan); skip cells covered by a merge.
    span_info: dict[tuple[int, int], tuple[int, int]] = {}
    skip: set[tuple[int, int]] = set()
    for mr in ws.merged_cells.ranges:
        rowspan = mr.max_row - mr.min_row + 1
        colspan = mr.max_col - mr.min_col + 1
        span_info[(mr.min_row, mr.min_col)] = (rowspan, colspan)
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != (mr.min_row, mr.min_col):
                    skip.add((r, c))

    # Build the rendered table.
    rows_html: list[str] = []
    for r in range(1, n_rows + 1):
        cells: list[str] = [f'<td class="rownum">{r}</td>']
        for c in range(1, n_cols + 1):
            if (r, c) in skip:
                continue
            val = ws.cell(row=r, column=c).value
            spans = span_info.get((r, c))
            attrs: list[str] = []
            if spans:
                rowspan, colspan = spans
                if rowspan > 1:
                    attrs.append(f'rowspan="{rowspan}"')
                if colspan > 1:
                    attrs.append(f'colspan="{colspan}"')

            issues = cell_issues(r, c, val, HEADER_ROW)
            classes: list[str] = []
            for i in issues:
                classes.append(f"issue-{i}")
            if r == HEADER_ROW and c >= 5:
                classes.append("header-cell")
            if r < HEADER_ROW and c <= 4:
                classes.append("legend-cell")
            if val is None:
                classes.append("empty-cell")

            if classes:
                attrs.append(f'class="{" ".join(classes)}"')

            display = "" if val is None else str(val)
            display_html = html.escape(display).replace("\n", "<br>")
            tip = ""
            if issues:
                labels = {
                    "trailing-ws": "trailing whitespace",
                    "internal-double-space": "internal double space",
                    "msno-typo": "likely typo: 4223 should be 4523?",
                    "test-near-dup": f"near-duplicate test name → canonical \"{VARIANT_TO_CANONICAL.get(val.strip() if isinstance(val, str) else '', '')}\"",
                    "final-report-truncation": "first character truncated (\"م تحليل…\" should be \"تم تحليل…\")",
                    "dilution-garbled": "garbled value in dilution column — meaning unclear",
                    "col-empty-in-file": "column is 100% empty in this file",
                }
                tip = " | ".join(labels.get(i, i) for i in issues)
                attrs.append(f'title="{html.escape(tip)}"')

            cells.append(f'<td {" ".join(attrs)}>{display_html}</td>')
        rows_html.append("<tr>" + "".join(cells) + "</tr>")

    table_html = "\n".join(rows_html)

    page = f"""<!DOCTYPE html>
<html lang="en" dir="ltr">
<head>
<meta charset="utf-8">
<title>Review — {html.escape(xlsx_path.name)}</title>
<style>
  :root {{
    --bg: #fafbfc;
    --fg: #1f2328;
    --muted: #57606a;
    --accent: #0969da;
    --border: #d0d7de;
    --header-bg: #ddf4ff;
    --legend-bg: #f6f8fa;

    --c-trailing: #fff8c5;        /* yellow */
    --c-double-space: #fff1b8;    /* yellow-orange */
    --c-msno: #ff8182;            /* red */
    --c-near-dup: #ffd8b5;        /* orange */
    --c-truncation: #ff8182;      /* red */
    --c-dilution: #d8b9ff;        /* purple */
    --c-empty: #eaeef2;           /* grey */
  }}
  html, body {{ background: var(--bg); color: var(--fg); margin: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", system-ui, sans-serif; padding: 24px; }}
  h1, h2, h3 {{ margin: 0 0 8px; }}
  h1 {{ font-size: 22px; }}
  h2 {{ font-size: 17px; margin-top: 28px; padding-bottom: 4px; border-bottom: 1px solid var(--border); }}
  h3 {{ font-size: 14px; margin-top: 20px; color: var(--muted); }}
  .meta {{ color: var(--muted); font-size: 13px; margin-bottom: 16px; }}
  .grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 24px; }}
  .panel {{ background: white; border: 1px solid var(--border); border-radius: 6px; padding: 16px; }}
  .legend-bar {{ display: flex; flex-wrap: wrap; gap: 10px; margin: 8px 0 16px; font-size: 12px; }}
  .legend-bar span {{ padding: 3px 8px; border-radius: 3px; border: 1px solid var(--border); }}
  table.xlsx {{ border-collapse: collapse; font-size: 12px; background: white; direction: rtl; }}
  table.xlsx td {{ border: 1px solid var(--border); padding: 4px 6px; vertical-align: middle; min-width: 32px; max-width: 360px;
                   white-space: pre-wrap; word-break: break-word; }}
  table.xlsx td.rownum {{ background: var(--legend-bg); color: var(--muted); text-align: center; font-family: ui-monospace, monospace;
                          direction: ltr; min-width: 36px; }}
  table.xlsx td.legend-cell {{ background: var(--legend-bg); color: var(--muted); font-style: italic; }}
  table.xlsx td.header-cell {{ background: var(--header-bg); font-weight: 600; }}
  table.xlsx td.empty-cell {{ background: #fdfdfd; }}

  td.issue-trailing-ws {{ background: var(--c-trailing) !important; }}
  td.issue-internal-double-space {{ background: var(--c-double-space) !important; }}
  td.issue-msno-typo {{ background: var(--c-msno) !important; color: white; font-weight: bold; }}
  td.issue-test-near-dup {{ background: var(--c-near-dup) !important; }}
  td.issue-final-report-truncation {{ background: var(--c-truncation) !important; color: white; }}
  td.issue-dilution-garbled {{ background: var(--c-dilution) !important; font-weight: bold; }}
  td.issue-col-empty-in-file {{ background: var(--c-empty) !important; }}

  .scroll {{ overflow: auto; max-height: 78vh; border: 1px solid var(--border); border-radius: 6px; padding: 4px; background: white; }}
  ul {{ margin: 4px 0 12px 18px; padding: 0; }}
  li {{ margin: 6px 0; }}
  code {{ background: #f6f8fa; padding: 1px 5px; border-radius: 3px; font-size: 0.92em; }}
  .q {{ background: #fff8c5; padding: 8px 12px; border-radius: 6px; margin: 8px 0; }}
  .q b {{ color: #9a6700; }}
  .answer {{ display: block; margin-top: 6px; color: var(--muted); font-style: italic; }}
  .pill {{ display: inline-block; padding: 1px 8px; border-radius: 999px; background: #ddf4ff; color: var(--accent); font-size: 11px; font-weight: 600; }}
</style>
</head>
<body>

<h1>Review — <code>{html.escape(xlsx_path.name)}</code></h1>
<div class="meta">
  Sheet <b>{html.escape(sheet_name)}</b> · {n_rows} rows × {n_cols} cols ·
  225 merged ranges · no formulas · no hidden rows/cols/sheets · no colour-coded cells.
  <br>This page preserves the original layout. Cells are coloured by issue. Hover for tooltips.
</div>

<div class="legend-bar">
  <span style="background: var(--c-msno); color: white;">M.S.No typo (4223 → 4523?)</span>
  <span style="background: var(--c-truncation); color: white;">Final Report truncation (م تحليل)</span>
  <span style="background: var(--c-near-dup);">Test name near-duplicate</span>
  <span style="background: var(--c-trailing);">Trailing whitespace</span>
  <span style="background: var(--c-double-space);">Internal double-space</span>
  <span style="background: var(--c-dilution);">Garbled dilution-column value</span>
  <span style="background: var(--c-empty);">Column 100% empty in this file</span>
  <span style="background: var(--header-bg);">Real header row (row 6)</span>
  <span style="background: var(--legend-bg); font-style: italic;">Legend block (rows 1–7, cols A–D)</span>
</div>

<h2>The data — exactly as stored</h2>
<div class="scroll">
<table class="xlsx">
{table_html}
</table>
</div>

<h2>Recommendations &amp; open questions</h2>
<div class="grid">

  <div class="panel">
    <h3>Structural recommendations <span class="pill">need confirm</span></h3>
    <ul>
      <li><b>Drop the legend block.</b> Rows 1–7, cols A–D are decoration (test names + reference limits). They duplicate info that's already in the data block. → ignore on import.</li>
      <li><b>Pin the header row to row 6 (Excel) / index 5 (pandas).</b> Auto-detected confirmed.</li>
      <li><b>Extract the date</b> from the title banner cell <code>E4</code> (<code>'نتائج يوم الاثنين 01/01/2024'</code>) — that's the only place it lives inside the file. Filename gives <code>010124</code> as a backup.</li>
      <li><b>Forward-fill sample-level columns</b> (<code>M.S.No</code>, <code>S.No</code>, <code>GSO code</code>, <code>Sample Name</code>, <code>Restaurant name</code>, <code>Barcodes</code>, <code>Final Report</code>) down across their merged test rows. Without this, 60–70% of those cells read as <code>NaN</code>.</li>
      <li><b>Reshape to long format</b>: one row per <em>(sample, test)</em> with columns <code>date, m_s_no, s_no, gso_code, sample_name, restaurant, test, limit, result, validity, final_report</code>. The current wide layout exists only for printing.</li>
    </ul>

    <h3>Per-column proposals</h3>
    <ul>
      <li><code>M.S.No</code> → <code>int64 nullable</code>; flag <code>4223</code> for review.</li>
      <li><code>S.No</code> → <code>int64 nullable</code>; document the 6 missing numbers (1–37 with gaps).</li>
      <li><code>GSO code</code> → string categorical; allowed pattern <code>^[A-Z]-\\d+(/\\d+)?$</code> + literal <code>ISO</code>.</li>
      <li><code>Sample Name</code> → string, <code>strip()</code> + collapse internal whitespace; preserve Arabic as-is.</li>
      <li><code>Test</code> → categorical against the legend's 7 canonical names + Pseudomonas; map all variants.</li>
      <li><code>GSO 1016 Limits</code> → keep as <code>string</code>. Cannot be coerced to numeric (mixes <code>100</code>, <code>&lt;100/25cm2</code>, <code>غير موجودة</code>).</li>
      <li><code>Results</code> → keep as <code>string</code> for the same reason; add a derived <code>result_numeric</code> column where parseable.</li>
      <li><code>Validity</code> → boolean: <code>صالح</code> = True, <code>غير صالح</code> = False.</li>
      <li><code>Final Report</code> → string; map the 10 unique phrases to a finite set <code>final_report_code</code>; fix the <code>'م تحليل'</code> truncation.</li>
      <li><code>Barcodes</code>, <code>Restaurant name</code> → keep in schema as nullable strings; populated in other files probably.</li>
    </ul>

    <h3>Test-name canonicalization (proposed)</h3>
    <ul>
      <li><code>كولي فورم</code> → <code>كوليفورم</code></li>
      <li><code>باصلص سيرس</code> → <code>باصلص سيرز</code></li>
      <li><code>الخماير والاعفان</code> → <code>الخمائر والاعفان</code></li>
      <li><code>خمائر</code> + <code>اعفان</code> → <code>الخمائر والاعفان</code> <em>(only if you confirm — they appear separately on some rows)</em></li>
      <li><code>سيدومومناس</code> → <em>keep as-is until you confirm canonical Arabic spelling for Pseudomonas</em></li>
    </ul>
  </div>

  <div class="panel">
    <h3>Questions for you <span class="pill">blocking Phase 2</span></h3>

    <div class="q"><b>Q1.</b> What do columns <code>10 1</code>, <code>10 2</code>, <code>10 3</code> (K, L, M) actually mean? They look like dilution labels (10¹, 10², 10³) but carry mixed counts and presence/absence words. What are the rules?
      <span class="answer">your answer:</span></div>

    <div class="q"><b>Q2.</b> Cleaned output shape — <b>long</b> (one row per sample-test) or <b>wide</b> (one row per sample, tests as columns)?
      <span class="answer">your answer:</span></div>

    <div class="q"><b>Q3.</b> Empty <code>Barcodes</code> &amp; <code>Restaurant name</code> — keep in schema (other files have them?) or drop?
      <span class="answer">your answer:</span></div>

    <div class="q"><b>Q4.</b> Test-name merges — confirm:
      <ul>
        <li>كولي فورم → كوليفورم?</li>
        <li>باصلص سيرس → باصلص سيرز?</li>
        <li>الخماير والاعفان → الخمائر والاعفان?</li>
        <li>خمائر / اعفان (when separate rows) — keep separate, or merge into الخمائر والاعفان?</li>
        <li>سيدومومناس — canonical Arabic spelling?</li>
      </ul>
      <span class="answer">your answer:</span></div>

    <div class="q"><b>Q5.</b> <code>M.S.No = 4223</code> sitting between 4522 and 4524 — auto-correct to 4523, or just flag and leave?
      <span class="answer">your answer:</span></div>

    <div class="q"><b>Q6.</b> <code>GSO code = 'ISO'</code> on row 9 — what does it mean? (others follow <code>letter-digit</code>)
      <span class="answer">your answer:</span></div>

    <div class="q"><b>Q7.</b> Confirm forward-fill of sample-level columns through merged test rows is correct?
      <span class="answer">your answer:</span></div>

    <div class="q"><b>Q8.</b> Do you want missing rows in <code>S.No</code> (gaps in 1–37) treated as data-loss flags, or just ignored?
      <span class="answer">your answer:</span></div>

    <div class="q"><b>Q9.</b> Should <code>Final Report</code> be <em>code-mapped</em> (10 phrases → e.g. <code>VALID</code>, <code>INVALID_SALMONELLA</code>, …) so analysis is easy, while keeping the raw text in a side column?
      <span class="answer">your answer:</span></div>

    <h3>What I'm <em>not</em> changing without your green light</h3>
    <ul>
      <li>Anything in the legend block (rows 1–7).</li>
      <li>Any value in <code>Final Report</code> beyond the obvious truncation fix.</li>
      <li>The <code>10 1/10 2/10 3</code> columns until I understand them.</li>
    </ul>
  </div>
</div>

<div class="meta" style="margin-top: 24px;">
  Source: <code>{html.escape(str(xlsx_path))}</code>.<br>
  Generated read-only by <code>scripts/render_review_html.py</code>; the workbook itself is untouched.
</div>

</body>
</html>
"""

    html_path.parent.mkdir(parents=True, exist_ok=True)
    html_path.write_text(page, encoding="utf-8")
    print(f"wrote {html_path} ({len(page)} bytes, {n_rows} rows × {n_cols} cols)")


def main() -> None:
    if len(sys.argv) != 3:
        print("Usage: python render_review_html.py <input.xlsx> <output.html>", file=sys.stderr)
        sys.exit(2)
    render(Path(sys.argv[1]).resolve(), Path(sys.argv[2]).resolve())


if __name__ == "__main__":
    main()
