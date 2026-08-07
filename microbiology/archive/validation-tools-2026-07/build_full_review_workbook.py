"""Comprehensive, fillable classification review — EVERY distinct sample name and
EVERY sample, both years, so Muhannad can spot any mis-classification (not just
the flagged ones). Replays the dashboard's exact GSO derivation per sample.

Sheets:
  • Numbers   — our vs official (validation)
  • Review    — every distinct sample_name -> GSO (grouped BY GSO category so a
                wrong item stands out in its group); yellow corrected column + dropdown
  • All_samples — every sample (id/year/month/name/category/bucket/GSO/source/valid)
  • Valid_GSO_categories — dropdown source

Run:  microbiology/.venv/bin/python microbiology/scripts/build_full_review_workbook.py
Out:  microbiology/reports/full_classification_review_to_fill.xlsx
"""
from __future__ import annotations
from collections import defaultdict
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

from build_classification_table import classify, _val

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "reports" / "full_classification_review_to_fill.xlsx"
OFF = {2024: (9108, 6399), 2025: (11404, 8345)}
MN = {f"{i:02d}": m for i, m in enumerate(
    ["", "Jan", "Feb", "Mar", "Apr", "May", "June", "July", "Aug", "Sep", "Oct", "Nov", "Dec"])}

HDR = PatternFill("solid", fgColor="1C2742")
YEL = PatternFill("solid", fgColor="F59E0B")
YELCELL = PatternFill("solid", fgColor="FFF3C7")
WB = Font(color="FFFFFF", bold=True)
CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build():
    valid_gso = set()
    counts = {}
    # distinct-name review: (name) -> {gso, source, n, years, cats}
    byname = defaultdict(lambda: {"gso": None, "src": None, "n": 0, "years": set(), "cats": set()})
    all_rows = []
    for y in (2024, 2025):
        d = pd.read_parquet(ROOT / "cleaned" / f"data{y}.parquet")
        counts[y] = (len(d), int((d["is_failure"] != True).sum()))  # noqa: E712
        for r in d.to_dict("records"):
            gso, src = classify(r)
            valid_gso.add(gso)
            nm = str(_val(r.get("sample_name")) or "(no name)")
            raw = str(_val(r.get("category_canonical")) or _val(r.get("gso_category_name_en")) or "—")
            bucket = str(_val(r.get("sample_type")) or "—")
            valid = r.get("is_valid")
            g = byname[nm]
            g["gso"] = gso; g["src"] = src; g["n"] += 1; g["years"].add(y)
            if raw != "—":
                g["cats"].add(raw)
            all_rows.append([
                str(_val(r.get("barcode")) or _val(r.get("sample_id")) or ""),
                y, str(r["year_month"])[-2:], nm, raw, bucket, gso, src,
                ("" if pd.isna(valid) else bool(valid)),
            ])

    review = sorted(
        [{"gso_category": g["gso"], "sample_name": nm, "raw_category": " · ".join(sorted(g["cats"])[:2]) or "—",
          "source": g["src"], "samples": g["n"], "years": "/".join(str(x) for x in sorted(g["years"]))}
         for nm, g in byname.items()],
        key=lambda r: (r["gso_category"], -r["samples"]))

    wb = Workbook()

    # Numbers
    ws = wb.active; ws.title = "Numbers"
    ws.append(["year", "metric", "official", "our", "delta"])
    for y in (2024, 2025):
        os_, oc = OFF[y]; rs, rc = counts[y]
        ws.append([y, "Samples", os_, rs, rs - os_])
        ws.append([y, "Compliant", oc, rc, rc - oc])
        ws.append([y, "Non-compliant", os_ - oc, rs - rc, (rs - rc) - (os_ - oc)])
        ws.append([y, "Compliance %", round(100 * oc / os_, 2), round(100 * rc / rs, 2),
                   round(100 * rc / rs - 100 * oc / os_, 2)])

    # Review (by distinct name)
    rv = wb.create_sheet("Review")
    rv.append(["gso_category (current)", "sample_name", "raw_category", "source",
               "samples", "year(s)", "corrected_gso_category", "notes"])
    for r in review:
        rv.append([r["gso_category"], r["sample_name"], r["raw_category"], r["source"],
                   r["samples"], r["years"], "", ""])

    # All samples
    al = wb.create_sheet("All_samples")
    al.append(["sample_id", "year", "month", "sample_name", "raw_category",
               "sample_type_bucket", "gso_category", "source", "valid"])
    for row in all_rows:
        row[2] = MN.get(row[2], row[2])
        al.append(row)

    # Valid GSO list
    vg = wb.create_sheet("Valid_GSO_categories")
    vg.append(["valid_gso_category"])
    for g in sorted(valid_gso):
        vg.append([g])

    # ---- styling ----
    for name in wb.sheetnames:
        ws = wb[name]
        for c in ws[1]:
            c.fill = HDR; c.font = WB; c.alignment = CEN
        ws.freeze_panes = "A2"
    rv.cell(row=1, column=7).fill = YEL
    rv.cell(row=1, column=8).fill = YEL
    # one DV over the whole corrected column (efficient)
    dv = DataValidation(type="list",
                        formula1=f"Valid_GSO_categories!$A$2:$A${len(valid_gso) + 1}",
                        allow_blank=True)
    rv.add_data_validation(dv)
    dv.add(f"G2:G{len(review) + 1}")
    for i in range(2, len(review) + 2):
        rv.cell(row=i, column=7).fill = YELCELL
        rv.cell(row=i, column=8).fill = YELCELL
    for col, w in {"A": 36, "B": 30, "C": 30, "D": 20, "E": 9, "F": 9, "G": 36, "H": 30}.items():
        rv.column_dimensions[col].width = w
    for col, w in {"A": 20, "B": 7, "C": 8, "D": 28, "E": 28, "F": 16, "G": 34, "H": 20, "I": 8}.items():
        al.column_dimensions[col].width = w
    for col, w in {"A": 8, "B": 16, "C": 12, "D": 12, "E": 10}.items():
        wb["Numbers"].column_dimensions[col].width = w
    wb["Valid_GSO_categories"].column_dimensions["A"].width = 40

    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  distinct names to review: {len(review)} | all samples: {len(all_rows)}")


if __name__ == "__main__":
    build()
