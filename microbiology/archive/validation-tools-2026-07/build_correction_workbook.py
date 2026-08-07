"""Build a fillable Excel to (a) validate numbers vs the official annual figures
and (b) let Muhannad correct the samples that are unclassified / classification-
doubtful. Doubtful = classified only by name keyword or fell to Miscellaneous.

Rows are grouped by distinct sample_name so each product is classified once; the
YELLOW 'corrected_gso_category' column has a dropdown of the valid GSO categories.
Send it back and I apply the corrections (name rule / bucket / per-sample).

Run:  microbiology/.venv/bin/python microbiology/scripts/build_correction_workbook.py
Out:  microbiology/reports/classification_corrections_to_fill.xlsx
"""
from __future__ import annotations
from collections import defaultdict
from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

from build_classification_table import classify, _val
from build_dashboard_combined import SAMPLE_TYPE_TO_GSO

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "reports" / "classification_corrections_to_fill.xlsx"
OFF = {2024: (9108, 6399), 2025: (11404, 8345)}     # (samples, compliant)
DOUBTFUL = {"name-keyword", "Miscellaneous-fallback"}


def build():
    valid_gso = set()
    counts = {}
    review_rows = []   # (year, level, key, samples, current_gso, source, example_raw_category)
    for y in (2024, 2025):
        d = pd.read_parquet(ROOT / "cleaned" / f"data{y}.parquet")
        counts[y] = (len(d), int((d["is_failure"] != True).sum()))  # noqa: E712
        by_name = defaultdict(lambda: {"n": 0, "gso": None, "src": None, "cats": set()})
        by_bucket = defaultdict(lambda: {"n": 0, "gso": None, "names": set()})
        for r in d.to_dict("records"):
            gso, src = classify(r)
            valid_gso.add(gso)
            nm = str(_val(r.get("sample_name")) or "(no name)")
            st = str(_val(r.get("sample_type")) or "")
            if src in DOUBTFUL:                       # uncertain name-keyword / Misc
                g = by_name[nm]
                g["n"] += 1; g["gso"] = gso; g["src"] = src
                cat = _val(r.get("category_canonical")) or _val(r.get("gso_category_name_en"))
                if cat:
                    g["cats"].add(str(cat))
            if st:                                    # bucket-classified (2025)
                b = by_bucket[st]
                b["n"] += 1; b["gso"] = gso; b["names"].add(nm)
        # per-year doubtful sample names (both years; 2025 has ~none)
        for nm, g in by_name.items():
            review_rows.append((y, "sample name", nm, g["n"], g["gso"], g["src"],
                                " · ".join(sorted(g["cats"])[:3])))
        # 2025 is bucket-classified — list the bucket→GSO mappings to validate too
        if y == 2025:
            for st, b in by_bucket.items():
                bg = SAMPLE_TYPE_TO_GSO.get(st, "(varies — name-classified, see examples)")
                review_rows.append((y, "sample_type bucket", st, b["n"], bg,
                                    "sample_type-bucket", " · ".join(sorted(b["names"])[:4])))
    review_rows.sort(key=lambda t: (t[0], t[1], -t[3]))

    # ---- write ----
    HDR = PatternFill("solid", fgColor="1C2742")
    YEL = PatternFill("solid", fgColor="F59E0B")
    YELCELL = PatternFill("solid", fgColor="FFF3C7")
    WB = Font(color="FFFFFF", bold=True)
    CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)

    numbers = pd.DataFrame([
        {"year": y, "metric": m, "official": o, "our": r, "delta": r - o}
        for y in (2024, 2025)
        for m, o, r in [
            ("Samples", OFF[y][0], counts[y][0]),
            ("Compliant", OFF[y][1], counts[y][1]),
            ("Non-compliant", OFF[y][0] - OFF[y][1], counts[y][0] - counts[y][1]),
            ("Compliance %", round(100 * OFF[y][1] / OFF[y][0], 2), round(100 * counts[y][1] / counts[y][0], 2)),
        ]])
    review_df = pd.DataFrame(review_rows, columns=[
        "year", "level", "key (name / bucket)", "samples", "current_gso", "source", "example_raw_category"])
    review_df["corrected_gso_category"] = ""
    review_df["notes"] = ""
    valid_df = pd.DataFrame(sorted(valid_gso), columns=["valid_gso_category"])

    # Pure openpyxl write (avoids the pandas ExcelWriter interaction bug).
    from openpyxl import Workbook
    wb = Workbook()

    def _clean(v):
        return None if (isinstance(v, float) and pd.isna(v)) else v

    def write_sheet(title, df, yellow_cols=()):
        ws = wb.active if title == "Numbers" else wb.create_sheet(title)
        ws.title = title
        for j, col in enumerate(df.columns, 1):
            c = ws.cell(row=1, column=j, value=col)
            c.fill = YEL if (j - 1) in yellow_cols else HDR
            c.font = WB; c.alignment = CEN
        for i, (_, row) in enumerate(df.iterrows(), 2):
            for j, val in enumerate(row.tolist(), 1):
                cc = ws.cell(row=i, column=j, value=_clean(val))
                if (j - 1) in yellow_cols:
                    cc.fill = YELCELL
        ws.freeze_panes = "A2"
        return ws

    write_sheet("Numbers", numbers)
    rv = write_sheet("To_review", review_df, yellow_cols=(7, 8))
    write_sheet("Valid_GSO_categories", valid_df)

    gso_list = sorted(valid_gso)
    dv = DataValidation(type="list",
                        formula1=f"Valid_GSO_categories!$A$2:$A${len(gso_list) + 1}",
                        allow_blank=True)
    rv.add_data_validation(dv)
    dv.add(f"H2:H{len(review_df) + 1}")
    for col, w in {"A": 7, "B": 18, "C": 32, "D": 9, "E": 30, "F": 20, "G": 30, "H": 32, "I": 40}.items():
        rv.column_dimensions[col].width = w
    for col, w in {"A": 8, "B": 16, "C": 12, "D": 22, "E": 10}.items():
        wb["Numbers"].column_dimensions[col].width = w
    wb["Valid_GSO_categories"].column_dimensions["A"].width = 40
    wb.save(OUT)

    n2024 = sum(1 for r in review_rows if r[0] == 2024)
    n2025 = sum(1 for r in review_rows if r[0] == 2025)
    print(f"wrote {OUT}")
    print(f"  review rows: {len(review_rows)}  (2024 names: {n2024}, 2025 buckets: {n2025})")


if __name__ == "__main__":
    build()
