"""Per-section chemistry-category validation workbook.

One sheet per (section, year) — 12 tabs total. Each tab lists every distinct
sample_category value in that section with row counts, % of section, validity
breakdown, and sample-name examples for verification.

Output: clean/chemistry_categories_by_section.xlsx
"""
from __future__ import annotations

from pathlib import Path
import pandas as pd
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
CHEM = ROOT / "chemistry"
OUT  = ROOT / "chemistry_categories_by_section.xlsx"

# Section order (canonical)
SECTIONS = [
    "aflatoxins",
    "food_chemistry",
    "heavy_metals",
    "honey",
    "hormones_antibiotics",
    "pesticides",
    "water_analysis",
]
YEARS = [2024, 2025]

# Styles
HEADER_FILL = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEAD_FILL = PatternFill(start_color="EEF2F7", end_color="EEF2F7", fill_type="solid")
SUBHEAD_FONT = Font(bold=True, color="1C2742", size=11)
ALT_FILL = PatternFill(start_color="F7F8FB", end_color="F7F8FB", fill_type="solid")
WATER_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
JUNK_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
INVALID_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
GAP_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
THIN_BORDER = Border(bottom=Side(style="thin", color="D8DEE9"))

JUNK_VALUES = {"NA"} | {f"{n} عينة" for n in ("11","18","20","3")} | {"3 عينات"}


def is_water_text(s):
    if not s: return False
    sl = str(s).lower()
    if any(k in sl for k in ('water','tap','bottle','drink')): return True
    if any(k in str(s) for k in ('مياه','مياة','ماء','شرب','حنفية','حنفيه','فلتر','معبأ','موية','مويه')): return True
    return False


def style_header(ws, row, last_col):
    for ci in range(1, last_col + 1):
        cell = ws.cell(row=row, column=ci)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER


def style_subhead(ws, row, last_col):
    for ci in range(1, last_col + 1):
        cell = ws.cell(row=row, column=ci)
        cell.fill = SUBHEAD_FILL
        cell.font = SUBHEAD_FONT


def fit_columns(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def category_examples(df: pd.DataFrame, category, n: int = 5) -> str:
    """Top-N distinct sample_name values for one category."""
    sub = df[df["sample_category"] == category]
    if "sample_name" not in sub.columns:
        return ""
    names = sub["sample_name"].dropna().astype(str).value_counts().head(n)
    return " | ".join(f"{name} ({c}×)" for name, c in names.items())


wb = Workbook()
wb.remove(wb.active)

# ─── Overview sheet ──────────────────────────────────────────────────────────
ws = wb.create_sheet("Overview")
ws.append(["Chemistry categories — per-section validation"])
ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="3B82F6")
ws.append(["One tab per (section, year). Validate each section's category list independently."])
ws.append([])
ws.append(["Section · Year", "Parquet rows", "Categorized rows", "Distinct categories", "Notes"])
style_header(ws, 4, 5)

for sec in SECTIONS:
    for year in YEARS:
        p = CHEM / f"chem_{sec}_{year}.parquet"
        if not p.exists():
            ws.append([f"{sec} · {year}", "—", "—", "—", "(no parquet for this year)"])
            continue
        df = pd.read_parquet(p)
        if "sample_category" not in df.columns:
            ws.append([f"{sec} · {year}", len(df), 0, 0, "no sample_category column"])
            continue
        nn = df["sample_category"].notna().sum()
        distinct = df["sample_category"].dropna().nunique()
        if nn == 0:
            note = "2024 source has no sample_category column"
        else:
            note = ""
        ws.append([f"{sec} · {year}", len(df), nn, distinct, note])

ws.append([])
ws.append(["Tab color key", ""])
style_subhead(ws, ws.max_row, 2)
ws.append(["Light blue row", "Water-related category (Tap/Unbottled/Bottled/sub-type)"])
ws.append(["Light red row", "Junk value (NA, '11 عينة', etc.)"])
ws.append(["Gray", "2024 sections (no category data in source)"])
fit_columns(ws, [30, 16, 18, 22, 50])
ws.freeze_panes = "A5"


# ─── One sheet per (section, year) ──────────────────────────────────────────
for sec in SECTIONS:
    for year in YEARS:
        p = CHEM / f"chem_{sec}_{year}.parquet"
        if not p.exists():
            continue
        sheet_name = f"{sec[:24]}_{year}"  # max 31 chars
        ws = wb.create_sheet(sheet_name[:31])
        df = pd.read_parquet(p)

        # Title block
        ws.append([f"Section: {sec}  ·  Year: {year}"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="3B82F6")

        if "sample_category" not in df.columns or df["sample_category"].notna().sum() == 0:
            ws.append([])
            ws.append([f"⚠ No sample_category data in this section."])
            if year == 2024:
                ws.append(["The 2024 chemistry source xlsx files DO NOT contain a Sample Category column."])
                ws.append([f"All {len(df):,} rows in this section have NULL category."])
            else:
                ws.append([f"All {len(df):,} rows in this section have NULL category."])
            ws.cell(row=2, column=1).fill = GAP_FILL
            fit_columns(ws, [80])
            continue

        nn = int(df["sample_category"].notna().sum())
        nu = int(df["sample_category"].isna().sum())
        distinct = int(df["sample_category"].dropna().nunique())
        ws.append([f"Total rows: {len(df):,}   ·   Categorized: {nn:,}   ·   Null: {nu:,}   ·   Distinct categories: {distinct}"])
        ws.cell(row=2, column=1).font = Font(color="6B7894")
        ws.append([])

        # Column headers — now includes canonical column from the parquet
        has_canonical = "sample_category_canonical" in df.columns
        headers = ["#", "Sample category (raw)", "Canonical (applied)" if has_canonical else "Canonical",
                   "Rows", "% of section", "Valid", "Invalid", "Unknown",
                   "Fail %", "Distinct sample_names", "Top 5 sample_name examples (with counts)"]
        ws.append(headers)
        header_row = ws.max_row
        style_header(ws, header_row, len(headers))

        # Per-category breakdown
        sub_cat = df[df["sample_category"].notna()]
        counts = sub_cat["sample_category"].value_counts()
        total_categorized = counts.sum()
        for i, (cat, n) in enumerate(counts.items(), start=1):
            cat_rows = sub_cat[sub_cat["sample_category"] == cat]
            # Get the canonical value (should be consistent within each raw cat)
            canonical = ""
            if has_canonical:
                canon_values = cat_rows["sample_category_canonical"].dropna().unique()
                canonical = " | ".join(str(c) for c in canon_values) if len(canon_values) else "(NULL)"
            valid = int((cat_rows.get("is_valid", pd.Series()) == True).sum())
            invalid = int((cat_rows.get("is_valid", pd.Series()) == False).sum())
            unk = len(cat_rows) - valid - invalid
            fail_pct = invalid * 100 / len(cat_rows) if len(cat_rows) else 0
            distinct_names = cat_rows["sample_name"].dropna().nunique() if "sample_name" in cat_rows.columns else 0
            examples = category_examples(df, cat)
            ws.append([i, cat, canonical, n, round(n*100/total_categorized, 1),
                       valid, invalid, unk, round(fail_pct, 1),
                       distinct_names, examples])
            last_row = ws.max_row
            if cat in JUNK_VALUES:
                for ci in range(1, len(headers)+1):
                    ws.cell(row=last_row, column=ci).fill = JUNK_FILL
            elif is_water_text(cat):
                for ci in range(1, len(headers)+1):
                    ws.cell(row=last_row, column=ci).fill = WATER_FILL
            elif i % 2 == 0:
                for ci in range(1, len(headers)+1):
                    ws.cell(row=last_row, column=ci).fill = ALT_FILL
            # Highlight invalid count in red when fail_pct >= 20%
            if fail_pct >= 20:
                ws.cell(row=last_row, column=6).fill = INVALID_FILL
                ws.cell(row=last_row, column=8).fill = INVALID_FILL

        # NULL row at bottom if any
        if nu > 0:
            null_rows = df[df["sample_category"].isna()]
            null_valid = int((null_rows.get("is_valid", pd.Series()) == True).sum())
            null_invalid = int((null_rows.get("is_valid", pd.Series()) == False).sum())
            null_unk = nu - null_valid - null_invalid
            ws.append(["—", "(NULL — no category set)", "", nu, round(nu*100/len(df), 1),
                       null_valid, null_invalid, null_unk,
                       round(null_invalid*100/nu, 1) if nu else 0, "—",
                       category_examples(df.assign(sample_category=df["sample_category"].fillna("__NULL__")),
                                         "__NULL__") if "sample_name" in df.columns else ""])
            for ci in range(1, len(headers)+1):
                ws.cell(row=ws.max_row, column=ci).fill = GAP_FILL

        # Total row
        ws.append([])
        ws.append(["", "TOTAL", "",
                   total_categorized,
                   100.0,
                   int((sub_cat.get("is_valid", pd.Series()) == True).sum()),
                   int((sub_cat.get("is_valid", pd.Series()) == False).sum()),
                   int(sub_cat.get("is_valid", pd.Series()).isna().sum()) if "is_valid" in sub_cat.columns else 0,
                   "", "", ""])
        style_subhead(ws, ws.max_row, len(headers))

        fit_columns(ws, [4, 50, 30, 8, 12, 9, 9, 10, 9, 18, 70])
        ws.freeze_panes = f"A{header_row+1}"
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{ws.max_row}"


# Save
wb.save(OUT)
print(f"wrote {OUT}  ({OUT.stat().st_size:,} bytes, {len(wb.sheetnames)} sheets)")
print("Sheets:")
for s in wb.sheetnames:
    print(f"  - {s}")
