"""Per-year microbiology category validation workbook.

Microbiology doesn't have "sections" like chemistry — it has one dataset per
year (2023/2024/2025). The category source column differs by year:

  - 2023 + 2024: use `gso_category_name_en` (derived from GSO 1016 codes —
    the source xlsx doesn't have a Sample Category column for these years).
  - 2025:        use `category_canonical` (direct from source's "Sample Category"
    column — 353 distinct values, much more granular than the GSO buckets).
  - 2025 sample_type tab: the cleaner's 12 functional buckets, easier to
    validate cross-year alongside 2023/2024.

Output: clean/microbiology_categories_by_year.xlsx
"""
from __future__ import annotations

from pathlib import Path
import pandas as pd
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
MICRO = ROOT / "microbiology"
OUT = ROOT / "microbiology_categories_by_year.xlsx"

YEARS = [2023, 2024, 2025]

# Styles (matching chemistry workbook)
HEADER_FILL = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEAD_FILL = PatternFill(start_color="EEF2F7", end_color="EEF2F7", fill_type="solid")
SUBHEAD_FONT = Font(bold=True, color="1C2742", size=11)
ALT_FILL = PatternFill(start_color="F7F8FB", end_color="F7F8FB", fill_type="solid")
WATER_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
JUNK_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
INVALID_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
GAP_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
PATHOGEN_FILL = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")
THIN_BORDER = Border(bottom=Side(style="thin", color="D8DEE9"))

WATER_FALSE_POSITIVES = {
    "شراب بنكهة و مركزاتها ( flavoured drink & its concentrates )",
    "البطيخ (watermelons)",
    "Pasteurized fruit juice and drink",
    "مشروب الصويا المبستر Pasteurized soya drink",
    "ملحمة سماء القاهرة",
}


def is_water_text(s):
    if not s: return False
    if s in WATER_FALSE_POSITIVES: return False
    sl = str(s).lower()
    if any(k in sl for k in ('water','tap','bottle','drink')): return True
    if any(k in str(s) for k in ('مياه','مياة','ماء','شرب','حنفية','حنفيه','فلتر','معبأ','موية','مويه')): return True
    return False


def style_header(ws, row, last_col):
    for ci in range(1, last_col + 1):
        cell = ws.cell(row=row, column=ci)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER


def style_subhead(ws, row, last_col):
    for ci in range(1, last_col + 1):
        cell = ws.cell(row=row, column=ci)
        cell.fill = SUBHEAD_FILL
        cell.font = SUBHEAD_FONT


def fit_columns(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def category_examples(df: pd.DataFrame, cat_col: str, category, n: int = 5) -> str:
    sub = df[df[cat_col] == category]
    if "sample_name" not in sub.columns:
        return ""
    names = sub["sample_name"].dropna().astype(str).value_counts().head(n)
    return " | ".join(f"{name} ({c}×)" for name, c in names.items())


def top_facilities(df: pd.DataFrame, cat_col: str, category, n: int = 3) -> str:
    sub = df[df[cat_col] == category]
    col = None
    for c in ("facility_chain", "facility_name", "restaurant_name"):
        if c in sub.columns and sub[c].notna().sum() > 0:
            col = c; break
    if col is None:
        return ""
    facs = sub[col].dropna().astype(str).value_counts().head(n)
    return " | ".join(f"{f} ({c}×)" for f, c in facs.items())


def severity_summary(df: pd.DataFrame, cat_col: str, category) -> tuple[int, int, int, int]:
    """Returns (none, indicator_only, pathogen, multi_pathogen) counts."""
    sub = df[df[cat_col] == category]
    if "severity_tier" not in sub.columns:
        return (0, 0, 0, 0)
    counts = sub["severity_tier"].fillna("none").value_counts().to_dict()
    return (
        int(counts.get("none", 0)),
        int(counts.get("indicator_only", 0)),
        int(counts.get("pathogen", 0)),
        int(counts.get("multi_pathogen", 0)),
    )


def pick_category_column(df: pd.DataFrame, year: int) -> tuple[str, str]:
    """Returns (column_name, source_description) for the best category column."""
    if year == 2025 and "category_canonical" in df.columns and df["category_canonical"].notna().sum() > 0:
        return "category_canonical", "category_canonical (from source's Sample Category column)"
    if "gso_category_name_en" in df.columns and df["gso_category_name_en"].notna().sum() > 0:
        return "gso_category_name_en", "gso_category_name_en (derived from GSO 1016 code)"
    if "category_canonical" in df.columns and df["category_canonical"].notna().sum() > 0:
        return "category_canonical", "category_canonical"
    return None, "no category column populated"


wb = Workbook()
wb.remove(wb.active)

# ─── Overview sheet ──────────────────────────────────────────────────────────
ws = wb.create_sheet("Overview")
ws.append(["Microbiology categories — per-year validation"])
ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="3B82F6")
ws.append(["One tab per year (2023/2024/2025) + a 2025 sample_type tab. Each tab lists every distinct category with counts."])
ws.append([])
ws.append(["Year", "Rows", "Category source column", "Distinct values", "Severity?", "Notes"])
style_header(ws, 4, 6)

# Pre-load per-year data
per_year = {}
for year in YEARS:
    p = MICRO / f"data{year}.parquet"
    if not p.exists():
        ws.append([year, "—", "—", "—", "no", "(no parquet)"])
        per_year[year] = None
        continue
    df = pd.read_parquet(p)
    per_year[year] = df
    col, desc = pick_category_column(df, year)
    distinct = df[col].dropna().nunique() if col else 0
    sev = "yes" if "severity_tier" in df.columns else "no"
    note = ""
    if year == 2023:
        note = "Aug–Dec only · no facility data · categories from GSO 1016 only"
    elif year == 2024:
        note = "Full year · no facility data · categories from GSO 1016 only"
    elif year == 2025:
        note = "Full year · has facility chain/branch + sector · direct from source"
    ws.append([year, len(df), desc, distinct, sev, note])

# Bonus row: 2025 sample_type
df25 = per_year.get(2025)
if df25 is not None and "sample_type" in df25.columns and df25["sample_type"].notna().sum() > 0:
    ws.append([
        "2025 (sample_type)", len(df25),
        "sample_type (12 functional buckets — cleaner-derived)",
        df25["sample_type"].dropna().nunique(),
        "yes", "Alternate view: aggregates the 353 category_canonical values into 12 high-level buckets",
    ])

ws.append([])
ws.append(["Tab colour key", ""])
style_subhead(ws, ws.max_row, 2)
ws.append(["Light blue row", "Water-related category (real water)"])
ws.append(["Light red row", "False-positive (water keyword matched but not water) OR ≥20% fail-rate row"])
ws.append(["Light orange row", "Category had pathogen-level failures (severity_tier=pathogen / multi_pathogen)"])
ws.append(["Gray (NULL row)", "Rows without category"])
fit_columns(ws, [16, 14, 60, 18, 14, 60])
ws.freeze_panes = "A5"


def build_year_sheet(wb, sheet_name: str, df: pd.DataFrame, cat_col: str, col_desc: str, year_label: str):
    """Generic builder — works for any year + any category column."""
    ws = wb.create_sheet(sheet_name[:31])
    ws.append([f"{year_label}  ·  Source column: {col_desc}"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="3B82F6")

    if cat_col is None:
        ws.append([])
        ws.append(["⚠ No category column populated for this year."])
        fit_columns(ws, [80])
        return

    nn = int(df[cat_col].notna().sum())
    nu = int(df[cat_col].isna().sum())
    distinct = int(df[cat_col].dropna().nunique())
    ws.append([f"Total rows: {len(df):,}   ·   Categorized: {nn:,}   ·   Null: {nu:,}   ·   Distinct categories: {distinct}"])
    ws.cell(row=2, column=1).font = Font(color="6B7894")
    ws.append([])

    headers = ["#", "Category", "Rows", "% of year",
               "Valid", "Invalid", "Unknown", "Fail %",
               "Severity: none / indic / pathogen / multi",
               "Sample types (top 3)", "Top 5 sample_name examples",
               "Top 3 facility/chain"]
    ws.append(headers)
    header_row = ws.max_row
    style_header(ws, header_row, len(headers))

    sub_cat = df[df[cat_col].notna()]
    counts = sub_cat[cat_col].value_counts()
    total_categorized = counts.sum()
    for i, (cat, n) in enumerate(counts.items(), start=1):
        cat_rows = sub_cat[sub_cat[cat_col] == cat]
        valid = int((cat_rows.get("is_valid", pd.Series()) == True).sum()) if "is_valid" in cat_rows.columns else 0
        invalid = int((cat_rows.get("is_valid", pd.Series()) == False).sum()) if "is_valid" in cat_rows.columns else 0
        unk = len(cat_rows) - valid - invalid
        fail_pct = invalid * 100 / len(cat_rows) if len(cat_rows) else 0
        none_n, ind_n, path_n, multi_n = severity_summary(df, cat_col, cat)
        sev_summary = f"{none_n}/{ind_n}/{path_n}/{multi_n}"
        if "sample_type" in cat_rows.columns:
            st = cat_rows["sample_type"].dropna().astype(str).value_counts().head(3)
            st_str = ", ".join(f"{s} ({c})" for s, c in st.items())
        else:
            st_str = ""
        examples = category_examples(df, cat_col, cat)
        facs = top_facilities(df, cat_col, cat)
        ws.append([i, cat, n, round(n*100/total_categorized, 1),
                   valid, invalid, unk, round(fail_pct, 1),
                   sev_summary, st_str, examples, facs])
        last_row = ws.max_row
        is_fp = cat in WATER_FALSE_POSITIVES
        is_water = is_water_text(cat)
        if is_fp:
            for ci in range(1, len(headers)+1):
                ws.cell(row=last_row, column=ci).fill = JUNK_FILL
        elif is_water:
            for ci in range(1, len(headers)+1):
                ws.cell(row=last_row, column=ci).fill = WATER_FILL
        elif (path_n + multi_n) > 0:
            for ci in range(1, len(headers)+1):
                ws.cell(row=last_row, column=ci).fill = PATHOGEN_FILL
        elif i % 2 == 0:
            for ci in range(1, len(headers)+1):
                ws.cell(row=last_row, column=ci).fill = ALT_FILL
        if fail_pct >= 20:
            ws.cell(row=last_row, column=6).fill = INVALID_FILL
            ws.cell(row=last_row, column=8).fill = INVALID_FILL

    if nu > 0:
        null_rows = df[df[cat_col].isna()]
        null_valid = int((null_rows.get("is_valid", pd.Series()) == True).sum()) if "is_valid" in null_rows.columns else 0
        null_invalid = int((null_rows.get("is_valid", pd.Series()) == False).sum()) if "is_valid" in null_rows.columns else 0
        null_unk = nu - null_valid - null_invalid
        ws.append(["—", f"(NULL — no {cat_col})", nu, round(nu*100/len(df), 1),
                   null_valid, null_invalid, null_unk,
                   round(null_invalid*100/nu, 1) if nu else 0,
                   "—", "—", "—", "—"])
        for ci in range(1, len(headers)+1):
            ws.cell(row=ws.max_row, column=ci).fill = GAP_FILL

    ws.append([])
    total_valid = int((sub_cat.get("is_valid", pd.Series()) == True).sum()) if "is_valid" in sub_cat.columns else 0
    total_invalid = int((sub_cat.get("is_valid", pd.Series()) == False).sum()) if "is_valid" in sub_cat.columns else 0
    total_unk = total_categorized - total_valid - total_invalid
    ws.append(["", "TOTAL", total_categorized, 100.0,
               total_valid, total_invalid, total_unk,
               round(total_invalid*100/total_categorized, 1) if total_categorized else 0,
               "", "", "", ""])
    style_subhead(ws, ws.max_row, len(headers))

    fit_columns(ws, [4, 55, 8, 11, 8, 9, 9, 9, 35, 35, 60, 50])
    ws.freeze_panes = f"A{header_row+1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{ws.max_row}"


# ─── One sheet per year ──────────────────────────────────────────────────────
for year in YEARS:
    df = per_year.get(year)
    if df is None:
        continue
    cat_col, desc = pick_category_column(df, year)
    build_year_sheet(wb, str(year), df, cat_col, desc, f"Year: {year}")

# Bonus: 2025 sample_type tab
df25 = per_year.get(2025)
if df25 is not None and "sample_type" in df25.columns and df25["sample_type"].notna().sum() > 0:
    build_year_sheet(wb, "2025_sample_type", df25, "sample_type",
                     "sample_type (12 functional buckets — cleaner-derived)",
                     "Year: 2025 · sample_type bucket view")


# Save
wb.save(OUT)
print(f"wrote {OUT}  ({OUT.stat().st_size:,} bytes, {len(wb.sheetnames)} sheets)")
print("Sheets:")
for s in wb.sheetnames:
    print(f"  - {s}")
