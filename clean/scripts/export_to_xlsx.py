"""Export every cleaned parquet to an xlsx workbook alongside it.

For each parquet in clean/chemistry/ and clean/microbiology/, writes a
corresponding .xlsx file with:
  - One main 'data' sheet (the table)
  - One 'summary' sheet (column inventory + non-null counts)
  - Bold header row, frozen top row, autofilter, sensible column widths

Run:
    .venv/bin/python clean/scripts/export_to_xlsx.py
Output: clean/{chemistry,microbiology}/*.xlsx
"""
from __future__ import annotations

from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
HEADER_FILL = PatternFill(start_color="EEF2F7", end_color="EEF2F7", fill_type="solid")
HEADER_FONT = Font(bold=True, color="1C2742", size=11)
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)
SUBHEAD_FILL = PatternFill(start_color="F7F8FB", end_color="F7F8FB", fill_type="solid")
THIN_BORDER = Border(bottom=Side(style="thin", color="D8DEE9"))

# Hard cap on column width to keep things sensible.
MAX_WIDTH = 60
MIN_WIDTH = 10


def write_workbook(df: pd.DataFrame, out_path: Path, label: str) -> None:
    """Write df to xlsx with header formatting + a summary sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "data"

    # Convert datetime/Timestamp columns to date-only or ISO string for readability
    df = df.copy()
    for c in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[c]):
            df[c] = df[c].dt.strftime("%Y-%m-%d")
        elif df[c].dtype == "object":
            # Stringify any non-scalar values (lists, dicts, arrays)
            df[c] = df[c].apply(
                lambda v: v if v is None or pd.api.types.is_number(v) or isinstance(v, (str, bool))
                else (",".join(str(x) for x in v) if hasattr(v, "__iter__") and not isinstance(v, str) else str(v))
            )

    # Write header
    for ci, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=ci, value=str(col))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Write data rows
    for ri, (_, row) in enumerate(df.iterrows(), start=2):
        for ci, col in enumerate(df.columns, start=1):
            v = row[col]
            if pd.isna(v):
                continue
            ws.cell(row=ri, column=ci, value=v)

    # Freeze + autofilter
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions

    # Column widths: clamp the longest string per column to MAX_WIDTH
    for ci, col in enumerate(df.columns, start=1):
        max_len = max(len(str(col)),
                      max((len(str(v)) for v in df[col].dropna().head(200)), default=10))
        width = min(MAX_WIDTH, max(MIN_WIDTH, max_len + 2))
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 22

    # === summary sheet ===
    sm = wb.create_sheet("summary")
    sm.append([f"Source: {label}"])
    sm.append([f"Rows: {len(df):,}"])
    sm.append([f"Columns: {df.shape[1]}"])
    sm.append([])
    sm.append(["Column", "Non-null count", "Null count", "Sample values"])
    for col in df.columns:
        nn = int(df[col].notna().sum())
        nu = int(df[col].isna().sum())
        sample = df[col].dropna().head(3).astype(str).tolist()
        sm.append([col, nn, nu, " | ".join(sample)[:100]])

    # Format summary header
    for ri in (1, 2, 3):
        sm.cell(row=ri, column=1).font = Font(bold=True, color="6B7894")
    for ci in range(1, 5):
        cell = sm.cell(row=5, column=ci)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    sm.freeze_panes = "A6"
    sm.column_dimensions["A"].width = 35
    sm.column_dimensions["B"].width = 16
    sm.column_dimensions["C"].width = 14
    sm.column_dimensions["D"].width = 60

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main():
    print("=" * 80)
    print("Exporting parquets → xlsx workbooks")
    print("=" * 80)

    for sub in ("chemistry", "microbiology"):
        src_dir = ROOT / sub
        for pq in sorted(src_dir.glob("*.parquet")):
            df = pd.read_parquet(pq)
            out = src_dir / (pq.stem + ".xlsx")
            print(f"  {pq.name:<40} → {out.name:<40} ({len(df):,} rows × {df.shape[1]} cols)")
            write_workbook(df, out, label=f"{sub}/{pq.name}")

    # Total summary
    total_xlsx = sum(1 for _ in (ROOT / "chemistry").glob("*.xlsx")) + \
                 sum(1 for _ in (ROOT / "microbiology").glob("*.xlsx"))
    print(f"\nWrote {total_xlsx} xlsx files.")


if __name__ == "__main__":
    main()
