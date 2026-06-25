"""Build a validation xlsx covering every sample_category value across both labs.

Output: clean/categories_validation.xlsx with multiple sheets:
  - Overview        — summary counts per source
  - Chem categories — every distinct value, total count, per-section breakdown
  - Micro categories — every distinct value, total count, per-year breakdown
  - Water deep-dive — water samples by category AND by sample_name
  - Canonical map   — suggested normalization (one row per variant → canonical)
"""
from __future__ import annotations

from pathlib import Path
import pandas as pd
from collections import Counter, defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
CHEM = ROOT / "chemistry"
MICRO = ROOT / "microbiology"
OUT = ROOT / "categories_validation.xlsx"

# Styles
HEADER_FILL = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEAD_FILL = PatternFill(start_color="EEF2F7", end_color="EEF2F7", fill_type="solid")
SUBHEAD_FONT = Font(bold=True, color="1C2742", size=11)
ALT_FILL = PatternFill(start_color="F7F8FB", end_color="F7F8FB", fill_type="solid")
WATER_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
JUNK_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
THIN_BORDER = Border(bottom=Side(style="thin", color="D8DEE9"))

WATER_EN = ['water', 'tap', 'bottle', 'drink']
WATER_AR = ['مياه', 'مياة', 'ماء', 'شرب', 'حنفية', 'حنفيه', 'فلتر', 'معبأ', 'موية', 'مويه']
WATER_FALSE_POSITIVES = {
    'شراب بنكهة و مركزاتها ( flavoured drink & its concentrates )',
    'البطيخ (watermelons)',
    'Pasteurized fruit juice and drink',
    'مشروب الصويا المبستر Pasteurized soya drink',
    'ملحمة سماء القاهرة',
}


def is_water_text(s):
    if s is None: return False
    try:
        if pd.isna(s): return False
    except Exception:
        pass
    s = str(s)
    if s in WATER_FALSE_POSITIVES: return False
    sl = s.lower()
    if any(k in sl for k in WATER_EN): return True
    if any(k in s for k in WATER_AR): return True
    return False


def style_header(ws, row, last_col):
    for ci in range(1, last_col + 1):
        cell = ws.cell(row=row, column=ci)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = LEFT
        cell.border = THIN_BORDER


def style_subhead(ws, row, last_col):
    for ci in range(1, last_col + 1):
        cell = ws.cell(row=row, column=ci)
        cell.fill = SUBHEAD_FILL
        cell.font = SUBHEAD_FONT
        cell.border = THIN_BORDER


def fit_columns(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# === Load all data ===
chem_per_section = defaultdict(Counter)  # section -> category -> count
chem_all = Counter()
for p in sorted(CHEM.glob('chem_*.parquet')):
    section = p.stem.replace('chem_', '')
    df = pd.read_parquet(p)
    if 'sample_category' not in df.columns:
        continue
    s = df['sample_category'].dropna().astype(str)
    for v, n in s.value_counts().items():
        chem_per_section[section][v] += n
        chem_all[v] += n

micro_per_year = defaultdict(Counter)
micro_all = Counter()
for year in (2023, 2024, 2025):
    p = MICRO / f'data{year}.parquet'
    if not p.exists():
        continue
    df = pd.read_parquet(p)
    if 'category_canonical' not in df.columns:
        continue
    s = df['category_canonical'].dropna().astype(str)
    for v, n in s.value_counts().items():
        micro_per_year[year][v] += n
        micro_all[v] += n

# === Canonical mapping (Chemistry: 12 canonical) ===
CHEM_CANONICAL = {
    "Fruit and Vegetables الفواكه والخضار": "Fruit and Vegetables",
    "Fruit and Vegetables": "Fruit and Vegetables",
    "Cereal and Legume products الحبوب والبقوليات\"": "Cereal and Legume products",
    "الحبوب والبقوليات\"": "Cereal and Legume products",
    "Cereal and Legume products": "Cereal and Legume products",
    "Spices and Sauces products البهارات والصوصات": "Spices and Sauces products",
    "Ready to Eat Foods الأطعمة الجاهزه للاكل": "Ready to Eat Foods",
    "\"Ready to Eat Foods الأطعمة الجاهزه للاكل\"": "Ready to Eat Foods",
    "Tap water مياه الحنفية": "Tap water",
    "مياة حنفية غسيل الادوات": "Tap water",
    "مياه شرب": "Tap water",
    "مياه متحركة": "Tap water",
    "Drinking water مياه الشرب": "Tap water",
    "مياه فلتر لغسيل الأدوات": "Tap water",
    "المياه الغير المعبأة (Unbottled water)": "Unbottled water",
    "المياه المعبأة (bot water)": "Bottled water",
    "Meat and Poultry products اللحوم والدواجن": "Meat and Poultry products",
    "Fish and Seafood products الأسماك والمأكولات البحرية": "Fish and Seafood products",
    "اعلاف Fodder": "Fodder",
    "Milk and Dairy products الحليب ومنتجات الالبان": "Milk and Dairy products",
    "Beverage المشروبات": "Beverage",
    "Sweets and Chocolate products الحلويات والشكولاته": "Sweets and Chocolate products",
    "Fats and Oils الدهون والزيوت": "Fats and Oils",
    # Junk values
    "NA": "(JUNK — placeholder)",
    "11 عينة": "(JUNK — sample count)",
    "18 عينة": "(JUNK — sample count)",
    "20 عينة": "(JUNK — sample count)",
    "3 عينات": "(JUNK — sample count)",
}


wb = Workbook()
wb.remove(wb.active)

# ─── SHEET 1: Overview ───────────────────────────────────────────────────────
ws = wb.create_sheet("Overview")
ws.append(["Categories validation workbook"])
ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="3B82F6")
ws.append([f"Source: clean/chemistry/ + clean/microbiology/"])
ws.append([f"Generated: 2026-06-11"])
ws.append([])
ws.append(["Section", "Notes"])
style_header(ws, 5, 2)
ws.append(["Chem categories",  f"{sum(chem_all.values()):,} categorized rows · {len(chem_all)} distinct values · 12 canonical · 7 sections"])
ws.append(["Micro categories", f"{sum(micro_all.values()):,} categorized rows · {len(micro_all)} distinct values · across 2023/2024/2025"])
ws.append(["Water deep-dive", "Every water-related category + sample-name keyword findings (Tap / Unbottled / Bottled + sub-types)"])
ws.append(["Canonical map",   "Suggested variant → canonical mapping for chemistry (use to normalize)"])
ws.append([])
ws.append(["Chemistry coverage", ""])
style_subhead(ws, 11, 2)
ws.append(["Total rows in chemistry", "15,786"])
ws.append(["Rows with category filled", "9,427 (60%)"])
ws.append(["Rows with NULL category", "6,359 (40% — all 2024 rows; column absent in 2024 source forms)"])
ws.append(["Distinct categories", "28"])
ws.append(["Canonical categories (after normalization)", "12"])
ws.append([])
ws.append(["Microbiology coverage", ""])
style_subhead(ws, 18, 2)
ws.append(["Total rows in microbiology (wide)", "22,596"])
ws.append(["Rows with category_canonical filled", f"{sum(micro_all.values()):,}"])
ws.append(["Distinct category_canonical values", f"{len(micro_all)}"])
fit_columns(ws, [42, 80])
ws.freeze_panes = "A5"

# ─── SHEET 2: Chemistry categories ──────────────────────────────────────────
ws = wb.create_sheet("Chem categories")
headers = ["#", "Sample category (chemistry)", "Total rows", "Canonical group", "Notes",
           "aflatoxins_2025", "food_chemistry_2025", "heavy_metals_2025",
           "honey_2025", "hormones_antibiotics_2025", "pesticides_2025", "water_analysis_2025"]
ws.append(headers)
style_header(ws, 1, len(headers))

chem_sections = ['aflatoxins_2025','food_chemistry_2025','heavy_metals_2025',
                 'honey_2025','hormones_antibiotics_2025','pesticides_2025','water_analysis_2025']

# Sort by total count desc
sorted_cats = chem_all.most_common()
for i, (cat, n) in enumerate(sorted_cats, start=1):
    canonical = CHEM_CANONICAL.get(cat, "(needs review)")
    notes = ""
    if canonical.startswith("(JUNK"):
        notes = "Junk value — exclude or remap"
    elif canonical == "(needs review)":
        notes = "Not in canonical map yet"
    elif cat != canonical and not canonical.startswith("(JUNK"):
        # Show what kind of variant
        if "\"" in cat: notes = "Variant: stray quote"
        elif " " in cat[:3] and not any(c.isalpha() for c in cat[:3]): notes = "Variant"
        elif cat.startswith("مياة") or cat.startswith("مياه") and canonical == "Tap water":
            notes = "Water sub-type"
    row = [i, cat, n, canonical, notes]
    for sec in chem_sections:
        row.append(chem_per_section[sec].get(cat, 0))
    ws.append(row)
    # Style row
    last_row = ws.max_row
    if canonical.startswith("(JUNK"):
        for ci in range(1, len(headers)+1):
            ws.cell(row=last_row, column=ci).fill = JUNK_FILL
    elif is_water_text(cat) and canonical in ("Tap water", "Unbottled water", "Bottled water"):
        for ci in range(1, len(headers)+1):
            ws.cell(row=last_row, column=ci).fill = WATER_FILL
    elif i % 2 == 0:
        for ci in range(1, len(headers)+1):
            ws.cell(row=last_row, column=ci).fill = ALT_FILL

# Total row
ws.append([])
totals = ["", "TOTAL CATEGORIZED ROWS", sum(chem_all.values()), "", ""]
for sec in chem_sections:
    totals.append(sum(chem_per_section[sec].values()))
ws.append(totals)
last_row = ws.max_row
style_subhead(ws, last_row, len(headers))

fit_columns(ws, [4, 55, 12, 30, 26, 16, 18, 16, 12, 22, 16, 18])
ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"

# ─── SHEET 3: Microbiology categories ───────────────────────────────────────
ws = wb.create_sheet("Micro categories")
headers = ["#", "category_canonical (microbiology)", "Total rows", "2023", "2024", "2025",
           "Water-related?", "False-positive?"]
ws.append(headers)
style_header(ws, 1, len(headers))

# Sort by total
sorted_m = micro_all.most_common()
for i, (cat, n) in enumerate(sorted_m, start=1):
    water = "yes" if is_water_text(cat) else ("yes (FP)" if cat in WATER_FALSE_POSITIVES else "")
    fp = "FALSE POSITIVE" if cat in WATER_FALSE_POSITIVES else ""
    row = [i, cat, n, micro_per_year[2023].get(cat, 0), micro_per_year[2024].get(cat, 0),
           micro_per_year[2025].get(cat, 0), water, fp]
    ws.append(row)
    last_row = ws.max_row
    if cat in WATER_FALSE_POSITIVES:
        for ci in range(1, len(headers)+1):
            ws.cell(row=last_row, column=ci).fill = JUNK_FILL
    elif is_water_text(cat):
        for ci in range(1, len(headers)+1):
            ws.cell(row=last_row, column=ci).fill = WATER_FILL
    elif i % 2 == 0:
        for ci in range(1, len(headers)+1):
            ws.cell(row=last_row, column=ci).fill = ALT_FILL

ws.append([])
ws.append(["", "TOTAL CATEGORIZED ROWS", sum(micro_all.values()),
           sum(micro_per_year[2023].values()),
           sum(micro_per_year[2024].values()),
           sum(micro_per_year[2025].values()), "", ""])
style_subhead(ws, ws.max_row, len(headers))

fit_columns(ws, [4, 60, 12, 10, 10, 10, 16, 18])
ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

# ─── SHEET 4: Water deep-dive ───────────────────────────────────────────────
ws = wb.create_sheet("Water deep-dive")
ws.append(["WATER CATEGORIES — across chemistry and microbiology"])
ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="3B82F6")
ws.append([])

# Chemistry water by category
ws.append(["Chemistry — water categories"])
style_subhead(ws, ws.max_row, 1)
ws.append(["Category", "Rows", "Notes"])
style_header(ws, ws.max_row, 3)
chem_water_cats = [(c, n) for c, n in chem_all.items() if is_water_text(c)]
chem_water_cats.sort(key=lambda x: -x[1])
for cat, n in chem_water_cats:
    canonical = CHEM_CANONICAL.get(cat, "Tap water")
    note = "★ canonical" if cat == "Tap water مياه الحنفية" else "sub-type variant"
    ws.append([cat, n, note])
ws.append(["TOTAL", sum(n for _, n in chem_water_cats), ""])
style_subhead(ws, ws.max_row, 3)
ws.append([])

# Microbio water by category
ws.append(["Microbiology — water categories"])
style_subhead(ws, ws.max_row, 1)
ws.append(["Category", "Rows", "Notes"])
style_header(ws, ws.max_row, 3)
mw = [(c, n, c in WATER_FALSE_POSITIVES) for c, n in micro_all.items()
      if is_water_text(c) or c in WATER_FALSE_POSITIVES]
mw.sort(key=lambda x: -x[1])
for cat, n, fp in mw:
    note = "FALSE POSITIVE — not water" if fp else ("★ canonical" if cat in ("المياه الغير معبأة (Unbottled water)", "المياه المعبأة (Bottled water)") else "sub-type")
    ws.append([cat, n, note])
ws.append(["TOTAL real water", sum(n for c, n, fp in mw if not fp), ""])
style_subhead(ws, ws.max_row, 3)
ws.append([])

# Sample-name water analysis
ws.append(["Sample-name keyword analysis (water-use sub-types)"])
style_subhead(ws, ws.max_row, 1)
ws.append(["Note: official sample_category collapses all of these to 'Tap water' / 'Unbottled water'."])
ws.append(["The functional sub-type lives in the sample_name field — use that for finer granularity."])
ws.append([])

# Get top water sample names from chemistry + microbio
chem_names = Counter()
for p in CHEM.glob('chem_*.parquet'):
    df = pd.read_parquet(p)
    if 'sample_name' not in df.columns: continue
    sn = df['sample_name'].dropna().astype(str)
    for v in sn[sn.apply(is_water_text)]:
        chem_names[v] += 1
micro_names = Counter()
for year in (2023, 2024, 2025):
    p = MICRO / f'data{year}.parquet'
    if not p.exists(): continue
    df = pd.read_parquet(p)
    if 'sample_name' not in df.columns: continue
    sn = df['sample_name'].dropna().astype(str)
    for v in sn[sn.apply(is_water_text)]:
        micro_names[v] += 1

# Top 30 by chemistry
ws.append(["Top 30 water sample names — Chemistry"])
style_subhead(ws, ws.max_row, 1)
ws.append(["Sample name", "Chemistry rows", "Microbio rows (same name)", "Functional sub-type"])
style_header(ws, ws.max_row, 4)
def classify(name):
    s = str(name)
    if 'غسيل الادوات' in s or 'غسيل الأدوات' in s or 'غسيل ادوات' in s: return 'utensil-wash'
    if 'للطبخ' in s: return 'cooking'
    if 'للعجانة' in s or 'للعجانه' in s: return 'dough-mixer'
    if 'لغسيل الارز' in s or 'لغسيل الأرز' in s: return 'rice-wash'
    if 'لغسيل الاسماك' in s or 'لغسيل الأسماك' in s: return 'fish-wash'
    if 'حوض' in s: return 'basin / tank'
    if 'حنفية' in s or 'حنفيه' in s: return 'tap water'
    if 'فلتر' in s: return 'filtered water'
    if 'شرب' in s: return 'drinking water'
    if 'معبأ' in s: return 'bottled / packaged'
    return 'other'
for name, n in chem_names.most_common(30):
    ws.append([name, n, micro_names.get(name, 0), classify(name)])

ws.append([])
ws.append(["Top 30 water sample names — Microbiology"])
style_subhead(ws, ws.max_row, 1)
ws.append(["Sample name", "Microbio rows", "Chemistry rows (same name)", "Functional sub-type"])
style_header(ws, ws.max_row, 4)
for name, n in micro_names.most_common(30):
    ws.append([name, n, chem_names.get(name, 0), classify(name)])

fit_columns(ws, [55, 14, 22, 22])
ws.freeze_panes = "A2"

# ─── SHEET 5: Canonical mapping ─────────────────────────────────────────────
ws = wb.create_sheet("Canonical map")
ws.append(["Suggested chemistry canonical mapping (variant → canonical)"])
ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="3B82F6")
ws.append(["Edit the 'Canonical' column to refine the mapping; the cleaner can apply it after."])
ws.append([])
ws.append(["#", "Variant value (as in raw)", "Rows", "Canonical (suggested)", "Type"])
style_header(ws, 4, 5)
sorted_map = sorted(chem_all.most_common(), key=lambda x: (CHEM_CANONICAL.get(x[0], "z"), -x[1]))
for i, (variant, n) in enumerate(sorted_map, start=1):
    canonical = CHEM_CANONICAL.get(variant, "(needs review)")
    typ = ""
    if canonical.startswith("(JUNK"): typ = "JUNK"
    elif variant == canonical: typ = "canonical"
    elif canonical == "(needs review)": typ = "unmapped"
    else: typ = "variant"
    ws.append([i, variant, n, canonical, typ])
    last_row = ws.max_row
    if typ == "JUNK":
        for ci in range(1, 6): ws.cell(row=last_row, column=ci).fill = JUNK_FILL
    elif typ == "variant":
        for ci in range(1, 6): ws.cell(row=last_row, column=ci).fill = WATER_FILL if "water" in canonical.lower() else ALT_FILL
fit_columns(ws, [4, 60, 12, 30, 14])
ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A4:E{ws.max_row}"

# Save
wb.save(OUT)
print(f"wrote {OUT}  ({OUT.stat().st_size:,} bytes, {len(wb.sheetnames)} sheets)")
print(f"Sheets: {wb.sheetnames}")
